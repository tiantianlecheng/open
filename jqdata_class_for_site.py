# coding: UTF-8

from jqdata import *

import os

import datetime

import time

import shelve

import csv

import re

import math

import pprint

import openpyxl

from openpyxl.utils import get_column_letter,column_index_from_string

import os.path

import pandas as pd 

from scipy import stats

import numpy as np

from tabulate import tabulate

import matplotlib.pyplot as plt

import matplotlib.dates as mdates

import statistics

from dateutil.relativedelta import relativedelta



#让dataframe全部展示所有的数据。

pd.set_option('display.max_rows', None, 'display.max_columns', None)

#让dataframe显示小数点后三位。
pd.set_option('display.float_format', lambda x: '%.3f' % x)


#this script only works on joinquant.com, copy and paste it on the website and find all the strong sec you want. ALL index!!!!


all_index_list=['399295.SZ','399296.SZ',
                    '399983.SZ','399804.SZ','399814.SZ','399807.SZ','399986.SZ','399673.SZ','000819.SH','399966.SZ',
                     '399987.SZ','399997.SZ','399932.SZ','399998.SZ','399395.SZ','399440.SZ','399809.SZ','399393.SZ',
                     '399967.SZ','000016.SH', '399805.SZ','399971.SZ','399975.SZ','399992.SZ','000979.SH','000805.SH',
                     '399412.SZ','399812.SZ','399396.SZ','399330.SZ','399006.SZ','000998.SH','000978.SH','000827.SH',
                      '000992.SH','000991.SH','000903.SH','399300.SZ','000993.SH','000987.SH','399437.SZ','000922.SH',
                      '399811.SZ','399995.SZ','000989.SH','399905.SZ','399976.SZ','000688.SH',]


def check_code_exist(target_list_added,code_added):  #用来检验某个代码的信息是否已经生成表中表。

	code_existing_list=[]

	for code_list in target_list_added:
		code_existing=code_list[0]
		code_existing_list.append(code_existing)

	if code_added not in code_existing_list:

		return True


#如下两个函数，建议date加上当天的日期。否则退市的也跟来捣乱。
def get_all_index(end_day):

	all_index_list=get_all_securities(types=['index'], date=end_day).index.tolist()

	return all_index_list

def get_all_sec(end_day):

	all_sec_list=get_all_securities(types=['stock'], date=end_day).index.tolist()

	return all_sec_list

def get_all_etf(end_day):

	all_etf_list=get_all_securities(types=['etf'],date=end_day).index.tolist()

	return all_etf_list


class CreateCsv():

	def __init__(self,csv_name):

		self.csv_name=csv_name
		

	def csv_add_mode(self,content_line):
		csvFile=open('{}.csv'.format(self.csv_name),'a',newline='',encoding='gbk')
		csvWriter=csv.writer(csvFile)

		write_content_list=[]
		write_content_list.append(content_line)
		csvWriter.writerow(write_content_list)


		blank_line_list=[]
		blank_line='                               '
		blank_line_list.append(blank_line)
		csvWriter.writerow(blank_line_list)

		csvFile.close()

		return '{}.csv'.format(self.csv_name)


	def csv_write_mode(self,content_line):

		csvFile=open('{}.csv'.format(self.csv_name),'w',newline='',encoding='gbk')
		csvWriter=csv.writer(csvFile)

		write_content_list=[]
		write_content_list.append(content_line)
		csvWriter.writerow(write_content_list)


		blank_line_list=[]
		blank_line='                               '
		blank_line_list.append(blank_line)
		csvWriter.writerow(blank_line_list)

		csvFile.close()


		return '{}.csv'.format(self.csv_name)

	def create_csv_from_list(self,target_list):  #从一个列表的数据创建csv文件，这样的文件后续可以轻易做成execl,这样也好排序。pandas 读取excel数据排序。

		csvFile=open('{}.csv'.format(self.csv_name),'a',newline='',encoding='gbk')
		csvWriter=csv.writer(csvFile)

		csvWriter.writerow(target_list)

		csvFile.close()

		return '{}.csv'.format(self.csv_name)



class CreateTxt():

	def __init__(self,txt_name):
		self.txt_name=txt_name


	def txt_add_mode(self,txt_line):

		txtFile=open('{}.txt'.format(self.txt_name),'a',encoding='gbk')
		
		txtFile.write(txt_line)
		txtFile.write('\n')
		txtFile.write('')
		txtFile.write('\n')


		txtFile.close()

		return '{}.txt'.format(self.txt_name)


	def txt_write_mode(self,txt_line):

		txtFile=open('{}.txt'.format(self.txt_name),'w',encoding='gbk')
		
		txtFile.write(txt_line)
		txtFile.write('\n')
		txtFile.write('')
		txtFile.write('\n')

		txtFile.close()

		return '{}.txt'.format(self.txt_name)


class DateJq():
	'''create date object for jointquant API use'''
	def __init__(self):
		self.end_date = datetime.datetime.now().strftime('%Y-%m-%d')
		self.onedaydelta = datetime.timedelta(days=1)
		self.threedaydelta = datetime.timedelta(days=3)
		self.now_time = datetime.datetime.now()
		self.past_end_date=(self.now_time - self.onedaydelta).strftime('%Y-%m-%d')
		self.previous_three_day=(self.now_time - self.threedaydelta).strftime('%Y-%m-%d')
		self.next_trade_day=(self.now_time + self.onedaydelta).strftime('%Y-%m-%d')

	def get_previousday(self,end_day):#end_day 的格式必须是2022-06-29这样的
		#在指定的日期上获取前一天的日期。

		target_day=datetime.datetime.strptime(end_day,'%Y-%m-%d')

		previous_day=(target_day - self.onedaydelta).strftime('%Y-%m-%d')

		return previous_day  #in the format like 2022-06-29


	def get_next_day(self,end_day):#end_day 的格式必须是2022-06-29这样的
		#在指定的日期上，获取下一天的日期。

		target_day=datetime.datetime.strptime(end_day,'%Y-%m-%d')

		next_day=(target_day + self.onedaydelta).strftime('%Y-%m-%d')

		return next_day  #in the format like 2022-06-29


	def complete_date(self,date_input):

		if len(date_input)==8:

			string_year=date_input[:4]
			string_month=date_input[4:6]
			string_date=date_input[6:]

			complete_date='-'.join([string_year,string_month,string_date])

		elif len(date_input)==10 and date_input[4]=='-' and date_input[7]=='-':

			complete_date=date_input

		else:

			print(" 请输入日期格式如 '2022-06-19' 或 '20220619',重试!")


		return complete_date




class Jqdata():

	'''create a class from jointquant API for easy use!!!'''

	def __init__(self,code):
		self.code = code
		self.jqcode = normalize_code(code)  #把代码变成是jq需要的代码。
		self.frequency = 'daily'            #获取的频率是以天为单位的！
		self.volume_count_days = 25         #默认是25
		self.money_count_days = 25          #默认是25
		self.close_count_days = 6           #获取收盘点位数据的间隔天数。默认6天，一般计算过去5天etf或者指数的表现，找出强势的指数,5天的表现，必须用6天的数据。
		#self.point_count_days=21            #获取盘中最高最低，收盘开盘数据的间隔天数，一般为21天，计算过去20天的数据。这个主要是用来计算ATR的数据的。(不好控制，2022-07-01取消使用)
		self.risk_value=0.0010               #让每只股票每天对总资产的影响在0.15%,20只股票，对资产的影响将是0.15*20=3%个点。调低为10个点，发现自己最近买股票太多，止损太大。
		self.high_low_point_count_days = 2  #获取最高点，最低点数据的间隔天数，一般观察过去2，3天的数据。默认是2.
		self.max_high_close_count_days = 20 #获取20个交易日内，两个月最高点的价位和最高的收盘价。以便于观察一个趋势可能出现的回撤程度。
		self.zhenfu_count_days=30           #获取30个交易日内，一个股票/指数/基金的每日振幅的数据。
		self.unit = '1d'                    #设置获取数据的单位，通常是看日线，那么就是'1d'
		self.extra_days=150                  #这个是专为获取100天平均线时候，启用的，就是算100天的数据，会多给出50天的数据，这样会有比较多的数据来算平均线，数据结果会更准确。
		self.score_days=90                  #这个就是为动量排名定的，90天，根据书上来定的90天，一般不更改！
		self.gapdays=91                     #计算缺口的天数，设定为91天，因为需要比较数据大小计算等等，所以第一天的数据最后要删除，所以就多留一天，应该是90天。
		self.gapsize=10                     #这个是设定的股票行情的缺口的大小为，10%，大于10%的缺口的股票，一律不考虑，但可以随时更改！
		self.gapsize2=20                    #这个是设定的创业板和科创板股票行情的缺口的大小为，20%，大于20%的缺口的股票，一律不考虑，但可以随时更改！
		self.ATR_days=21                    #为了计算20天内ATR的平均值，第一天的数据最后会去掉，所以为了计算考虑，取21天。
		self.peak_price_stocks={}
		self.trend_days=3                   #为了找出过去几天股票高点低点不断走高的股票。定为3天，可以随时调整。
		self.lowchange_sec_list_week=[]     #设置横盘的股票列表为空,周线观察
		self.lowchange_sec_list_month=[]    #设置横盘的股票列表为空,月线观察
		self.lowchange_rate=1.3              #这个值是横盘的最高点与最低点的比值，目前比例是定位30%，即高点/低点小于等于1.3，可以随时改。
		self.all_index_list=all_index_list  #全市场所有指数代码
		self.week_per_month=15              #横盘看周线，横盘看至少3个月，所以要一个这个数据，每个月按5周计算。所以是15
		self.low_change_month=3             #横盘看月线，横盘至少看3个月，这是基础，也可以改，写个方法做下设置。


	def transform_code(self,code_transformed):   #把代码变成正常的代码，不带任何的字母，就是6位数而已。
		codeRegex=re.compile(r'(\d{6}).(\w{4})')
		code_part=codeRegex.search(code_transformed)[1]
		character_part=codeRegex.search(code_transformed)[2]

		return code_part

	def security_name(self):
		security_name = get_security_info(self.jqcode).display_name
		return security_name


	def get_security_public_date(self):  #获得一个股票的上市日期。

		public_date=get_security_info(self.jqcode).start_date.strftime('%Y-%m-%d')

		return public_date

	def judge_recent_public_stock(self,present_day): #判断一个股票是否上市时间少于一年。

		public_date=self.get_security_public_date()

		original_public_date=datetime.datetime.strptime(public_date,'%Y-%m-%d')

		#date_standard=original_public_date.replace(year=original_public_date.year + 1)  

		date_standard=original_public_date + relativedelta(years=1)  #如果两年就是+2，现在是不要上市低于1年的。

		comapring_date=date_standard.strftime('%Y-%m-%d')

		if comapring_date > present_day:

			return True  #说明这个股票上市低于一年。

		else:

			return False

	def judge_index_point(self, endday):  #如果得不到一个指数当天或某天的收盘点位，说明这个指数的数据观察不到，那么就要自己手动观察。

	    df = get_price(self.jqcode, frequency=self.frequency, end_date=endday, count=1, fields=['close'])

	    return not pd.isna(df.close).values[0]


	def security_industry(self,end_day):

		industry_name=get_industry(self.jqcode,date=end_day)[self.jqcode]['sw_l3']['industry_name']  #使用的是申万三级的行业名称。通常分为申万，聚宽或者证监会分类，见下面实例参考。#数据获取出错，sw_l3改为jq_l1

		return industry_name


	def set_volume_count_days(self,countdays=25):  #设定要获取成交量数据的时间段的天数。
		self.volume_count_days = countdays
		return self.volume_count_days

	def get_volume(self,end_day):  #获取给定天数（时间段）的日成交量和启用函数当天的成交量。
	#从聚宽API获取成交量的数据。可以设置天数。默认是25天。返回现在的成交量，过去25天最大的量，最小的量，和25天平均成交量。
		#end_date=DateJq().end_date
		past_end_date=get_previous_days(end_day,2)[0]  #从指定日期的前一天开始计算。

		df=get_price(self.jqcode,frequency=self.frequency,end_date=past_end_date,count=self.volume_count_days,fields=['volume']) #不包括程序运行当天
		max_volume=df.values.max()         # 过去几天成交量最大的一天量
		min_volume=df.values.min()         # 过去几天成交量最小的一天量
		average_volume=df.values.mean()    # 过去几天的平均成交量

		return max_volume,min_volume,average_volume


	def get_present_volume(self,end_day):

		volume_df=get_price(self.jqcode,frequency=self.frequency,skip_paused=True,fields=['volume'],count=2,end_date=end_day) 

		present_volume=volume_df.iloc[-1,0]

		return present_volume



	def store_volume_data(self,end_day): #存储的是指定日前之前的数据。不包括指定日期当天。
		'''store the obtained volume data for future use'''
		volume_data='myvolume'+str(self.volume_count_days)
		shelveFile=shelve.open(volume_data)
		volume_data_list=[]
		#present_volume,max_volume,min_volume,average_volume=self.get_volume()
		max_volume,min_volume,average_volume=self.get_volume(end_day)


		#volume_data_list.append(present_volume)
		volume_data_list.append(max_volume)
		volume_data_list.append(min_volume)
		volume_data_list.append(average_volume)
		key_shelveFile=self.code  #这个跟后面的volume_compare里面的代码要一致
		shelveFile[key_shelveFile]=volume_data_list
		shelveFile.close()

	def restore_volume_data(self,end_day):  #指定是哪天，用end_day来设定。
		'''restore the stored volume data for use now'''
		volume_data='myvolume'+str(self.volume_count_days)
		shelveFile=shelve.open(volume_data)
		key_shelveFile=self.code #跟前面的store_volume_data相一致。否则出错。

		#end_date=DateJq().end_date

		present_df=get_price(self.jqcode,frequency=self.frequency,fields=['volume'],count=1,end_date=end_day)  #就是获取当天的数据。
		present_volume=present_df.values[0][0]

		max_volume=shelveFile[key_shelveFile][0]
		min_volume=shelveFile[key_shelveFile][1]
		average_volume=shelveFile[key_shelveFile][2]


		max_compare=round(present_volume/max_volume,2)    #与过去多少天的最大成交量相比,是一个比值。
		min_compare=round(present_volume/min_volume,2)    #与过去多少天的最小成交量相比
		average_compare=round(present_volume/average_volume,2)   #与过去多少天的平均成交量相比

		return max_compare,min_compare,average_compare   #返回当前成交量与最大，最小成交量的比较。

	def get_volume_change(self,end_day):

		volume_df=get_price(self.jqcode,frequency=self.frequency,skip_paused=True,fields=['volume'],count=2,end_date=end_day) 

		#print(volume_df)

		present_volume=volume_df.iloc[-1,0]

		previous_volume=volume_df.iloc[0,0]

		#if previous_volume!=0:

		volume_change=round((present_volume/previous_volume),2)

		#else:

			#volume_change=-101  #说明这个数据不存在。

		return volume_change


	def get_volume_ranking(self,end_day):  #得到包括当天某只股票的成交量排名，从1到5，值越大越好，最大为5. volume_change越小越好，配合起来使用。改成观察10天内的排名，count=10


		volume_df=get_price(self.jqcode,frequency=self.frequency,skip_paused=True,fields=['volume'],count=10,end_date=end_day)   # 测试一个股票在过去5天中的排名。	

		volume_df['date']=volume_df.index  #将index的日期值变成一列。

		dataframe_ranking(volume_df,'volume')

		volume_df.reset_index(inplace=True)

		volume_df['ranking']=volume_df.index+1    #这样就按照从高到低，排名，并给予排名。

		volume_df=volume_df.loc[:,['date','volume','ranking']]

		#print(volume_df)

		try:

			target_position=np.where(volume_df['date']==end_day)[0][0]

			ranking_number_for_volume=volume_df.loc[target_position,'ranking']

		except:

			#print(code)

			ranking_number_for_volume=0  #说明没有排名。因为当天没有成交量，所以没有成交量的数据，也就无法排名。比如停牌造成的。

			pass

		ranking_number_for_volume=int(ranking_number_for_volume)

		return ranking_number_for_volume


	def volume_compare(self):   #暂时不使用，有store,restore两个方法可以解决目前的需求。跟restore_volume_data是一样的内容，完全一致。
		'''compare the present volume data with past trading days'''
		#self.store_volume_data()  #generate the volume data and store it.
		volume_data='myvolume'+str(self.volume_count_days)
		shelveFile=shelve.open(volume_data)
		key_shelveFile=self.code #跟前面的store_volume_data相一致。否则出错。
		present_volume=shelveFile[key_shelveFile][0]
		max_volume=shelveFile[key_shelveFile][1]
		min_volume=shelveFile[key_shelveFile][2]
		average_volume=shelveFile[key_shelveFile][3]

		max_compare=round(present_volume/max_volume,2)    #与过去多少天的最大成交量相比,是一个比值。
		min_compare=round(present_volume/min_volume,2)    #与过去多少天的最小成交量相比
		average_compare=round(present_volume/average_volume,2)   #与过去多少天的平均成交量相比

		return max_compare,min_compare,average_compare   #返回当前成交量与最大，最小成交量的比较。

	def store_compare_volume(self):
		"""store the volume data and then compare the volume data within this same method"""
		self.store_volume_data()
		max_compare,min_compare,average_compare=self.volume_compare()

		return max_compare,min_compare,average_compare   #如果没有上面的一行，和这一行，那么返回的就是None,这是基本概念！！！


	#below is the method to get the real money transaction for each day or appointed days!


	def set_money_count_days(self,countdays=25):  #设定要获取成交量数据的时间段的天数。
		self.money_count_days = countdays
		return self.money_count_days

	def get_money(self,end_day):  #获取给定天数（时间段）的日成交量和启用函数当天的成交量。
	#从聚宽API获取成交量的数据。可以设置天数。默认是25天。返回现在的成交量，过去25天最大的量，最小的量，和25天平均成交量。
		past_end_date=get_previous_days(end_day,2)[0]

		df=get_price(self.jqcode,frequency=self.frequency,end_date=past_end_date,count=self.money_count_days,fields=['money']) #不包括程序运行当天
		max_money=df.values.max()         # 过去几天成交量最大的一天量
		min_money=df.values.min()         # 过去几天成交量最小的一天量
		average_money=df.values.mean()    # 过去几天的平均成交量

		return max_money,min_money,average_money


	def get_present_money(self,end_day): #只计算当天的成交金额数据。

		present_date=end_day

		df=get_price(self.jqcode,frequency=self.frequency,start_date=present_date,end_date=present_date,fields=['money']) #只包括程序运行当天

		present_money=round(df.values[0][0]/100000000,4) #以亿为单位。

		return present_money


	def get_average_money(self,end_day): #计算20天以内的平均日成交额，以便于淘汰成交量太低的目标。

		df=get_price(self.jqcode,frequency=self.frequency,count=20,end_date=end_day,fields=['money']) #只包括程序运行当天

		average_money=round(df['money'].mean()/100000000,3) #以亿为单位的成交额

		return average_money


	def get_stock_closechange_in_days(self,count_days,end_day):#计算过去多少天那的收盘价。

		stock_df=get_price(self.jqcode,count=count_days, end_date=end_day, frequency='1d', fields='close')

		stock_closechange_df=stock_df.pct_change()

		stock_closechange_df=stock_closechange_df*100  #获得涨跌幅的百分比。

		return stock_closechange_df


	def get_money_change(self,end_day):

		money_df=get_price(self.jqcode,frequency=self.frequency,skip_paused=True,fields=['money'],count=2,end_date=end_day) 

		present_money=money_df.iloc[-1,0]

		previous_money=money_df.iloc[0,0]

		money_change=round((present_money/previous_money),2)

		return money_change



	def store_money_data(self,end_day):
		'''store the obtained money data for future use'''
		money_data='mymoney'+str(self.money_count_days)
		shelveFile=shelve.open(money_data)
		money_data_list=[]
		#present_money,max_money,min_money,average_money=self.get_money()
		max_money,min_money,average_money=self.get_money(end_day)


		#money_data_list.append(present_money)
		money_data_list.append(max_money)
		money_data_list.append(min_money)
		money_data_list.append(average_money)
		key_shelveFile=self.code  #这个跟后面的money_compare里面的代码要一致
		shelveFile[key_shelveFile]=money_data_list
		shelveFile.close()

	def restore_money_data(self,end_day):  #指定当天的数据
		'''restore the stored money data for use now'''
		money_data='mymoney'+str(self.money_count_days)
		shelveFile=shelve.open(money_data)
		key_shelveFile=self.code #跟前面的store_money_data相一致。否则出错。

		#end_date=DateJq().end_date
		end_date=end_day

		present_df=get_price(self.jqcode,frequency=self.frequency,fields=['money'],count=1,end_date=end_date)
		present_money=present_df.values[0][0]

		max_money=shelveFile[key_shelveFile][0]
		min_money=shelveFile[key_shelveFile][1]
		average_money=shelveFile[key_shelveFile][2]


		max_compare=round(present_money/max_money,2)    #与过去多少天的最大成交量相比,是一个比值。
		min_compare=round(present_money/min_money,2)    #与过去多少天的最小成交量相比
		average_compare=round(present_money/average_money,2)   #与过去多少天的平均成交量相比

		return max_compare,min_compare,average_compare   #返回当前成交量与最大，最小成交量的比较。

	def create_restore_money_data(self,end_day): #生成这个数据，然后提炼这个数据。  目前看用处不大。
		self.store_money_data(end_day)
		max_compare,min_compare,average_compare=self.restore_money_data(end_day)



	def money_compare(self):   #暂时不使用，有store,restore两个方法可以解决目前的需求。跟restore_money_data是一样的内容，完全一致。
		'''compare the present money data with past trading days'''
		#self.store_money_data()  #generate the money data and store it.
		money_data='mymoney'+str(self.money_count_days)
		shelveFile=shelve.open(money_data)
		key_shelveFile=self.code #跟前面的store_money_data相一致。否则出错。
		present_money=shelveFile[key_shelveFile][0]
		max_money=shelveFile[key_shelveFile][1]
		min_money=shelveFile[key_shelveFile][2]
		average_money=shelveFile[key_shelveFile][3]

		max_compare=round(present_money/max_money,2)    #与过去多少天的最大成交量相比,是一个比值。
		min_compare=round(present_money/min_money,2)    #与过去多少天的最小成交量相比
		average_compare=round(present_money/average_money,2)   #与过去多少天的平均成交量相比

		return max_compare,min_compare,average_compare   #返回当前成交量与最大，最小成交量的比较。

	def store_compare_money(self):
		"""store the money data and then compare the money data within this same method"""
		self.store_money_data()
		max_compare,min_compare,average_compare=self.money_compare()

		return max_compare,min_compare,average_compare   #如果没有上面的一行，和这一行，那么返回的就是None,这是基本概念！！！



	def get_sec_market_size(self,end_day):

		q = query (valuation).filter(valuation.code == self.jqcode)

		df = get_fundamentals(q, end_day)


		market_size=df['market_cap'][0]

		liutong_size=df['circulating_market_cap'][0]

		#exchange_rate=df['turnover_ratio'][0]


		return market_size,liutong_size



	def set_close_count_days(self,countdays=6):  #获取收盘数据的时间段的天数，如果要获取过去5天的数据，那么时间段要设置为6天。因为获取包含运行程序当天的数据。
		
		self.close_count_days = countdays

		return self.close_count_days



	def get_close_change(self,end_day,count_days): #计算多少天内的价格变化百分比。
		'''get the close change precent in the appointed days!!!'''
		#end_date=DateJq().end_date
		#onedaydelta=DateJq().onedaydelta
		#now_time=DateJq().now_time
		endday=end_day
		countdays=count_days + 1    #这里要加1，因为比如说考虑过去5天的价格变化，就要从6天前的收盘价算起。

		df=get_price(self.jqcode,frequency=self.frequency,end_date=endday,count=countdays,skip_paused=True,fields=['close']) #包括运行程序的当天。

		close_start=df.values[0][0]  #开始日收盘价
		close_end=df.values[-1][0]   #最后一天收盘价

		#if close_start!=0:

		change_percent=round((close_end - close_start)/close_start,4)*100   #增长的百分比

		#else:

			#change_percent=-101   #说明这个数据不存在。

		return change_percent



    

	def get_close_change_within_days(self,start_day,end_day):

		'''get the close change percent in the days range,for example from 2022-03-01 to 2022-03-29'''

		df=get_price(self.jqcode,start_date=start_day,end_date=end_day,fields=['close'])   #start from one day to the end day. 


		close_start=df.values[0][0]   #开始日收盘价

		close_end=df.values[-1][0]   #最后一天收盘价

		change_percent=round((close_end - close_start)/close_start,4)*100   #增长的百分比

		return change_percent

	
	def get_close_change_during_days(self,end_day,count_days):  #功能和get_close_change 是一样的，这个需要重新测试下。

		'''get the close change percent during the time'''

		countdays=(count_days + 1)

		df=get_price(self.jqcode,count=countdays,end_date=end_day,fields=['close'])   #start from one day to the end day. 


		close_start=df.values[0][0]   #开始日收盘价

		close_end=df.values[-1][0]   #最后一天收盘价

		change_percent=round((close_end - close_start)/close_start,4)*100   #增长的百分比

		return change_percent


	def get_close_within_days(self,start_day,end_day):  #返回pandas数据结果。

		"get the close data within a period days" 

		df=get_price(self.jqcode,frequency=self.frequency,start_date=start_day,end_date=end_day,fields=['close'])  

		return df


	def get_close_for_days(self,end_day,count_days): #以后自己输入count_days,免得出错！   #返回pandas数据结果。

		"get the stock close within a period"

		df=get_price(self.jqcode,frequency=self.frequency,end_date=end_day,count=count_days,fields=['close'])

		return df



     #计算动量策略里面的R2,选择出动量好的股票，R2介于0-1之间，越大越好。
	def momentum_score(self,end_day):

		code=self.jqcode

		#print(code)

		score_days=self.score_days

		df=self.get_close_for_days(end_day,score_days)

		ts=df.close #这是一个pd.series 类型的数据，如果不是这样，下面就会出错。

		"""
		Input:  Price time series.
		Output: Annualized exponential regression slope, 
		multiplied by the R2
		"""
		# Make a list of consecutive numbers
		x = np.arange(len(ts)) 
		# Get logs
		log_ts = np.log(ts) 
		# Calculate regression values
		slope, intercept, r_value, p_value, std_err = stats.linregress(x, log_ts)
		#print("R2="+str(r_value**2))
		# Annualize percent
		annualized_slope = (np.power(np.exp(slope), 252) - 1) * 100
		#print("annualized_rate="+str(annualized_slope))
		#Adjust for fitness
		score = round(annualized_slope * (r_value ** 2),2)

		scoredays=score_days

		return score,scoredays  #这个分数越高越好。


	def get_close(self,end_day,count_days):
		'''get the close change precent in the appointed days!!!'''
		#end_date=DateJq().end_date
		#onedaydelta=DateJq().onedaydelta
		#now_time=DateJq().now_time
		endday=end_day

		countdays=count_days

		df=get_price(self.jqcode,frequency=self.frequency,end_date=endday,count=countdays,fields=['close']) #包括运行程序的当天。

		close_yesterday=df.values[-2][0]  #当前日前一天的收盘价。
		close_present=df.values[-1][0]   #最后一天收盘价

		return close_yesterday,close_present
    
	def get_present_close(self,end_day):#获取的是指定的某天当天的收盘价。

		'''get the present close point of index or sec'''
		#end_date=DateJq().end_date

		present_close=get_price(self.jqcode,frequency=self.frequency,end_date=end_day,count=1,fields=['close']).values[-1][0]

		return present_close

	def get_present_low(self,end_day): #get the lowest price of the day

		present_low=get_price(self.jqcode,frequency=self.frequency,end_date=end_day,count=1,fields=['low']).values[-1][0]

		return present_low

	def get_present_high(self,end_day): #get the highest price of the day

		present_high=get_price(self.jqcode,frequency=self.frequency,end_date=end_day,count=1,fields=['high']).values[-1][0]

		return present_high

	def get_yesterday_close(self,end_day):

		'''get the close point of index or sec for yesterday'''

		target_day=datetime.datetime.strptime(end_day,'%Y-%m-%d')

		yesterday=(target_day - DateJq().onedaydelta).strftime('%Y-%m-%d')	

		yesterday_close=get_price(self.jqcode,frequency=self.frequency,end_date=yesterday,count=1,fields=['close']).values[-1][0]

		return yesterday_close


	def store_yesterday_close_data(self,end_date=DateJq().end_date):

		'''store the sec or index yesterday close data in shelve format'''

		yesterday_close_data='yesterday_close_shelve'
		shelveFile=shelve.open(yesterday_close_data)
		yesterday_close=self.get_yesterday_close(end_date)
		key_shelveFile=self.code
		#print(key_shelveFile)
		shelveFile[key_shelveFile]=yesterday_close
		shelveFile.close()

	def restore_yesterday_close_data(self):
		'''restore the sec or index yesterday close data from shelve format'''
		yesterday_close_data='yesterday_close_shelve'
		shelveFile=shelve.open(yesterday_close_data)
		key_shelveFile=self.code

		yesterday_close=shelveFile[key_shelveFile]

		shelveFile.close()

		return yesterday_close

	def get_daily_close_change(self,end_date=DateJq().end_date):

		'''only get the present day change, not the same as get_present_close_change() as it can set the days between past days and today'''
		yesterday_close=self.get_yesterday_close(end_date)  #get yesterday close data

		present_close=self.get_present_close(end_date)

		change_percent=round((present_close - yesterday_close)/yesterday_close,4)*100 #增长的百分比

		return change_percent


	def get_present_close_change(self,end_day):

		'''get the present day close change for sec or index'''


		close_change=self.get_close_change(end_day,1)  #这里只能写一天。因为里面要加1的。

		return close_change

	def get_present_high_low_change(self,end_day): #获取市场当日跌倒最低点的跌幅和涨到最高点的涨幅。

		present_low=self.get_present_low(end_day)  #获取当天最低点的价格。

		present_high=self.get_present_high(end_day) #获取当天最高点的价格。

		yesterday_date=get_yesterday_date(end_day)

		yesterday_close=self.get_present_close(yesterday_date)

		present_max_high_change=round((present_high - yesterday_close)/yesterday_close,4)*100

		present_max_low_change=round((present_low - yesterday_close)/yesterday_close,4)*100

		return present_max_high_change,present_max_low_change


	def get_close_change_YTD(self,startdate):
		'''fix the start date to the present day, the close change'''

		stock_public_date=self.get_security_public_date() #get stock public date

		now_time=DateJq().now_time
		#now_time='2023-12-30'

		if stock_public_date > startdate:  #说明这个股票是在年内上市的，那么计算该股票年内上市的时间，就是上市日开始算起。

			startdate = stock_public_date

		df=get_price(self.jqcode,start_date=startdate,end_date=now_time,fields=['close'])

		close_start=df.values[0][0]  #开始日收盘价
		close_end=df.values[-1][0]   #最后一天收盘价

		#if close_start!=0:

		increase_percent=round((close_end - close_start)/close_start,4)*100   #增长的百分比

		#else:

			#increase_percent=-101   #代表没有开始日的数据。

		return increase_percent
    

	def get_point(self):#只获取当天的四个点的数据。

		df=get_price(self.jqcode,frequency=self.frequency,fields=['close','open','high','low',],count=1,end_date=DateJq().end_date)

		present_point=df.values[0][0]

		present_open=df.values[0][1]

		present_high_point=df.values[0][2]

		present_low_point=df.values[0][3]

		#print(present_point,present_open,present_high_point,present_low_point)

		return present_point,present_open,present_high_point,present_low_point


	def get_stock_point(self,end_day):#只获取制定日的四个点的数据。

		df=get_price(self.jqcode,frequency=self.frequency,fields=['close','open','high','low',],count=1,end_date=end_day)

		present_point=df.values[0][0]

		present_open=df.values[0][1]

		present_high_point=df.values[0][2]

		present_low_point=df.values[0][3]

		#print(present_point,present_open,present_high_point,present_low_point)

		return present_point,present_open,present_high_point,present_low_point	


	def is_doji(self,end_day):#判断一个股票当天是否是十字星K线图。

		close_price,open_price,high_price,low_price=self.get_stock_point(end_day)
		body_size = abs(close_price - open_price)
		wick_size = high_price - max(open_price, close_price, key=lambda x: abs(x - (open_price + close_price) / 2))

		return body_size < wick_size * 0.5  # 根据需要调整阈值


	def get_point_within_days(self,end_day,count_days): #获取指定日期包括当天的数据，数据包括开盘收盘，盘中最高最低的数据,

		df=get_price(self.jqcode,frequency=self.frequency,fields=['close','open','high','low'],count=count_days,end_date=end_day)

		return df        #return the dataframe of these four data.  #count_days是多少天，就是多少天的数据。


	def get_ma(self,ma_days,end_day):

		#calculate the moving average value like ma20,ma100,normally will calculate more than the target days.

		#extra_days,like 50 or more.normally will add more than the target.

		#end_day通常以当天为最后结束的那天。

		extra_days=self.extra_days

		code=self.jqcode

		countdays=ma_days + extra_days

		df=self.get_close_for_days(end_day,countdays)

		ma_rolling=df.rolling(window=ma_days).mean() #the target days avarage value

		ma=ma_rolling.values[-1][0]

		return ma


	def judge_stock_up_trend(self,end_day):

		n1, n2, n3 = 5,10,20

		ma1,ma2,ma3 = self.get_ma(n1, end_day),self.get_ma(n2, end_day),self.get_ma(n3, end_day)

		presentclose = self.get_present_close(end_day)

		return presentclose > ma1 > ma2 > ma3   #判断是否多头排列， 是的话就是True


	def judge_above_ma(self,ma_days,end_day):

		code=self.jqcode

		ma=self.get_ma(ma_days,end_day)

		present_close=self.get_present_close(end_day)

		return present_close > ma    #这样写更简洁高效，如果大于就返回True,否则就是False 


	def find_above_ma_sec(self,ma_days,end_day): #为了找出昨天没过100天线，今天过100天线的股票，就是在临界点找出这些股票。

		code=self.jqcode

		present_day=end_day

		previous_day=get_previous_days(end_day,2)[0]

		present_ma=self.get_ma(ma_days,present_day)  #这是指定当天的数据。这里应该列出昨天的100天线的值，和今天的100天线的值。

		previous_ma=self.get_ma(ma_days,previous_day)  #这是指定日前一天的数据。

		df=self.get_close_for_days(end_day,2) #只获取某天和前一天，所以这里应该是2,

		yesterday_close=df.values[0][0]

		present_close=df.values[1][0]

		#print(yesterday_close,present_close)

		#present_close=self.get_present_close(end_date=end_day)  #这个错误，这个永远是从当天出发，算当天的。而我这里是想指定算那天的就算那天的。

		#yesterday_close=self.get_yesterday_close(end_date=end_day)  #这个错误，这个永远是从当天出发，算昨天的。

		if present_close > present_ma and yesterday_close < previous_ma:  #当天的收盘价，大于当天的100天线，昨天的收盘价，小于昨天的100天线，才是合理的。

			#return code

			return True

		else:

			#return None
			return False



	def judge_index_above_ma(self,ma_days,end_day):

		bool_value=self.judge_above_ma(ma_days,end_day)

		warning_word=self.security_name()+' '+self.jqcode+' 目前点位低于 '+str(ma_days)+' 天均线,请警惕风险！'

		if bool_value ==False:

			print(warning_word)

		return bool_value  #需要用来判断不好的股票有多少。


	def index_sec_close_change_ranking_within_days(self,count_days,end_day):

		code=self.jqcode

		countdays=(count_days + 1)

		securities_list=self.get_index_securities()

		all_securities_info=[]

		for security in securities_list:

			security_info_list=[]

			sec_name=get_security_info(security).display_name

			df=get_price(security,fields=['close'],count=countdays,end_date=end_day)

			close_start=df.values[0][0]   #开始日收盘价

			close_end=df.values[-1][0]   #最后一天收盘价

			change_percent=round((close_end - close_start)/close_start,4)*100   #增长的百分比

			security_info_list=[security,sec_name,change_percent,count_days,end_day]

			all_securities_info.append(security_info_list)

		df=pd.DataFrame(all_securities_info)

		df.columns=['Code','Name','Close_change','Days','End_date']

		df.sort_values(by=['Close_change'],ascending=False,inplace=True)

		return df

	def judge_stock_gapsize(self,end_day): #找出缺口口大于9.9%或者小于-9.9%的股票，弃之不用。

		code=self.jqcode

		#self.set_point_count_days(gapdays)

		gapdays=self.gapdays

		df=self.get_point_within_days(end_day,gapdays)

		if code.startswith('30') or code.startswith('688'):

			gap_size=self.gapsize2   #创业板，科创板 设置为20%

		else:

			gap_size=self.gapsize      #其他的设置为10% 


		df['yesterday_high']=df['high'].shift(1)  #昨天的最高价，单独提出来。

		df['yesterday_low'] =df['low'].shift(1)   #昨天的最低价，也单独提出来。

		df=df.drop(df.index[0]) #去掉第一行，因为第一行的 yesterday_high 是没有值的，没有前一天的数据，所以不在比较之列。

		conditionlist=[
		              (df['low'] > df['yesterday_high']),
		              (df['high'] < df['yesterday_low']),
		              ]

		high_gap=round((df['low']/df['yesterday_high']-1)*100,2)

		low_gap=round((df['high']/df['yesterday_low']-1)*100,2)

		choicelist=[high_gap,low_gap]

		df['gap_size']=np.select(conditionlist,choicelist,default=0)   #this block of code is very new to me, to compare the different column, and then make a new column after action.

		bool_value_list=((df['gap_size'] > gap_size) | (df['gap_size'] < (-gap_size))).tolist() #A股缺口大于9.99的，就放弃。这个语法以前没用过，用 | 表示或在dataframe中。

		gap_days=gapdays

		if True in bool_value_list:

			return True, gap_days,gap_size #这样的股票就要放弃。缺口过大。

		else:

			return False,gap_days,gap_size 


	def judge_stock_data(self,end_day):#比较过去三天收盘价，最高价，最低价，是否在逐步提高的数据，最后通过yes no 来表达。

		code=self.jqcode

		df=self.get_point_within_days(end_day,self.trend_days) #设置三天。可以随时调整。 

		df['yesterday_high']=df['high'].shift(1)  #昨天的最高价，单独提出来。

		df['yesterday_low'] =df['low'].shift(1)   #昨天的最低价，也单独提出来。

		df['yesterday_close']=df['close'].shift(1) 

		df=df.drop(df.index[0]) #去掉第一行，因为第一行的 yesterday_high 是没有值的，没有前一天的数据，所以不在比较之列。

		bool_value_list_close_up=(df['close'] > df['yesterday_close']).tolist()    #就是单纯的比较收盘价。

		#print(bool_value_list)

		if False in bool_value_list_close_up:   #只要有不满足条件的，那么就是不是连续上涨的股票。就返回 false

			close_up='no'

		else:

			close_up='yes'

		bool_value_list_close_down=(df['close'] < df['yesterday_close']).tolist()    #就是单纯的比较收盘价。

		#print(bool_value_list)

		if False in bool_value_list_close_down:   #只要有不满足条件的，那么就是不是连续下跌的股票。就返回 false，确定不是连续下跌的股票。

			close_down='no'

		else:

			close_down='yes'



		bool_value_list_high_up=(df['high'] > df['yesterday_high']).tolist()    #就是单纯的比较最高价。

		#print(bool_value_list)

		if False in bool_value_list_high_up:   #只要有不满足条件的，那么就是不是连续上涨的股票。就返回 false

			high_up='no'

		else:

			high_up='yes'

		bool_value_list_high_down=(df['high'] < df['yesterday_high']).tolist()    #就是单纯的比较最高价。

		#print(bool_value_list)

		if False in bool_value_list_high_down:   #只要有不满足条件的，那么就是不是连续下跌的股票。就返回 false，确定不是连续下跌的股票。

			high_down='no'

		else:

			high_down='yes'


		bool_value_list_low_up=(df['low'] > df['yesterday_low']).tolist()    #就是单纯的比较最低价。

		#print(bool_value_list)

		if False in bool_value_list_low_up:   #只要有不满足条件的，那么就是不是连续上涨的股票。就返回 false

			low_up='no'

		else:

			low_up='yes'

		bool_value_list_low_down=(df['low'] < df['yesterday_low']).tolist()    #就是单纯的比较最低价。

		#print(bool_value_list)

		if False in bool_value_list_low_down:   #只要有不满足条件的，那么就是不是连续下跌的股票。就返回 false，确定不是连续下跌的股票。

			low_down='no'

		else:

			low_down='yes'


		return close_up,close_down,high_up,high_down,low_up,low_down


	def judge_stock_potential(self,end_day): #就是发现如果包括当天在内，满足过去三天高点低点都上升的股票，就说明有潜力。

		code=self.jqcode

		close_up,close_down,high_up,high_down,low_up,low_down=self.judge_stock_data(end_day)

		if high_up == 'yes' and low_up == 'yes':  #获得过去三天高点低点连续上升的判断。没有要求收盘价也必须连涨。

			return True

		else:

			return False



	def judge_stock_trend(self,end_day): #找出过去三天低点和高点每天都上涨的股票，实际上就是比较包括今天在内三天的数据。

		code=self.jqcode

		#self.set_point_count_days(gapdays)

		#gapdays=self.gapdays

		df=self.get_point_within_days(end_day,self.trend_days)  #设置三天。可以随时调整。 


		df['yesterday_high']=df['high'].shift(1)  #昨天的最高价，单独提出来。

		df['yesterday_low'] =df['low'].shift(1)   #昨天的最低价，也单独提出来。

		df=df.drop(df.index[0]) #去掉第一行，因为第一行的 yesterday_high 是没有值的，没有前一天的数据，所以不在比较之列。

		#conditionlist=[
		              #(df['low'] > df['yesterday_high']),
		              #(df['high'] < df['yesterday_low']),
		              #]

		#high_gap=round((df['low']/df['yesterday_high']-1)*100,2)

		#low_gap=round((df['high']/df['yesterday_low']-1)*100,2)

		#choicelist=[high_gap,low_gap]

		#df['gap_size']=np.select(conditionlist,choicelist,default=0)   #this block of code is very new to me, to compare the different column, and then make a new column after action.

		bool_value_list=((df['high'] > df['yesterday_high']) & (df['low'] > df['yesterday_low'])).tolist() #高点低点走高就。这个语法以前没用过，用 & 表示或在dataframe中。

		#print(bool_value_list)

		if False in bool_value_list:  #bool_value_list 是一个datafram所有行，通过筛选得到一个个true,false的值的列表

			return False   #这样的股票，可能不好

		else:

			return True   #满足条件，就没有 False 在列表中，这样的股票可能在趋势中，要保留



	def judge_stock_bad_trend(self,end_day): #找出过去三天低点和高点每天都下降的股票，实际上就是比较包括今天在内三天的数据。

		code=self.jqcode

		#self.set_point_count_days(gapdays)

		#gapdays=self.gapdays

		df=self.get_point_within_days(end_day,self.trend_days)  #设置三天。可以随时调整。 


		df['yesterday_high']=df['high'].shift(1)  #昨天的最高价，单独提出来。

		df['yesterday_low'] =df['low'].shift(1)   #昨天的最低价，也单独提出来。

		df=df.drop(df.index[0]) #去掉第一行，因为第一行的 yesterday_high 是没有值的，没有前一天的数据，所以不在比较之列。

		#conditionlist=[
		              #(df['low'] > df['yesterday_high']),
		              #(df['high'] < df['yesterday_low']),
		              #]

		#high_gap=round((df['low']/df['yesterday_high']-1)*100,2)

		#low_gap=round((df['high']/df['yesterday_low']-1)*100,2)

		#choicelist=[high_gap,low_gap]

		#df['gap_size']=np.select(conditionlist,choicelist,default=0)   #this block of code is very new to me, to compare the different column, and then make a new column after action.

		bool_value_list=((df['high'] < df['yesterday_high']) & (df['low'] < df['yesterday_low'])).tolist() #高点低点走高就。这个语法以前没用过，用 & 表示或在dataframe中。

		#print(bool_value_list)

		if False in bool_value_list:  #bool_value_list 是一个datafram所有行，通过筛选得到一个个true,false的值的列表

			return False   #这样的股票，

		else:

			return True   #满足条件，就没有 False 在列表中，这样的股票可能在三天连续下跌的趋势中，要注意风险！


	def judge_stock_increase(self,end_day):   #找出连续三天股价上涨的股票。实际是连续两天上涨，按下面的条件选择出来的。

		df=get_price(self.jqcode,frequency=self.frequency,end_date=end_day,count=self.trend_days,fields=['close'])  #时间设置为3天，包括end_day 当天。

		df['yesterday_close']=df['close'].shift(1) 

		df=df.drop(df.index[0])

		#print(df)

		#print(df)

		bool_value_list=(df['close'] > df['yesterday_close']).tolist()    #就是单纯的比较收盘价。

		#print(bool_value_list)

		if False in bool_value_list:   #只要有不满足条件的，那么就是不是连续上涨的股票。就返回 false

			return False

		else:

			return True


	def judge_stock_down(self,end_day):   #找出连续三天股价上涨的股票。实际是连续两天下跌的股票，下面的条件就是这么算的。

		df=get_price(self.jqcode,frequency=self.frequency,end_date=end_day,count=self.trend_days,fields=['close'])  #时间设置为3天，包括end_day 当天。

		df['yesterday_close']=df['close'].shift(1) 

		df=df.drop(df.index[0])

		#print(df)

		bool_value_list=(df['close'] < df['yesterday_close']).tolist()    #就是单纯的比较收盘价。

		#print(bool_value_list)

		if False in bool_value_list:   #只要有不满足条件的，那么就是不是连续下跌的股票。就返回 false，确定不是连续下跌的股票。

			return False

		else:

			return True    #说明是坏股票，连跌三天。
		

	def get_ATR(self,end_date):#normally get ATR for the past 20 days, but will get the data for the past 21 days.

		code=self.jqcode

		#self.set_point_count_days(ATR_days) #这样的话，可以自己选定设定计算多少天的ATR值。

		ATR_days=self.ATR_days

		df=self.get_point_within_days(end_date,ATR_days)

		df['yes_close']=df['close'].shift(1)  #这是获取上一行的close值的语法。shift.

		df=df.drop(df.index[0]) #去掉第一行，因为第一行的 yes_close 是没有值的，没有前一天的数据，所以不在比较之列。 所以最后只有20行。

		df['a1']=df['high'] - df['low']

		df['a2'] = abs(df['high'] - df['yes_close'])

		df['a3'] = abs(df['low'] - df['yes_close'])

		#ATR 公式就是  max((present_high - present_low),abs(present_high - yesterday_close),abs(present_low - yesterday_close))
		#三个值当中选择一个最大的。

		df['ATR'] = df[['a1','a2','a3']].max(axis=1)  #每行当中都比较这三个列的数据，获取最大值

		ATR_value=round(df['ATR'].mean(),4)

		#print(code)

		#print(ATR_value)

		#print(df)

		#print('ATR值为:'+str(ATR_value))

		return ATR_value,ATR_days  #就得到了一个代码在过去多少天之内的ATR平均值，作为风险因子来参考。


	def buy_stocks_volume(self,total_money,end_date):#计算某一支股票/基金ETF到底该买多少股。

		code=self.jqcode

		#ATR_days=self.ATR_days

		ATR_value,ATR_days=self.get_ATR(end_date)

		risk_value=self.risk_value

		max_loss_day=total_money*risk_value

		if ATR_value !=0:

			buying_stock_volume=round(max_loss_day/ATR_value,2)

			present_point,x,y,z=self.get_point()

			buying_money=round(buying_stock_volume*present_point,2)

			sec_name=self.security_name()

			sec_industry=self.security_industry(end_date)

			#print('买入 {},买入数量为 {},买入金额为 {}'.format(sec_name,buying_stock_volume,buying_money))

			return code,sec_name,sec_industry,ATR_value,ATR_days,risk_value,present_point,buying_stock_volume,buying_money

		else:

			buying_stock_volume=0

			present_point=0

			buying_money=0

			sec_name=self.security_name()

			sec_industry=self.security_industry(end_date)

			return code,sec_name,sec_industry,ATR_value,ATR_days,risk_value,present_point,buying_stock_volume,buying_money


	def buy_single_stock_volume(self,total_money,end_date):#计算某一支股票/基金ETF到底该买多少股。与上面的程序类似，但这个只返回一个结果。

		code=self.jqcode

		#ATR_days=self.ATR_days

		ATR_value,ATR_days=self.get_ATR(end_date)

		risk_value=self.risk_value

		max_loss_day=total_money*risk_value

		if ATR_value !=0:

			buying_stock_volume=round(max_loss_day/ATR_value,2)

		else:

			buying_stock_volume=0


		#buying_stock_volume=int(buying_stock_volume/100)*100

		buying_stock_volume=round(buying_stock_volume,-2) #变成100的整数倍的一个数字，是实际的购买量

		buying_stock_volume=int(buying_stock_volume/100)  #变成多少手，1手就是100股，数字就更简化了。


		return buying_stock_volume


	def find_stock_highest_price_time(self,present_day):

		#start_date=present_day+' '+'9:30:00'  #使用get_price,1m 分钟为时间单位获取数据时候，必须要带上时间。

		#end_date=present_day+' '+'15:00:00'

		start_date=present_day+' '+'9:30:00'  #使用get_price,1m 分钟为时间单位获取数据时候，必须要带上时间。

		end_date=present_day+' '+'15:00:00'



		df=get_price(self.jqcode, start_date=start_date, end_date=end_date, frequency='1m')


		highest_price_time=df['high'].idxmax()  #说明准确运行程序，找到了具体的时间。

		#print(highest_price_time)

		#final_time=':'.join([str(highest_price_time.hour),str(highest_price_time.minute),str(highest_price_time.second)])

		if type(highest_price_time)!=float:

			#final_time=datetime.time(highest_price_time.hour,highest_price_time.minute,highest_price_time.second)

			final_time=datetime.time(highest_price_time.hour,highest_price_time.minute).strftime('%H:%M')

		else:

			#final_time=datetime.time(15,0,0) #这种情况下，说明找不到最高点的时间。所以把它设置为收盘时间。

			final_time='15:00'



		return final_time



	def judge_stock_highest_price_time_risk(self,present_day):

		highest_price_time=self.find_stock_highest_price_time(present_day)


		if highest_price_time < '10:00':

		#if highest_price_time < datetime.time(10,0,0):

			warningword='危'

		else:

			warningword=' '


		return warningword



	def get_zhenfu_chart(self,end_day):#得到一个股票，指数，基金的30个交易日内的日内振幅数据，并画出柱形图表示。


		close_list=get_price(self.jqcode,frequency=self.frequency,end_date=end_day,count=(self.zhenfu_count_days+1),fields=['high','low','close'])

		close_list['previous_close']=close_list['close'].shift(1)   #获取昨天的收盘价。

		close_list=close_list.drop(close_list.index[0]) #去掉第一行，因为第一行的 previous_close 是没有值的，没有前一天的数据，所以不在比较之列。

		close_list['zhenfu']=round((close_list['high']-close_list['low'])/close_list['previous_close'],4)*100  # 计算当天的振幅。

		close_list.index = close_list.index.strftime('%m-%d')  #将索引值的日期格式改成月日就行了。这样更简单。

		close_list['zhenfu'].plot.bar(figsize=(12, 7.2))

		#下面这两行是将日期倾斜。

		#ax = close_list['zhenfu'].plot.bar()

		#ax.set_xticklabels(close_list.index.strftime('%m-%d'), rotation=45)

		plt.show()


		return close_list


	def get_zhenfu(self,end_day):  #获得当天的振幅数据


		close_list=get_price(self.jqcode,frequency=self.frequency,end_date=end_day,count=2,fields=['high','low','close'])

		close_list['previous_close']=close_list['close'].shift(1)   #获取昨天的收盘价。

		close_list=close_list.drop(close_list.index[0]) #去掉第一行，因为第一行的 previous_close 是没有值的，没有前一天的数据，所以不在比较之列。

		close_list['zhenfu']=round((close_list['high']-close_list['low'])/close_list['previous_close'],4)*100  # 计算当天的振幅。

		zhenfu_data=list(close_list['zhenfu'])[0]

		return zhenfu_data




	def get_zhenfu_ranking(self,end_day):  #得到包括当天某只股票/指数/etf的振幅排名，从1到30。指的是30天内的振幅表现排名。


		close_list=get_price(self.jqcode,frequency=self.frequency,end_date=end_day,count=(self.zhenfu_count_days+1),fields=['high','low','close'])

		close_list['previous_close']=close_list['close'].shift(1)   #获取昨天的收盘价。

		close_list=close_list.drop(close_list.index[0]) #去掉第一行，因为第一行的 previous_close 是没有值的，没有前一天的数据，所以不在比较之列。

		close_list['zhenfu']=round((close_list['high']-close_list['low'])/close_list['previous_close'],4)*100  # 计算当天的振幅。


		close_list['date']=close_list.index #将index的日期值变成一列。

		dataframe_ranking(close_list,'zhenfu')

		close_list.reset_index(inplace=True)

		close_list['ranking']=close_list.index+1    #这样就按照从高到低，排名，并给予排名。

		close_list=close_list.loc[:,['date','zhenfu','ranking']]


		target_position=np.where(close_list['date']==end_day)[0][0]

		ranking_number_for_zhenfu=close_list.loc[target_position,'ranking']

		ranking_number_for_zhenfu=int(ranking_number_for_zhenfu)

		return ranking_number_for_zhenfu




	def buy_etf_volume(self,total_money,end_date):#计算某一支股票/基金ETF到底该买多少股。

		code=self.jqcode

		#ATR_days=self.ATR_days

		ATR_value,ATR_days=self.get_ATR(end_date)

		risk_value=self.risk_value

		max_loss_day=total_money*risk_value

		if ATR_value !=0:

			buying_stock_volume=round(max_loss_day/ATR_value,2)

			present_point,x,y,z=self.get_point()

			buying_money=round(buying_stock_volume*present_point,2)

			sec_name=self.security_name()

			#sec_industry=self.security_industry(end_date)

			#print('买入 {},买入数量为 {},买入金额为 {}'.format(sec_name,buying_stock_volume,buying_money))

			return code,sec_name,ATR_value,ATR_days,risk_value,present_point,buying_stock_volume,buying_money

		else:

			buying_stock_volume=0

			present_point=0

			buying_money=0

			sec_name=self.security_name()

			#sec_industry=self.security_industry(end_date)

			return code,sec_name,ATR_value,ATR_days,risk_value,present_point,buying_stock_volume,buying_money


	def buy_single_etf_volume(self,total_money,end_date):#计算某一支股票/基金ETF到底该买多少股。

		code=self.jqcode

		#ATR_days=self.ATR_days

		ATR_value,ATR_days=self.get_ATR(end_date)

		risk_value=self.risk_value

		max_loss_day=total_money*risk_value

		if ATR_value !=0 and not math.isnan(ATR_value):  #判断在有数据，并且不是0的情况下。

		#if not math.isnan(ATR_value):  #判断在有数据的情况下。

			buying_stock_volume=round(max_loss_day/ATR_value,2)

		#buying_stock_volume=int(buying_stock_volume/100)*100
			buying_stock_volume=round(buying_stock_volume,-2)  #100整数倍，得到的数字是实际的购买量。

			buying_stock_volume=int(buying_stock_volume/100)  #得到的数字就是多少手，一手100股，数字更简化了。

		else:

			print('找不到ATR的值的代码：或者ATR==0')

			print(code)

			buying_stock_volume=1000000 #使用一个极值，以便于好鉴别。


		return buying_stock_volume



	def set_lowchange_rate(self,target_lowchange_rate=1.3):

		self.lowchange_rate=target_lowchange_rate

		return self.lowchange_rate


    #获取过去几周的证券最高点，最低点的数据。
	def set_high_low_week(self,count=15): #默认是15周，约等于3月。可以自己改。self.high_low_point_count_days 如果要改默认值的话，这里也要改。
		self.week_per_month = count
		return self.week_per_month

	def get_week_high_low(self,check_date): #按周为单位计算，跟交易软件是一样的周线。include_now=True就包括当周。 check_date,就是结束的时间。
		close_list=get_bars(self.jqcode,count=self.week_per_month,unit='1w',fields=['high','low'],end_dt=check_date,include_now=True)
		high_point_pastweeks=close_list['high']
		low_point_pastweeks=close_list['low']
		high_max=high_point_pastweeks.max()
		low_min=low_point_pastweeks.min()
		high_list=list(close_list['high'])
		low_list=list(close_list['low'])

		return high_list,low_list,high_max,low_min


	def get_lowchange_sec_week(self,check_date):

		high_list,low_list,high_max,low_min=self.get_week_high_low(check_date)

		lowchange_rate_standard=self.set_lowchange_rate()  #默认是1.3

		if high_max/low_min <= lowchange_rate_standard:  #三个月,15周内，最高点与最低点波动小雨等于30%，就算是低波动，横盘股。

			#print(self.code)

			return self.code   #返回代码或者是None

	#	if checking_result != None:

	#		self.lowchange_sec_list_week.append(checking_result)

	#	return self.lowchange_sec_list_week   #返回的就是一个周线观察的横盘低波动的股票列表

    #获取过去几月的证券最高点，最低点的数据。
	def set_high_low_month(self,count=3): #默认是三月，可以自己改。self.high_low_point_count_days 如果要改默认值的话，这里也要改。
		self.low_change_month = count
		return self.low_change_month

	def get_month_high_low(self,check_date): #按月为单位计算，跟交易软件是一样的月线。include_now=True就包括当月。决定不包括当月，去掉inlucde_now,默认是false.
		close_list=get_bars(self.jqcode,count=self.low_change_month,unit='1M',fields=['high','low'],end_dt=check_date) #check_date,就是结束的时间
		high_point_months=close_list['high']
		low_point_months=close_list['low']
		high_max=high_point_months.max()
		low_min=low_point_months.min()
		high_list=list(close_list['high'])
		low_list=list(close_list['low'])

		return high_list,low_list,high_max,low_min

	def get_lowchange_sec_month(self,check_date): #观察过去三个月的数据，看是否低波动。

		high_list,low_list,high_max,low_min=self.get_month_high_low(check_date)

		lowchange_rate_standard=self.set_lowchange_rate()  #默认是1.3

		if high_max/low_min <= lowchange_rate_standard:  #三个月内，最高点与最低点波动小雨等于30%，就算是低波动，横盘股。

			return self.code   #返回代码或者是None

    #获取过去几天的证券最高点，最低点的数据。
	def set_high_low_count_days(self,countdays=2):
		self.high_low_point_count_days = countdays
		return self.high_low_point_count_days

	def get_high_low(self):
		close_list=get_bars(self.jqcode,self.high_low_point_count_days,unit=self.unit,fields=['high','low'],end_dt=DateJq().end_date)
		high_point_pastdays=close_list['high']
		low_point_pastdays=close_list['low']
		high_max=high_point_pastdays.max()
		low_min=low_point_pastdays.min()
		high_list=list(close_list['high'])
		low_list=list(close_list['low'])

		return high_list,low_list,high_max,low_min


	def get_max_high_close(self,end_day):  #这个函数中我用了check_date=DateJq().end_date作为参数，不知道是不是一个隐患。20天内的最高价和最高收盘价，不是15天。

		#close_list=get_bars(self.jqcode,self.max_high_close_count_days,unit=self.unit,fields=['high','close'],end_dt=DateJq().end_date)

		close_list=get_price(self.jqcode,frequency=self.frequency,end_date=end_day,count=self.max_high_close_count_days,fields=['high','close'])

		max_high=close_list['high'].max()

		max_close=close_list['close'].max()

		return max_high,max_close


	def get_max_high_min_low(self,start_day,end_day):

		close_list=get_price(self.jqcode,frequency=self.frequency,start_date=start_day,end_date=end_day,fields=['high','low'])

		lowest_price_date = close_list[close_list['low'] == close_list['low'].min()].index[0].strftime('%Y-%m-%d')  #最低价的日期

		highest_price_date = close_list[close_list['high'] == close_list['high'].max()].index[0].strftime('%Y-%m-%d') #最高价的日期

		max_high=close_list['high'].max()

		min_low=close_list['low'].min()

		return max_high,min_low,highest_price_date,lowest_price_date


	def store_high_low_point(self):  
		high_low_data='high_low_point'+str(self.high_low_point_count_days)
		shelveFile=shelve.open(high_low_data)
		indexName=self.security_name()

		highList,lowList,highPoint,lowPoint=self.get_high_low()



		high_low_dic={}

		high_low_dic.setdefault('index_name',indexName)
		high_low_dic.setdefault('high_point',highPoint)
		high_low_dic.setdefault('low_point',lowPoint)
		high_low_dic.setdefault('high_list',highList)
		high_low_dic.setdefault('low_list',lowList)
		#high_low_dic.setdefault('highList_s',highList_s)
		#high_low_dic.setdefault('highList_S',highList_S)
		#high_low_dic.setdefault('lowList_s',lowList_s)
		#high_low_dic.setdefault('lowList_S',lowList_S)

		shelveFile[self.code]=high_low_dic

		shelveFile.close()

	def restore_high_low_point(self):
		high_low_data='high_low_point'+str(self.high_low_point_count_days)
		shelveFile=shelve.open(high_low_data)
		index_code=self.code  #直接使用这个未经过处理的原始代码，作为key，去找到值。
		code_value=shelveFile[index_code]
		index_name=code_value['index_name']
		highList=code_value['high_list']
		lowList=code_value['low_list']
		lowPoint=code_value['low_point']
		highPoint=code_value['high_point']

		highList_s=sorted(highList)                # 正向排序，由小到大的排序
		highList_S=sorted(highList,reverse=True)   # 逆向排序，从大到小排序
		lowList_s=sorted(lowList)
		lowList_S=sorted(lowList,reverse=True)

		#highList_s=code_value['highList_s']             # 正向排序，由小到大的排序
		#highList_S=code_value['highList_S']   # 逆向排序，从大到小排序
		#lowList_s=code_value['lowList_s']
		#lowList_S=code_value['lowList_S']

		shelveFile.close()

		return index_name,highPoint,lowPoint,highList,lowList,highList_s,highList_S,lowList_s,lowList_S

	def create_restore_high_low_point(self): #创立然后使用创立的数据。
		self.store_high_low_point()
		index_name,highPoint,lowPoint,highList,lowList,highList_s,highList_S,lowList_s,lowList_S=self.restore_high_low_point()

		return index_name,highPoint,lowPoint,highList,lowList,highList_s,highList_S,lowList_s,lowList_S

	def get_index_securities(self):
		index_securities=get_index_stocks(self.jqcode)
		return index_securities  #这是一个list，包含所有的代码。当然是符合聚宽的代码标准的代码。

	def get_index_sec_daily_change(self,check_date=DateJq().end_date): #获取一个指数当天的所有股票的涨跌幅的数据的方法。默认是当天，也可以调为指定日期。
		#end_date=DateJq().end_date
		self.set_close_count_days(2)  #获取当天和前一天的数据，所以是两天。
		securities_list=self.get_index_securities()

		

		lower_positive_1=[]  #0--1之间的涨幅
		lower_positive_2=[]  #1-2之间的涨幅
		lower_positive_3=[]  #2-3之间的涨幅
		lower_positive_4=[]  #3-4之间的涨幅
		lower_positive_5=[]  #4-5之间的涨幅
		lower_positive_6=[]  #5-6之间的涨幅
		lower_positive_7=[]  #6-7之间的涨幅
		higher_positive_7=[]  #大于7以上的涨幅

		higher_negative_1=[]  #-1--0之间的涨幅
		higher_negative_2=[]  #-1  -2之间的涨幅
		higher_negative_3=[]  #-2  -3之间的涨幅
		higher_negative_4=[]  #-3  -4之间的涨幅
		higher_negative_5=[]  #-4  -5之间的涨幅
		higher_negative_6=[]  #-5  -6之间的涨幅
		higher_negative_7=[]  #-6  -7之间的涨幅
		lower_negative_7=[] #小于-7以上的涨幅

		for security in securities_list:  #这个securities 就是符合聚宽使用的成份股代码.

			df=get_price(security,frequency=self.frequency,end_date=check_date,count=self.close_count_days,fields=['close'])
			close_start=df.values[0][0]  #开始日收盘价
			close_end=df.values[-1][0]   #结束日收盘价

			change_percent=round((close_end - close_start)/close_start,4)*100   #增长的百分比

			if 0 < change_percent <= 1:

				lower_positive_1.append(change_percent)

			if 1 < change_percent <= 2:

				lower_positive_2.append(change_percent)

			if 2 < change_percent <= 3:

				lower_positive_3.append(change_percent)

			if 3 < change_percent <= 4:

				lower_positive_4.append(change_percent)

			if 4 < change_percent <= 5:

				lower_positive_5.append(change_percent)

			if 5 < change_percent <= 6:

				lower_positive_6.append(change_percent)

			if 6 < change_percent <= 7:

				lower_positive_7.append(change_percent)

			if change_percent > 7:

				higher_positive_7.append(change_percent)

			if -1 < change_percent <= 0:

				higher_negative_1.append(change_percent)

			if -2 < change_percent <= -1:

				higher_negative_2.append(change_percent)

			if -3 < change_percent <= -2:

				higher_negative_3.append(change_percent)

			if -4 < change_percent <= -3:

				higher_negative_4.append(change_percent)

			if -5 < change_percent <= -4:

				higher_negative_5.append(change_percent)

			if -6 < change_percent <= -5:

				higher_negative_6.append(change_percent)

			if -7 < change_percent <= -6:

				higher_negative_7.append(change_percent)

			if change_percent < -7:

				lower_negative_7.append(change_percent)

		index_name=self.security_name()

		index_present_change=self.get_present_close_change(check_date)

		print("{} 当天目前涨跌幅为 {}".format(index_name,index_present_change))

		frame=pd.DataFrame(
			{'{}_close_change'.format(index_name):
		     [
		      len(lower_negative_7),
		      len(higher_negative_7),
		      len(higher_negative_6),
		      len(higher_negative_5), 
		      len(higher_negative_4),
		      len(higher_negative_3), 
		      len(higher_negative_2),
		      len(higher_negative_1), 
		      len(lower_positive_1),
		      len(lower_positive_2),
		      len(lower_positive_3),
		      len(lower_positive_4), 
		      len(lower_positive_5), 
		      len(lower_positive_6),
		      len(lower_positive_7),
		      len(higher_positive_7), 
		      ]
		      },
		      index=[
		      '<-7',
		      '-7 -6',
		      '-6 -5',
		      '-5 -4',
		      '-4 -3',
		      '-3 -2',
		      '-2 -1',
		      '-1 0',
		      '0-1',
		      '1-2',
		      '2-3',
		      '3-4',
		      '4-5',
		      '5-6',
		      '6-7',
		      '>7'])

		all_positive_closechange=lower_positive_1+lower_positive_2+lower_positive_3+lower_positive_4+lower_positive_5+lower_positive_6+lower_positive_7+higher_positive_7
		all_negative_closechange=higher_negative_1+higher_negative_2+higher_negative_3+higher_negative_4+higher_negative_5+higher_negative_6+higher_negative_7+lower_negative_7

		positive_sec_number=len(all_positive_closechange)
		negative_sec_number=len(all_negative_closechange)

		#print(all_positive_closechange)  #打印出涨幅为正的数据

		#print(all_negative_closechange)  #打印出涨幅为负的数据

		all_sec_closechange=all_positive_closechange+all_negative_closechange  #指数所有股票涨跌幅的数据

		all_closechange_median=statistics.median(all_sec_closechange)  #获取所有涨幅的中位数
		

		all_closechange_mean=statistics.mean(all_sec_closechange)  #获取所有涨幅的平均数
		

		#positive_sec_number=len(lower_positive_1)+len(lower_positive_2)+len(lower_positive_3)+len(lower_positive_4)+len(lower_positive_5)+len(lower_positive_6)+len(lower_positive_7)+len(higher_positive_7)
		#negative_sec_number=len(higher_negative_1)+len(higher_negative_2)+len(higher_negative_3)+len(higher_negative_4)+len(higher_negative_5)+len(higher_negative_6)+len(higher_negative_7)+len(lower_negative_7)
		total_sec_number=positive_sec_number+negative_sec_number

		print('{} 涨跌家数比为 {}:{}'.format(index_name,positive_sec_number,negative_sec_number))

		print('{} 所有股票的涨跌幅的中位数 {}'.format(index_name,all_closechange_median))

		print('{} 所有股票的平均涨幅为{}'.format(index_name,all_closechange_mean))


		#frame=pd.DataFrame(data)

		frame.plot(kind='bar')

		print(frame)


        #以下这小段暂时没有用。
		data={
		'0-1':[len(lower_positive_1)],
		'1-2':[len(lower_positive_2)],
		'2-3':[len(lower_positive_3)], 
		'3-4':[len(lower_positive_4)], 
		'4-5':[len(lower_positive_5)], 
		'5-7':[len(lower_positive_7)], 
		'>7':[len(higher_positive_7)], 
		'-1 0':[len(higher_negative_1)], 
		'-2 -1':[len(higher_negative_2)], 
		'-3 -2':[len(higher_negative_3)], 
		'-4 -3':[len(higher_negative_4)], 
		'-5 -4':[len(higher_negative_5)], 
		'-7 -5':[len(higher_negative_7)], 
		'<-7':[len(lower_negative_7)],
		}

	def index_watch(self,check_date=DateJq().end_date): #获取一个指数当天的所有股票的涨跌幅的数据的方法。默认是当天，也可以调为指定日期。
		#end_date=DateJq().end_date
		self.set_close_count_days(2)  #获取当天和前一天的数据，所以是两天。
		securities_list=self.get_index_securities()

		

		lower_positive_1=[]  #0--1之间的涨幅
		lower_positive_2=[]  #1-2之间的涨幅
		lower_positive_3=[]  #2-3之间的涨幅
		lower_positive_4=[]  #3-4之间的涨幅
		lower_positive_5=[]  #4-5之间的涨幅
		lower_positive_6=[]  #5-6之间的涨幅
		lower_positive_7=[]  #6-7之间的涨幅
		higher_positive_7=[]  #大于7以上的涨幅

		higher_negative_1=[]  #-1--0之间的涨幅
		higher_negative_2=[]  #-1  -2之间的涨幅
		higher_negative_3=[]  #-2  -3之间的涨幅
		higher_negative_4=[]  #-3  -4之间的涨幅
		higher_negative_5=[]  #-4  -5之间的涨幅
		higher_negative_6=[]  #-5  -6之间的涨幅
		higher_negative_7=[]  #-6  -7之间的涨幅
		lower_negative_7=[] #小于-7以上的涨幅

		for security in securities_list:  #这个securities 就是符合聚宽使用的成份股代码.

			df=get_price(security,frequency=self.frequency,end_date=check_date,count=self.close_count_days,fields=['close'])
			close_start=df.values[0][0]  #开始日收盘价
			close_end=df.values[-1][0]   #结束日收盘价

			change_percent=round((close_end - close_start)/close_start,4)*100   #增长的百分比

			if 0 < change_percent <= 1:

				lower_positive_1.append(change_percent)

			if 1 < change_percent <= 2:

				lower_positive_2.append(change_percent)

			if 2 < change_percent <= 3:

				lower_positive_3.append(change_percent)

			if 3 < change_percent <= 4:

				lower_positive_4.append(change_percent)

			if 4 < change_percent <= 5:

				lower_positive_5.append(change_percent)

			if 5 < change_percent <= 6:

				lower_positive_6.append(change_percent)

			if 6 < change_percent <= 7:

				lower_positive_7.append(change_percent)

			if change_percent > 7:

				higher_positive_7.append(change_percent)

			if -1 < change_percent <= 0:

				higher_negative_1.append(change_percent)

			if -2 < change_percent <= -1:

				higher_negative_2.append(change_percent)

			if -3 < change_percent <= -2:

				higher_negative_3.append(change_percent)

			if -4 < change_percent <= -3:

				higher_negative_4.append(change_percent)

			if -5 < change_percent <= -4:

				higher_negative_5.append(change_percent)

			if -6 < change_percent <= -5:

				higher_negative_6.append(change_percent)

			if -7 < change_percent <= -6:

				higher_negative_7.append(change_percent)

			if change_percent < -7:

				lower_negative_7.append(change_percent)

		index_name=self.security_name()

		index_present_change=self.get_present_close_change(check_date)

		my_txt=CreateTxt('各指数当天涨跌及成份股涨跌情况明细')

		index_status="{} 当天目前涨跌幅为 {}".format(index_name,index_present_change)

		final_txt=my_txt.txt_add_mode(index_status)

		print(index_status)

		frame=pd.DataFrame(
			{'{}_close_change'.format(index_name):
		     [
		      len(lower_negative_7),
		      len(higher_negative_7),
		      len(higher_negative_6),
		      len(higher_negative_5), 
		      len(higher_negative_4),
		      len(higher_negative_3), 
		      len(higher_negative_2),
		      len(higher_negative_1), 
		      len(lower_positive_1),
		      len(lower_positive_2),
		      len(lower_positive_3),
		      len(lower_positive_4), 
		      len(lower_positive_5), 
		      len(lower_positive_6),
		      len(lower_positive_7),
		      len(higher_positive_7), 
		      ]
		      },
		      index=[
		      '<-7',
		      '-7 -6',
		      '-6 -5',
		      '-5 -4',
		      '-4 -3',
		      '-3 -2',
		      '-2 -1',
		      '-1 0',
		      '0-1',
		      '1-2',
		      '2-3',
		      '3-4',
		      '4-5',
		      '5-6',
		      '6-7',
		      '>7'])

		all_positive_closechange=lower_positive_1+lower_positive_2+lower_positive_3+lower_positive_4+lower_positive_5+lower_positive_6+lower_positive_7+higher_positive_7
		all_negative_closechange=higher_negative_1+higher_negative_2+higher_negative_3+higher_negative_4+higher_negative_5+higher_negative_6+higher_negative_7+lower_negative_7

		positive_sec_number=len(all_positive_closechange)
		negative_sec_number=len(all_negative_closechange)

		positive_vs_negative=round((positive_sec_number/negative_sec_number),2)

		#print(all_positive_closechange)  #打印出涨幅为正的数据

		#print(all_negative_closechange)  #打印出涨幅为负的数据

		all_sec_closechange=all_positive_closechange+all_negative_closechange  #指数所有股票涨跌幅的数据

		all_closechange_median=statistics.median(all_sec_closechange)  #获取所有涨幅的中位数
		

		all_closechange_mean=statistics.mean(all_sec_closechange)  #获取所有涨幅的平均数
		

		#positive_sec_number=len(lower_positive_1)+len(lower_positive_2)+len(lower_positive_3)+len(lower_positive_4)+len(lower_positive_5)+len(lower_positive_6)+len(lower_positive_7)+len(higher_positive_7)
		#negative_sec_number=len(higher_negative_1)+len(higher_negative_2)+len(higher_negative_3)+len(higher_negative_4)+len(higher_negative_5)+len(higher_negative_6)+len(higher_negative_7)+len(lower_negative_7)
		total_sec_number=positive_sec_number+negative_sec_number

		up_down_number='{} 涨跌家数比为 {}:{},上涨/下跌家数比为 {}'.format(index_name,positive_sec_number,negative_sec_number,positive_vs_negative)

		#up_vs_down='{} 上涨/下跌家数比为 {}'.format(index_name,positive_vs_negative)

		middle_sec_change='{} 所有股票的涨跌幅的中位数 {},所有股票的平均涨幅为 {} '.format(index_name,all_closechange_median,all_closechange_mean)

		#average_sec_change='{} 所有股票的平均涨幅为{}'.format(index_name,all_closechange_mean)

		final_txt=my_txt.txt_add_mode(up_down_number)

		#final_txt=my_txt.txt_add_mode(up_vs_down)

		final_txt=my_txt.txt_add_mode(middle_sec_change)

		#final_txt=my_txt.txt_add_mode(average_sec_change)

		frame_csv=frame.to_csv()

		final_txt=my_txt.txt_add_mode(frame_csv)

		return final_txt



	def create_peak_point_stock_sheet_for_index(self):

		present_day=DateJq().end_date

		filename='peak_point_stocks_{}.xlsx'.format(present_day)

		if os.path.isfile(filename):
			pass
		else:
			wb=openpyxl.Workbook()
			sheet=wb.active
			sheet['A1'].value='代码'
			sheet['B1'].value='证券名称'
			sheet['C1'].value='最高收盘价'
			sheet['D1'].value='创新高日期'
			sheet['E1'].value='超过时间标准'
			wb.save(filename)

		return filename

	def create_peak_point_stock_sheet_for_sec(self):

		present_day=DateJq().end_date

		filename='single_sec_peak_point_{}.xlsx'.format(present_day)

		if os.path.isfile(filename):
			pass
		else:
			wb=openpyxl.Workbook()
			sheet=wb.active
			sheet['A1'].value='代码'
			sheet['B1'].value='证券名称'
			sheet['C1'].value='最高收盘价'
			sheet['D1'].value='创新高日期'
			sheet['E1'].value='超过时间标准'
			wb.save(filename)

		return filename


	def find_peak_time_for_sec(self,date_start,target_date): #这个适用于指数代码或者股票代码，就是寻找指数或者股票一个时期内是否有突破新高的记录！
		
		target_sec_code=self.jqcode

		target_sec_name=self.security_name() #指数名称。

		sec_list_xlsx=[]

		research_date=DateJq().end_date

		mycsv_name='个股统计：突破创新高列表{}'.format(research_date)

		mycsv=CreateCsv(mycsv_name)

		mycsv.csv_add_mode('{} {} 创新高:   统计时间是 {}'.format(target_sec_code,target_sec_name,research_date))

		pause_line='______________________________'

		mycsv.csv_add_mode(pause_line)

		#target_csv='个股统计：突破创新高列表{}.csv'.format(research_date)
		#csvFile=open(target_csv,'a',newline='',encoding='gbk')
		#csvWriter=csv.writer(csvFile)

		#write_head_list=[]
		
		#head_content='{} {} 创新高:   统计时间是 {}'.format(target_sec_code,target_sec_name,research_date)
		#write_head_list.append(head_content)
		#csvWriter.writerow(write_head_list)
		#pause_line_list=[]
		#pause_line='______________________________'
		#pause_line_list.append(pause_line)
		#csvWriter.writerow(pause_line_list)

		strictcsv_name='个股统计：优先关注的创新高名单(精选){}'.format(research_date)
		strictcsv=CreateCsv(strictcsv_name)

		strictcsv.csv_add_mode('{} {} 强势创新高:  统计时间是 {} '.format(target_sec_code,target_sec_name,research_date))
		strictcsv.csv_add_mode(pause_line)



		#strict_sec_csv='个股统计：优先关注的创新高名单(精选){}.csv'.format(research_date)
		#strFile=open(strict_sec_csv,'a',newline='',encoding='gbk')
		#strWriter=csv.writer(strFile)

		#str_write_head_list=[]
		#str_head_content='{} {} 强势创新高:  统计时间是 {} '.format(target_sec_code,target_sec_name,research_date)
		#str_write_head_list.append(str_head_content)
		#strWriter.writerow(write_head_list)

		#strWriter.writerow(pause_line_list)

		#filename_store=self.create_peak_point_stock_sheet_for_sec()

		close_df=get_price(target_sec_code,frequency='daily',start_date=date_start,end_date=DateJq().end_date,fields=['close']).sort_values('close',ascending=False)
		peak_close_date=close_df.index[0].strftime('%Y-%m-%d')
		peak_close=close_df.values[0][0]
		second_peak_close_date=close_df.index[1].strftime('%Y-%m-%d')

		
		if peak_close_date > second_peak_close_date or peak_close_date > target_date:  # 通常情况下，新高的时间要大于前高，也又时候，已经创新高了，这两天又回落了，所以可能把最高的时间定在当月就可以。这个不能100%准，定在大盘创新高的那个月份就可以抓住90%以上创新高股票。
			target_sec_name=get_security_info(target_sec_code).display_name
			target_sec_info='{} {} 最高价是 {},时间是 {}'.format(target_sec_code,target_sec_name,peak_close,peak_close_date)
			break_line='                                   '
			#print(target_sec_info)

			mycsv.csv_add_mode(target_sec_info)
			mycsv.csv_add_mode(break_line)

			#write_content_list=[]
			#write_content_list.append(target_sec_info)
			#csvWriter.writerow(write_content_list)
			#write_line_list=[]
			#write_line_list.append(break_line)
			#csvWriter.writerow(write_line_list)


			if peak_close_date > target_date:



				#str_content_list=[]

				strict_target_sec_info='{} {} 最高价是 {},时间是 {} 大于这个 {} 时间的  更值得关注,统计时间是 {}'.format(target_sec_code,target_sec_name,peak_close,peak_close_date,target_date,research_date)

				print(strict_target_sec_info)

				code_normal_use=self.transform_code(target_sec_code)

				self.peak_price_stocks.setdefault(code_normal_use,target_sec_name)

				strictcsv.csv_add_mode(strict_target_sec_info)

				#str_content_list.append(strict_target_sec_info)

				#strWriter.writerow(str_content_list)

				#blank_line_list=[]

				blank_line='                               '

				strictcsv.csv_add_mode(blank_line)

				#blank_line_list.append(blank_line)

				#strWriter.writerow(blank_line_list)

				if check_code_exist(sec_list_xlsx,code_normal_use):

					sub_sec_list=[]
					sub_sec_list.append(code_normal_use)
					sub_sec_list.append(target_sec_name)

					sub_sec_list.append(peak_close)
					sub_sec_list.append(peak_close_date)
					sub_sec_list.append(target_date)

					sec_list_xlsx.append(sub_sec_list)



				#wb=openpyxl.load_workbook(filename_store)
				#sheet=wb.active
				#row_num=sheet.max_row+1
				#existing_code_list=[]
				#for row in range(2,sheet.max_row+1):
				#	existing_code=sheet['A'+str(row)].value
				#	existing_code_list.append(existing_code)

				#if code_normal_use not in existing_code_list:

				#	sheet['A'+str(row_num)].value=code_normal_use
				#	sheet['B'+str(row_num)].value=target_sec_name
				#	sheet['C'+str(row_num)].value=peak_close
				#	sheet['D'+str(row_num)].value=peak_close_date
				#	sheet['E'+str(row_num)].value=target_date

				#	wb.save(filename_store)


		#end_line_list=[]
		end_line='________________________________'
		#end_line_list.append(end_line)

		#csvWriter.writerow(end_line_list)

		mycsv.csv_add_mode(end_line)

		strictcsv.csv_add_mode(end_line)

		#csvFile.close()

		#strWriter.writerow(end_line_list)

		#strFile.close()

		#wb=openpyxl.load_workbook(filename_store)
		#sheet=wb.active

		#for column in range(1,sheet.max_column+1):

		#	sheet.column_dimensions[get_column_letter(column)].width = 15
		#wb.save(filename_store)

		return sec_list_xlsx


	def find_peak_time_for_index(self,date_start,target_date):  #data_start是测量的开始时间日期，target_date是指定一个时间日期，必须在大于这个时间日期的才会理出来。比如20200101
        #主要是为了寻找指数下面的成份股突破新高的个股，设定开始时间和突破的具体时间。

		securities_list=self.get_index_securities()
		if securities_list != []: #首先得是一个不是空的表格，是空的就不做下面的工作了。因为有的指数不是股票指数，所以成份股的列表是空的，这样就把空的指数忽略掉！

			target_index_code=self.jqcode  #指数代码
			target_index_name=self.security_name() #指数名称。

			sec_list_xlsx=[]

			research_date=DateJq().end_date

			mycsv_name='突破创新高个股列表{}'.format(research_date)

			mycsv=CreateCsv(mycsv_name)

			mycsv.csv_add_mode('{} {} 创新高重点关注股票名单如下:   统计时间是 {}'.format(target_index_code,target_index_name,research_date))

			pause_line='______________________________'

			mycsv.csv_add_mode(pause_line)



			#research_date=DateJq().end_date

			#target_csv='突破创新高个股列表{}.csv'.format(research_date)
			#csvFile=open(target_csv,'a',newline='',encoding='gbk')
			#csvWriter=csv.writer(csvFile)

			#write_head_list=[]
			
			#head_content='{} {} 创新高重点关注股票名单如下:   统计时间是 {}'.format(target_index_code,target_index_name,research_date)
			#write_head_list.append(head_content)
			#csvWriter.writerow(write_head_list)
			#pause_line_list=[]
			#pause_line='______________________________'
			#pause_line_list.append(pause_line)
			#csvWriter.writerow(pause_line_list)

			strictcsv_name='优先关注的创新高名单(精选){}'.format(research_date)
			strictcsv=CreateCsv(strictcsv_name)

			strictcsv.csv_add_mode('{} {} 最强创新高的个股名单如下:  统计时间是 {} '.format(target_index_code,target_index_name,research_date))
			strictcsv.csv_add_mode(pause_line)



			#strict_sec_csv='优先关注的创新高名单(精选){}.csv'.format(research_date)
			#strFile=open(strict_sec_csv,'a',newline='',encoding='gbk')
			#strWriter=csv.writer(strFile)

			#str_write_head_list=[]
			#str_head_content='{} {} 最强创新高的个股名单如下:  统计时间是 {} '.format(target_index_code,target_index_name,research_date)
			#str_write_head_list.append(str_head_content)
			#strWriter.writerow(write_head_list)

			#strWriter.writerow(pause_line_list)

			

			#filename_store=self.create_peak_point_stock_sheet_for_index()

			




			for securities in securities_list:
				close_df=get_price(securities,frequency='daily',start_date=date_start,end_date=DateJq().end_date,fields=['close']).sort_values('close',ascending=False)
				peak_close_date=close_df.index[0].strftime('%Y-%m-%d')
				peak_close=close_df.values[0][0]
				second_peak_close_date=close_df.index[1].strftime('%Y-%m-%d')

				
				if peak_close_date > second_peak_close_date or peak_close_date > target_date:  # 通常情况下，新高的时间要大于前高，也又时候，已经创新高了，这两天又回落了，所以可能把最高的时间定在当月就可以。这个不能100%准，定在大盘创新高的那个月份就可以抓住90%以上创新高股票。
					target_security_name=get_security_info(securities).display_name
					target_security_info='{} {} 最高价是 {},时间是 {}'.format(securities,target_security_name,peak_close,peak_close_date)
					break_line='                                   '
					#print(target_security_info)

					mycsv.csv_add_mode(target_security_info)
					mycsv.csv_add_mode(break_line)



					#write_content_list=[]
					#write_content_list.append(target_security_info)
					#csvWriter.writerow(write_content_list)
					#write_line_list=[]
					#write_line_list.append(break_line)
					#csvWriter.writerow(write_line_list)


					if peak_close_date > target_date:



						#str_content_list=[]

						strict_target_security_info='{} {} 最高价是 {},时间是 {} 大于这个 {} 时间的  更值得关注,统计时间是 {}'.format(securities,target_security_name,peak_close,peak_close_date,target_date,research_date)

						print(strict_target_security_info)

						strictcsv.csv_add_mode(strict_target_security_info)


						code_normal_use=self.transform_code(securities)

						self.peak_price_stocks.setdefault(code_normal_use,target_security_name)

						#str_content_list.append(strict_target_security_info)

						#strWriter.writerow(str_content_list)

						#blank_line_list=[]

						blank_line='                               '
						strictcsv.csv_add_mode(blank_line)

						#blank_line_list.append(blank_line)   #这一句写不写无所谓，效果是一样的，可能是因为blank_line是一行空的，没有内容。

						#strWriter.writerow(blank_line_list)

						if check_code_exist(sec_list_xlsx,code_normal_use):

							sub_sec_list=[]
							sub_sec_list.append(code_normal_use)
							sub_sec_list.append(target_security_name)

							sub_sec_list.append(peak_close)
							sub_sec_list.append(peak_close_date)
							sub_sec_list.append(target_date)

							sec_list_xlsx.append(sub_sec_list)

						#wb=openpyxl.load_workbook(filename_store)
						#sheet=wb.active
						#row_num=sheet.max_row+1
						#existing_code_list=[]
						#for row in range(2,sheet.max_row+1):
						#	existing_code=sheet['A'+str(row)].value
						#	existing_code_list.append(existing_code)

						#if code_normal_use not in existing_code_list:

						#	sheet['A'+str(row_num)].value=code_normal_use
						#	sheet['B'+str(row_num)].value=target_security_name
						#	sheet['C'+str(row_num)].value=peak_close
						#	sheet['D'+str(row_num)].value=peak_close_date
						#	sheet['E'+str(row_num)].value=target_date

						#	wb.save(filename_store)


			#end_line_list=[]
			end_line='________________________________'
			mycsv.csv_add_mode(end_line)
			strictcsv.csv_add_mode(end_line)
			#end_line_list.append(end_line)

			#csvWriter.writerow(end_line_list)

			#csvFile.close()

			#strWriter.writerow(end_line_list)

			#strFile.close()

			#wb=openpyxl.load_workbook(filename_store)
			#sheet=wb.active

			#for column in range(1,sheet.max_column+1):

			#	sheet.column_dimensions[get_column_letter(column)].width = 15
			#wb.save(filename_store)

			return sec_list_xlsx








def get_lowchange_seclist(target_change_rate,target_enddate): #这个函数是用来找到过去15周和3个月，横盘高点低点涨跌幅不超过比如1.3，设定一个截止日期比如当前，或者2020年12.31.

	lowchange_sec_list_week=[]
	lowchange_sec_list_month=[]

	for sec in get_all_sec(target_enddate):
		my_jq=Jqdata(sec)
		my_jq.set_lowchange_rate(target_change_rate)
		week_data=my_jq.get_lowchange_sec_week(target_enddate)
		if week_data != None:

			lowchange_sec_list_week.append(sec)

		month_data=my_jq.get_lowchange_sec_month(target_enddate)

		if month_data != None:

			lowchange_sec_list_month.append(sec)

	print(lowchange_sec_list_month)
	print(lowchange_sec_list_week)

#get_lowchange_seclist(1.3,'2021-03-05')

#my_jq=Jqdata('000016sh')
#my_jq.get_index_sec_closechange()
#my_jq=Jqdata('sz399300')
#my_jq.get_index_sec_closechange()

def get_index_sec_closechange(index_code,start_date):  
	my_jq=Jqdata(index_code)
	index_closechange=my_jq.get_close_change_YTD(start_date)
	index_name=my_jq.security_name()
	print("{}{}从{}到现在，涨幅为{}".format(index_name,index_code,start_date,index_closechange))
	all_sec_closechange=[]
	securities_list=my_jq.get_index_securities()
	for security in securities_list:
		security_jq=Jqdata(security)
		present_price,present_open,present_high,present_low=security_jq.get_point()
		security_name=get_security_info(security).display_name
		security_closechange_YTD=security_jq.get_close_change_YTD(start_date)
		print("{} {}从{}到现在，涨幅为{},现价为{}".format(security_name,security,start_date,security_closechange_YTD,present_price))



def get_index_sec_close_change(index_code,start_day,end_day):  #这个函数是获取制定的指数，在制定的时间内，从起始日的收盘价，到截止日的收盘价，中间的所有成份的涨跌幅排序，以便选出强势股。

	my_jq=Jqdata(index_code)

	index_closechange=my_jq.get_close_change_within_days(start_day,end_day)

	index_name=my_jq.security_name()

	print("{}{}从{}开始到{}，涨了{}".format(index_name,index_code,start_day,end_day,index_closechange))

	all_sec_closechange=[]

	securities_list=my_jq.get_index_securities()

	csv_file_name="{}{}成份股涨跌幅".format(index_name,index_code)

	my_csv=CreateCsv(csv_file_name)

	csv_head_content=[]

	csv_head_content.append('公司')

	csv_head_content.append('涨跌幅')

	csv_head_content.append('起始日')

	csv_head_content.append('截止日')

	csv_head_content.append('行业涨 {}'.format(index_closechange))

	my_csv.create_csv_from_list(csv_head_content)
	

	for security in securities_list:

		sec_info_list=[]

		security_jq=Jqdata(security)

		prsent_price,present_open,present_high,prsent_low=security_jq.get_point()

		security_name=get_security_info(security).display_name

		security_closechange_within_days=security_jq.get_close_change_within_days(start_day,end_day)

		sec_info_list.append(security_name)

		sec_info_list.append(security_closechange_within_days)

		sec_info_list.append(start_day)

		sec_info_list.append(end_day)

		my_csv.create_csv_from_list(sec_info_list)

	csvToExcel(csv_file_name+'.csv')

	sort_excel(csv_file_name+'.xlsx','B')



def check_code_exist(target_list_added,code_added):  #用来检验某个代码的信息是否已经生成表中表。

	code_existing_list=[]

	for code_list in target_list_added:
		code_existing=code_list[0]
		code_existing_list.append(code_existing)

	if code_added not in code_existing_list:

		return True

def judge_list(target_list): #如果列表中有数据，则判断为True
	
	
	if len(target_list)!=0:
		return True


start_date='2021-12-01'   #起始时间日期

standard_date='2021-12-16'  #在该时间之后创新高的更有意义


def find_strong_sec():

	index_list=[]
	index_code_list=[]
	sec_info_list=[]

	for index in all_index_list:
		my_jq=Jqdata(index)
		index_list += my_jq.find_peak_time_for_sec(start_date,standard_date) #找出最强的突破新高的指数，然后再去发现它们的成份股有哪些在同样的时间段内有突破新高的表现！

	print(index_list)


	if judge_list(index_list):

		#sort_excel(new_peak_index_list,'D')  #按照降序排列

		#my_email=sending_email(['lisztony@163.com','lisztony@icloud.com'])

		#email_subject='率先创新高的指数名单 {}'.format(present_day)

		#email_content='please see the attached!'

		#my_email.send_attachment(email_subject,email_content,new_peak_index_list)


		#wb=openpyxl.load_workbook(new_peak_index_list)

		#sheet=wb.active

		#index_code_list=[]

		for sub_list in index_list:

			index_code=sub_list[0]

			index_code_list.append(index_code)

		print(index_code_list)

		for index_code in index_code_list:

			my_jq=Jqdata(index_code)

			index_stock_list=my_jq.get_index_securities()

			index_name=my_jq.security_name()

			print(index_code,index_name)

			if index_stock_list !=[]:

				new_sec_list=my_jq.find_peak_time_for_index(start_date,standard_date)

			for sec_list in new_sec_list:

				sec_code=sec_list[0]

				if check_code_exist(sec_info_list,sec_code):
					sec_info_list.append(sec_list)

			#file1='peak_point_stocks.xlsx'

			#file2='优先关注的创新高名单(精选){}.csv'.format(present_day)

			#email_subject1='最先创新高的指数成份股名单(excel版)'

			#email_subject2='最先创新高的指数成份股名单(详细文本版)'

			#email_content='See the attachment!'

			#my_email=sending_email(['lisztony@163.com','lisztony@icloud.com'])

			#my_email.send_attachment(email_subject1,email_content,file1)

			#my_email.send_attachment(email_subject2,email_content,file2)

		print(sec_info_list)



def judge_indexoflist_above_ma(index_list,ma_days,end_day):

	for index in index_list:

		my_jq=Jqdata(index)

		my_jq.judge_index_above_ma(ma_days,end_day)



def get_allsec_above_ma_number(ma_days,start_date,end_date): #找出超越全市场超越60天/20天线的股票的数量，以此判断市场的热度，牛熊状况。


	# set the start and end dates for the 30-day period
	#start_date = '2022-04-01'
	#end_date = '2022-04-28'

	# get all stocks in the China A-share market
	stocks = get_all_securities(['stock'], date=end_date).index

	# create an empty DataFrame to store the results
	results = pd.DataFrame(columns=['number_data'])

	# iterate over each day in the period
	for date in pd.date_range(start=start_date, end=end_date):
	    # initialize a counter for the number of stocks above their MA5
	    count = 0
	    
	    # iterate over each stock
	    for stock in stocks:
	        # get the stock's close price data up to the current date
	        prices = get_price(stock, end_date=date, count=ma_days, frequency='daily', fields=['close'])['close']
	        
	        # calculate the stock's 5-day moving average
	        ma5 = prices.mean()
	        
	        # check if the stock's close price is above its MA5
	        if prices[-1] > ma5:
	            count += 1
	    
	    # store the result for the current date
	    results.loc[date, 'number_data'] = count


	return results


def get_index_stocklist_above_ma_number(index_code,ma_days,start_date,end_date): #找出某个指数的所有成分股，超越全市场超越60天/20天线的股票的数量，以此判断市场的热度，牛熊状况。



	# set the start and end dates for the 30-day period
	#start_date = '2022-04-01'
	#end_date = '2022-04-28'

	# get all stocks in the China A-share market
	#stocks = get_all_securities(['stock'], date=end_date).index

	stocklist=Jqdata(index_code).get_index_securities()



	stocks = stocklist

	# create an empty DataFrame to store the results

	column_name='ma_'+str(ma_days)

	results = pd.DataFrame(columns=[column_name])

	# iterate over each day in the period
	for date in pd.date_range(start=start_date, end=end_date):
	    # initialize a counter for the number of stocks above their MA5
	    count = 0
	    
	    # iterate over each stock
	    for stock in stocks:
	        # get the stock's close price data up to the current date
	        prices = get_price(stock, end_date=date, count=ma_days, frequency='daily', fields=['close'])['close']
	        
	        # calculate the stock's 5-day moving average

	        	        
	        ma5= prices.mean()
	        
	        # check if the stock's close price is above its MA5
	        if prices[-1] > ma5:
	            count += 1


	    
	    # store the result for the current date
	    results.loc[date, column_name] = count


	return results   #得到pd的数据结果，日期和超越局限的个股数据。



def get_index_sec_above_ma_number_chart(index_code,ma_days_1,ma_days_2,start_date,end_date): #观察超越两个不同均线的指数成分股的数量，观察指数是否出现内部强弱分化，并画图出来。


	ma_1=get_index_stocklist_above_ma_number(index_code,ma_days_1,start_date,end_date)

	ma_2=get_index_stocklist_above_ma_number(index_code,ma_days_2,start_date,end_date)

	result = pd.concat([ma_1, ma_2], axis=1)

	result.plot(kind='line',figsize=(12, 7.2))


	plt.show()

	if end_date==datetime.datetime.now().strftime('%Y-%m-%d'):

		date_time=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  #如果要看的是当前日期的，就这么写，可以精确到具体时间。

	else:

		date_time=end_date

	print('数据日期：{}'.format(date_time))


	return result



def get_index_trend_by_ma_number_percent_chart(index_code,ma_days_1,ma_days_2,start_date,end_date):




	index_sec_number=len(Jqdata(index_code).get_index_securities())


	ma_1=get_index_stocklist_above_ma_number(index_code,ma_days_1,start_date,end_date)

	column_1='ma_'+str(ma_days_1)

	ma_1[column_1]=round(ma_1[column_1]/index_sec_number,4)*100


	ma_2=get_index_stocklist_above_ma_number(index_code,ma_days_2,start_date,end_date)

	column_2='ma_'+str(ma_days_2)

	ma_2[column_2]=round(ma_2[column_2]/index_sec_number,4)*100

	

	result = pd.concat([ma_1, ma_2], axis=1)


	result.plot(kind='line',figsize=(12, 7.2))


	plt.show()

	if end_date==datetime.datetime.now().strftime('%Y-%m-%d'):

		date_time=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  #如果要看的是当前日期的，就这么写，可以精确到具体时间。

	else:

		date_time=end_date

	print('数据日期：{}'.format(date_time))

	return result



def get_below_different_ma_df_stocklist(stocklist,total_money,end_day):

	#获得高于5天线的股票

	df_5=Jq_codelist(stocklist).get_below_ma_stocklist(total_money,5,end_day)

	#获得高于10天线的股票

	df_10=Jq_codelist(stocklist).get_below_ma_stocklist(total_money,10,end_day)

	#获得高于20天线的股票

	df_20=Jq_codelist(stocklist).get_below_ma_stocklist(total_money,20,end_day)

	return df_5,df_10,df_20



def get_above_different_ma_df_stocklist(stocklist,total_money,end_day):

	#获得高于5天线的股票

	df_5=Jq_codelist(stocklist).get_above_ma_stocklist(total_money,5,end_day)

	#获得高于10天线的股票

	df_10=Jq_codelist(stocklist).get_above_ma_stocklist(total_money,10,end_day)

	#获得高于10天线的股票

	df_20=Jq_codelist(stocklist).get_above_ma_stocklist(total_money,20,end_day)

	return df_5,df_10,df_20



def judge_stocksoflist_gapsize(stocklist,end_day): #找出股票列表中，缺口口大于9.9%或者小于-9.9%的股票，弃之不用。

	del_items_list=[]  #不符合条件，缺口过大的股票，将放到这个表里，下一步讲这些从原表中全部删除。

	if stocklist !=[]:

		for stock in stocklist:

			my_jq=Jqdata(stock)

			bool_value,gap_days,gap_size=my_jq.judge_stock_gapsize(end_day)

			if bool_value==True:

				del_items_list.append(stock)   

				#stocklist.remove(stock)   这种算法不对，如果边循环，边删除，那么第一个数据删除后，循环就会停止，所以不能对在循环中的list做list.remove()的动作。

		for item in del_items_list:

			stocklist.remove(item)

		#print(len(stocklist))

		return stocklist,gap_days #返回一个去掉缺口大于9.99%的所有股票的列表。



def filter_with_ma_gapsize(stockslist,madays,end_date):  

	stocks_list,ma_days=Jq_codelist(stockslist).judge_stocklist_above_ma(madays,end_date) #先用均线删除不必要的股票，过滤。

	if stocks_list!=[]:

		sec_list,gap_days=judge_stocksoflist_gapsize(stocks_list,end_date)  #再用缺口的大小去过滤不需要的股票。

		if sec_list!=[]:

			return sec_list,ma_days,gap_days

		else:

			return None,None,None
	else:

		return None,None,None


def filter_index_list(index_list,madays,end_date):

	del_index_list=[]

	for index in index_list:

		my_jq=Jqdata(index)

		securities_list=my_jq.get_index_securities()

		sec_list,ma_days,gap_days=filter_with_ma_gapsize(securities_list,madays,end_date)

		if sec_list==None:

			del_index_list.append(index)

	for item in del_index_list:

		index_list.remove(item)

		print(item+'没有一个股票满足100天线或者缺口小于10，被删除！！')

	return index_list





#计算出一个股票或者基金列表里面的数据，每一个应该买入多少,并以表格形式表现出来。
def buy_listofstocks_volume_filter(stockslist,madays,total_money,end_date):


	stocks_list,ma_days=Jq_codelist(stockslist).judge_stocklist_above_ma(madays,end_date) #先用均线删除不必要的股票，过滤。


	stocks_list,gap_days=judge_stocksoflist_gapsize(stocks_list,end_date)  #再用缺口的大小去过滤不需要的股票。


	#stocks_list,ma_days,gap_days,gap_size=filter_with_ma_gapsize(stockslist,madays,gapdays,end_date,gapsize)


	list_for_df=[]

	for stock in stocks_list:

		list_for_stock=[]

		my_jq=Jqdata(stock)

		ATR_days=my_jq.ATR_days  

		R2_score,score_days=my_jq.momentum_score(end_date)

		stock_code,stock_name,stock_industry,ATR_value,ATR_days,risk_value,present_point,buying_stock_volume,buying_money=my_jq.buy_stocks_volume(total_money,end_date)

		#stock_code=stock_code.split('.')[0] #这就是将聚宽认定的代码后四位去掉，只留下大家普遍使用的代码的格式。以便于后续的工作

		if stock_code.startswith('30') or stock_code.startswith('688'):

			gap_size=20

		else:

			gap_size=10

		stock_code=my_jq.transform_code(stock_code)  #这就是将聚宽认定的代码后四位去掉，只留下大家普遍使用的代码的格式。sz399905 sh000016,以便于后续的工作


		pct_money=(round(buying_money/total_money,4))*100

		data_date=end_date

		close_change=my_jq.get_close_change(data_date,1)

		money_change=my_jq.get_money_change(data_date)

		volume_change=my_jq.get_volume_change(data_date)

		list_for_stock=[
		               stock_code,
		               stock_name,
		               stock_industry,
		               R2_score,
		               ATR_value,
		               buying_stock_volume,
		               present_point,
		               buying_money,
		               pct_money,
		               close_change,
		               volume_change,
		               money_change,
		               data_date,
		               risk_value,
		               ma_days,
		               (ATR_days-1),
		               score_days,
		               (gap_days-1),
		               gap_size,
		               ]
		list_for_df.append(list_for_stock)


	df=pd.DataFrame(list_for_df)

	df.columns=['Code','Name','Industry','Score','ATR_value','buy_volume',
	            'price_now','buy_value','weight','closechange','volumechange','moneychange',
	            'Data_time','risk','Ma_d','ATR_d','Score_d','Gap_d','Gap_size']

	df.drop(df[pd.isna(df['Score'])].index,inplace=True)    #在本dataframe 中'Score'列中，分数为nan(这是一个非字符的float类型的数据，用pd.isna()方法判断这是一个nan数据类型的，确定后就去掉)

	#为什么会产生nan,因为score是90天的数据，而当我们获取超过20天均线的股票数据时，刚上市不到90的股票超过20天均线也倍选中了，但不到90天没有分数，没有分数的就要被放弃，因为打分是我们的核心。

	df.sort_values(by=['Score'],ascending=False,inplace=True)

	money_used=df['buy_value'].sum()

	print('TILL now '+str(money_used)+' invested in the stocks')

	print('************************************')

	

	return df




#计算出一个股票或者基金列表里面的数据，每一个应该买入多少,并以表格形式表现出来。
def buy_listofetf_volume_filter(etf_list,madays,total_money,end_date):


	etflist,ma_days=Jq_codelist(etf_list).judge_stocklist_above_ma(madays,end_date) #先用均线删除不必要的股票，过滤。


	etf_list,gap_days=judge_stocksoflist_gapsize(etflist,end_date)  #再用缺口的大小去过滤不需要的股票。


	#stocks_list,ma_days,gap_days,gap_size=filter_with_ma_gapsize(stockslist,madays,gapdays,end_date,gapsize)


	list_for_df=[]

	for stock in etf_list:

		list_for_stock=[]

		my_jq=Jqdata(stock)

		ATR_days=my_jq.ATR_days  

		R2_score,score_days=my_jq.momentum_score(end_date)

		stock_code,stock_name,ATR_value,ATR_days,risk_value,present_point,buying_stock_volume,buying_money=my_jq.buy_etf_volume(total_money,end_date)

		#stock_code=stock_code.split('.')[0] #这就是将聚宽认定的代码后四位去掉，只留下大家普遍使用的代码的格式。以便于后续的工作

		stock_code=my_jq.transform_code(stock_code)  #这就是将聚宽认定的代码后四位去掉，只留下大家普遍使用的代码的格式。sz399905 sh000016,以便于后续的工作

		gap_size=10


		pct_money=(round(buying_money/total_money,4))*100

		data_date=end_date

		close_change=my_jq.get_close_change(data_date,1)

		money_change=my_jq.get_money_change(data_date)

		volume_change=my_jq.get_volume_change(data_date)

		list_for_stock=[
		               stock_code,
		               stock_name,
		               R2_score,
		               ATR_value,
		               buying_stock_volume,
		               present_point,
		               buying_money,
		               pct_money,
		               close_change,
		               volume_change,
		               money_change,
		               data_date,
		               risk_value,
		               ma_days,
		               (ATR_days-1),
		               score_days,
		               (gap_days-1),
		               gap_size,
		               ]
		list_for_df.append(list_for_stock)


	df=pd.DataFrame(list_for_df)

	df.columns=['Code','Name','Score','ATR_value','buy_volume',
	            'price_now','buy_value','weight','closechange','volumechange','moneychange',
	            'Data_time','risk','Ma_d','ATR_d','Score_d','Gap_d','Gap_size']

	df.drop(df[pd.isna(df['Score'])].index,inplace=True)    #在本dataframe 中'Score'列中，分数为nan(这是一个非字符的float类型的数据，用pd.isna()方法判断这是一个nan数据类型的，确定后就去掉)

	#为什么会产生nan,因为score是90天的数据，而当我们获取超过20天均线的股票数据时，刚上市不到90的股票超过20天均线也倍选中了，但不到90天没有分数，没有分数的就要被放弃，因为打分是我们的核心。

	df.sort_values(by=['Score'],ascending=False,inplace=True)

	money_used=df['buy_value'].sum()

	print('TILL now '+str(money_used)+' invested in the stocks')

	print('************************************')

	

	return df



#计算出一个股票或者基金列表里面的数据，每一个应该买入多少,并以表格形式表现出来。
def buy_alletf_volume(etf_list,total_money,end_date):

	df=pd.DataFrame(etf_list,columns=etf_columns)

	etflist=list(df.code)


	list_for_df=[]

	for stock in etflist:

		list_for_stock=[]

		my_jq=Jqdata(stock)

		ATR_days=my_jq.ATR_days  

		R2_score,score_days=my_jq.momentum_score(end_date)

		stock_code,stock_name,ATR_value,ATR_days,risk_value,present_point,buying_stock_volume,buying_money=my_jq.buy_etf_volume(total_money,end_date)

		#stock_code=stock_code.split('.')[0] #这就是将聚宽认定的代码后四位去掉，只留下大家普遍使用的代码的格式。以便于后续的工作

		stock_code=my_jq.transform_code(stock_code)  #这就是将聚宽认定的代码后四位去掉，只留下大家普遍使用的代码的格式。sz399905 sh000016,以便于后续的工作


		pct_money=(round(buying_money/total_money,4))*100

		data_date=end_date

		close_change=my_jq.get_close_change(data_date,1)

		money_change=my_jq.get_money_change(data_date)

		volume_change=my_jq.get_volume_change(data_date)

		list_for_stock=[
		               stock_code,
		               stock_name,
		               R2_score,
		               ATR_value,
		               buying_stock_volume,
		               present_point,
		               buying_money,
		               pct_money,
		               close_change,
		               volume_change,
		               money_change,
		               data_date,
		               risk_value,
		               (ATR_days-1),
		               score_days,
		               ]
		list_for_df.append(list_for_stock)


	df=pd.DataFrame(list_for_df)

	df.columns=['Code','Name','Score','ATR_value','buy_volume',
	            'price_now','buy_value','weight','closechange','volumechange','moneychange',
	            'Data_time','risk','ATR_d','Score_d',]

	df.drop(df[pd.isna(df['Score'])].index,inplace=True)    #在本dataframe 中'Score'列中，分数为nan(这是一个非字符的float类型的数据，用pd.isna()方法判断这是一个nan数据类型的，确定后就去掉)

	#为什么会产生nan,因为score是90天的数据，而当我们获取超过20天均线的股票数据时，刚上市不到90的股票超过20天均线也倍选中了，但不到90天没有分数，没有分数的就要被放弃，因为打分是我们的核心。

	df.sort_values(by=['Score'],ascending=False,inplace=True)

	money_used=df['buy_value'].sum()

	print('TILL now '+str(money_used)+' invested in all etfs')

	print('************************************')

	df_list,df_columns=dataframe2list(df)

	print(''.join(['df_list_alletf='+str(df_list)]))

	print(''.join(['df_columns_alletf='+str(df_columns)]))

	return df_list,df_columns



def daily_etf_check(madays,total_money,end_date):

	df=pd.DataFrame(etf_list,columns=etf_columns)

	etf_code_list=list(df.code)

	etf_df=buy_listofetf_volume_filter(etf_code_list,madays,total_money,end_date)

	df_list,df_columns=dataframe2list(etf_df)

	print(''.join(['df_list_etf='+str(df_list)]))

	print(''.join(['df_columns_etf='+str(df_columns)]))

	#print(df_list)

	#print(df_columns)

	return df_list,df_columns





def checking_list_for_repeat(check_list):   #这个函数是为了避免给出的表里面有重复的元素，每次运行之前之前，检查一下这个。


    unique_items = list(set(check_list))  #集合是最好的一种处理重复元素的方法。

    return unique_items




def create_index_list_dic(index_list): #生成一个如上显示的指数名称和成分股个数的字典。
    
    index_list_dic={}

    for index in index_list:

        index_dic={}

        my_jq=Jqdata(index)

        index_name=my_jq.security_name()

        security_list=my_jq.get_index_securities()

        security_num=len(security_list)

        index_dic.setdefault('name',index_name)

        index_dic.setdefault('number',security_num)

        index_list_dic.setdefault(index,index_dic)

    print(index_list_dic)

    return index_list_dic

#需要监测的指数列表。

index_list=['sh000688','sh000682','sh000685', 'sh000852', 'sz399905', 'sh000016', 'sh000097', 'sh000805', 'sz399006', 'sz399004', 'sz399005', 'sz399300', 'sz399330', 'sz399333', 'sz399612', 'sz399673', 'sz399903', 'sh000922',
 'sz399324', 'sh000812', 'sh000811', 'sh000810', 'sh000809', 'sh000813', 'sh000814', 'sh000815', 'sh000816', 'sh000818', 'sh000819', 'sh000827', 'sh000901', 'sh000928', 'sh000929', 'sh000930', 'sh000931', 
 'sz399932', 'sh000933', 'sh000934', 'sh000935', 'sh000941', 'sh000977', 'sh000978', 'sh000979', 'sh000986', 'sh000987', 'sh000988', 'sh000989', 'sh000990', 'sh000991', 'sh000992', 'sh000993', 'sh000998', 
 'sz399030', 'sz399060', 'sz399295', 'sz399296', 'sz399393', 'sz399394', 'sz399395', 'sz399396', 'sz399412', 'sz399417', 'sz399419', 'sz399420', 'sz399423', 'sz399432', 'sz399435', 'sz399438', 'sz399439',
  'sz399440', 'sz399441', 'sz399608', 'sz399610', 'sz399646', 'sz399647', 'sz399653', 'sz399654', 'sz399669', 'sz399674', 'sz399675', 'sz399676', 'sz399677', 'sz399683', 'sz399684', 'sz399687', 'sz399693', 
  'sz399695', 'sz399699', 'sz399704', 'sz399705', 'sz399706', 'sz399804', 'sz399805', 'sz399806', 'sz399809', 'sz399807', 'sz399808', 'sz399810', 'sz399811', 'sz399812', 'sz399813', 'sz399814', 'sz399967', 
  'sz399970', 'sz399971', 'sz399973', 'sz399975', 'sz399976', 'sz399986', 'sz399987', 'sz399989', 'sz399993', 'sz399994', 'sz399995', 'sz399996', 'sz399997', 'sz399998','sz399966',]


index_list=checking_list_for_repeat(index_list)


def daily_check(index_list,madays,total_money,end_date):

	judge_indexoflist_above_ma(index_list,madays,end_date)

	index_list=filter_index_list(index_list,madays,end_date)  #先将没有一个股票满足100天线或者缺口大于15的指数删除。仅仅是为了过滤掉返回空值的指数。否则就会导致后面的程序出错。

	#但是上面的步骤只是删除了不满足条件的指数，虽然在这个过程中，对所有指数的股票进行了100天线和15缺口的过滤，但下面的buy_listofstocks_volume_filter还会将这个过程重新做一遍
	#因为这次过滤的结果就是将符合条件的股票选择出来，保存起来，并进行数据清理和分析，

	code_df_list_dic={}

	for index_code in index_list:

		print(index_code)

		list_name=index_code

		my_jq=Jqdata(index_code)

		stockslist=my_jq.get_index_securities()

		#df_name=index_code+'_df'

		

		df=buy_listofstocks_volume_filter(stockslist,madays,total_money,end_date)

		#df=df.reset_index()  #生成index从0开始的dataframe

		#df['ranking']=df.index + 1 

		df_list,df_columns=dataframe2list(df)

		#new_df=add_ranking_into_dataframe(df,'ranking',3)

		#df_list,df_columns=dataframe2list(new_df)

		code_df_list_dic.setdefault(list_name,df_list)



	print(''.join(['df_list='+str(code_df_list_dic)]))

	print(''.join(['df_columns='+str(df_columns)]))


	#print(code_df_list_dic)

		#print(list_name,df_list,sep='=')

		#print(df_columns)

	return df_list,df_columns


def daily_check_above_ma_seclist(madays,end_date): #生成一个当天突破100天线的股票的列表。

	stocklist=get_all_sec(end_date)

	above_ma_list=[]

	for stock in stocklist:

		my_jq=Jqdata(stock)

		if my_jq.find_above_ma_sec(madays,end_date):

			above_ma_code=stock

			above_ma_list.append(above_ma_code)


	return above_ma_list



def check_above_ma_seclist(stocklist,madays,end_date): #生成一个当天突破100天线的股票的列表。

	above_ma_list=[]

	for stock in stocklist:

		my_jq=Jqdata(stock)

		if my_jq.find_above_ma_sec(madays,end_date):

			above_ma_code=stock

			above_ma_list.append(above_ma_code)


	return above_ma_list



def daily_above_ma_gapsize(above_ma_list,end_date):

	sec_list,gap_days=judge_stocksoflist_gapsize(above_ma_list,end_date)  #用缺口的大小去过滤不需要的股票。

	if sec_list!=[]:

		return sec_list,gap_days


	else:

		return None,None



def daily_above_ma_score(madays,total_money,end_date):

	above_ma_list=daily_check_above_ma_seclist(madays,end_date)

	sec_list,gap_days=daily_above_ma_gapsize(above_ma_list,end_date)

	#code_df_list_dic={}

	if sec_list !=[]:

		#list_name='above_ma_today_sec_list'

		df=buy_listofstocks_volume_filter(sec_list,madays,total_money,end_date)

		df_list,df_columns=dataframe2list(df)

		#new_df=add_ranking_into_dataframe(df,'ranking',3)

		#df_list,df_columns=dataframe2list(new_df)

		print(''.join(['df_list_100='+str(df_list)]))

		print(''.join(['df_columns_100='+str(df_columns)]))

		#print(df_list)

		#print(df_columns)

		return df_list,df_columns




def daily_check_all_sec(madays,total_money,end_date):


	stockslist=get_all_sec(end_date)

	df=buy_listofstocks_volume_filter(stockslist,madays,total_money,end_date)

	df_list,df_columns=dataframe2list(df)

	print(''.join(['df_list_all_sec='+str(df_list)]))

	print(''.join(['df_columns='+str(df_columns)]))

	return df_list,df_columns


def find_stock_in_trend(ma_days,end_day):  #获取全市场中，过去3天，连续两天上涨的指数。

	stockslist=get_all_sec(end_day)

	# 找出连涨2天的股票。

	good_trend_stock_list=[good_stock for good_stock in stockslist if Jqdata(good_stock).judge_stock_increase(end_day)==True]  #股票代码没有变化，按平台来的。

	good_trend_stock_list_code_transformed=[Jqdata(sec).transform_code(sec) for sec in good_trend_stock_list ]  #股票代码发生变化，按腾讯来的。

	print(''.join(['stock_in_trend=',str(good_trend_stock_list_code_transformed)]))

	print(''.join(['good_stock_date=',str(end_day)]))

	#以下是将上面的代码，代码转换，获取股票名称和2日内的涨幅，排序，然后做成列表和列的数据，打印出来。 改成了两日内的涨幅，因为选择实际也是选择连续两天上涨的股票，那么就是计算这两天上涨的涨幅，而不应该是计算三天。

	good_trend_stock_list_with_closechange=[[Jqdata(good_stock).transform_code(good_stock),Jqdata(good_stock).security_name(),Jqdata(good_stock).get_close_change(end_day,2)] for good_stock in good_trend_stock_list]
	
	columns_good_trend_stock_list_with_closechange=['code','name','closechange']
	
	good_trend_stock_df=list2dataframe(good_trend_stock_list_with_closechange,columns_good_trend_stock_list_with_closechange)

	good_trend_stock_df.drop(good_trend_stock_df[pd.isna(good_trend_stock_df['closechange'])].index,inplace=True)    #在本dataframe closechange 列中，分数为nan(这是一个非字符的float类型的数据，用pd.isna()方法判断这是一个nan数据类型的，确定后就去掉)
	
	dataframe_ranking(good_trend_stock_df,'closechange')  #排序

	#print(len(good_trend_stock_df))

	good_trend_stock_list_with_closechange,columns_good_trend_stock_list_with_closechange=dataframe2list(good_trend_stock_df)

	print(''.join(['good_trend_stock_with_closechange_list=',str(good_trend_stock_list_with_closechange)]))

	print(''.join(['good_trend_stock_with_closechange_columns=',str(columns_good_trend_stock_list_with_closechange)]))

	#获取3天连涨的股票列表中的，超越5天线的股票。

	list_name=''.join(['above_ma_stock_list_',str(ma_days)])

	above_ma_stock_list=[goodsec for goodsec in good_trend_stock_list if Jqdata(goodsec).judge_above_ma(ma_days,end_day)==True]   #股票代码没有变化，按平台来的。

	above_ma_stock_list_code_transformed=[Jqdata(goodsec).transform_code(goodsec) for goodsec in above_ma_stock_list]  #股票代码发生变化，按腾讯来的。

	print(''.join([list_name,'=',str(above_ma_stock_list_code_transformed)]))

	#以下是将上面超越5天线的的代码，代码转换，获取股票名称和3日内的涨幅，排序，然后做成列表和列的数据，打印出来。改成了两日内的涨幅，因为选择实际也是选择连续两天上涨的股票，那么就是计算这两天上涨的涨幅，而不应该是计算三天。

	above_ma_stock_list_with_closechange=[[Jqdata(good_stock).transform_code(good_stock),Jqdata(good_stock).security_name(),Jqdata(good_stock).get_close_change(end_day,2)] for good_stock in above_ma_stock_list]
	
	columns_above_ma_stock_list_with_closechange=['code','name','closechange']
	
	above_ma_stock_df=list2dataframe(above_ma_stock_list_with_closechange,columns_above_ma_stock_list_with_closechange)

	above_ma_stock_df.drop(above_ma_stock_df[pd.isna(above_ma_stock_df['closechange'])].index,inplace=True)    #在本dataframe 中'closechange'列中，分数为nan(这是一个非字符的float类型的数据，用pd.isna()方法判断这是一个nan数据类型的，确定后就去掉)
	
	dataframe_ranking(above_ma_stock_df,'closechange')  #排序

	#print(len(above_ma_stock_df))

	above_ma_stock_list_with_closechange,columns_above_ma_stock_list_with_closechange=dataframe2list(above_ma_stock_df)

	print(''.join(['above_ma_stock_with_closechange_list=',str(above_ma_stock_list_with_closechange)]))

	print(''.join(['above_ma_stock_with_closechange_columns=',str(columns_above_ma_stock_list_with_closechange)]))	

 

def find_stock_in_trend_onsite(total_money,ma_days,end_day):  #获取全市场中，过去2天，每天都上涨的股票。

	stockslist=get_all_sec(end_day)

	# 找出连涨2天的股票。

	good_trend_stock_list=[good_stock for good_stock in stockslist if Jqdata(good_stock).judge_stock_increase(end_day)==True]  #股票代码没有变化，按平台来的。

	good_trend_stock_list_code_transformed=[Jqdata(sec).transform_code(sec) for sec in good_trend_stock_list ]  #股票代码发生变化，按腾讯来的。

	above_ma_stock_list=[goodsec for goodsec in good_trend_stock_list if Jqdata(goodsec).judge_above_ma(ma_days,end_day)==True]   #股票代码没有变化，按平台来的。

	above_ma_stock_list_code_transformed=[Jqdata(goodsec).transform_code(goodsec) for goodsec in above_ma_stock_list]  #股票代码发生变化，按腾讯来的。

	final_df=get_target_stocklist_closechange_data(above_ma_stock_list_code_transformed,total_money,end_day)  #将得到的高于5天线，连涨两天的股票的信息，列出来，作为参考。

	dataframe_ranking(final_df,'vlchg')   #以当天的成交量变化，来排序。

	final_df.reset_index(drop=True,inplace=True)   #又重新生成索引。

	return above_ma_stock_list_code_transformed,final_df    #返回最终的代码列表，和一个dataframe.

def find_stock_in_trend_onsite_with_codelist(ma_days,end_day):  #获取全市场中，过去2天，每天都上涨的股票。

	stockslist=get_all_sec(end_day)

	# 找出连涨2天的股票。

	good_trend_stock_list=[good_stock for good_stock in stockslist if Jqdata(good_stock).judge_stock_increase(end_day)==True]  #股票代码没有变化，按平台来的。

	#good_trend_stock_list_code_transformed=[Jqdata(sec).transform_code(sec) for sec in good_trend_stock_list ]  #股票代码发生变化，按腾讯来的。

	if good_trend_stock_list!=[]:

		above_ma_stock_list=[goodsec for goodsec in good_trend_stock_list if Jqdata(goodsec).judge_above_ma(ma_days,end_day)==True]   #股票代码没有变化，按平台来的。

		above_ma_stock_list_code_transformed=[Jqdata(goodsec).transform_code(goodsec) for goodsec in above_ma_stock_list]  #股票代码发生变化，按腾讯来的。

	else:

		above_ma_stock_list_code_transformed=[]

	return above_ma_stock_list_code_transformed   #返回最终的代码列表，

def find_index_stock_in_trend_onsite(index_code,total_money,ma_days,end_day):  #获取某个指数成分股中，过去2天，每天都上涨的股票。

	stockslist=Jqdata(index_code).get_index_securities()

	# 找出连涨2天的股票。

	good_trend_stock_list=[good_stock for good_stock in stockslist if Jqdata(good_stock).judge_stock_increase(end_day)==True]  #股票代码没有变化，按平台来的。

	good_trend_stock_list_code_transformed=[Jqdata(sec).transform_code(sec) for sec in good_trend_stock_list ]  #股票代码发生变化，按腾讯来的。

	above_ma_stock_list=[goodsec for goodsec in good_trend_stock_list if Jqdata(goodsec).judge_above_ma(ma_days,end_day)==True]   #股票代码没有变化，按平台来的。

	above_ma_stock_list_code_transformed=[Jqdata(goodsec).transform_code(goodsec) for goodsec in above_ma_stock_list]  #股票代码发生变化，按腾讯来的。

	final_df=get_target_stocklist_closechange_data(above_ma_stock_list_code_transformed,total_money,end_day)  #将得到的高于5天线，连涨两天的股票的信息，列出来，作为参考。

	dataframe_ranking(final_df,'vlchg')   #以当天的成交量变化，来排序。

	final_df.reset_index(drop=True,inplace=True)   #又重新生成索引。


	return above_ma_stock_list_code_transformed,final_df    #返回最终的代码列表，和一个dataframe.



def find_index_stock_in_trend_onsite_with_codelist(index_code,ma_days,end_day):  #获取某个指数成分股中，过去2天，每天都上涨的股票。


	stockslist=Jqdata(index_code).get_index_securities()

	# 找出连涨2天的股票。

	good_trend_stock_list=[good_stock for good_stock in stockslist if Jqdata(good_stock).judge_stock_increase(end_day)==True]  #股票代码没有变化，按平台来的。

	#good_trend_stock_list_code_transformed=[Jqdata(sec).transform_code(sec) for sec in good_trend_stock_list ]  #股票代码发生变化，按腾讯来的。

	#print(''.join(['stock_in_trend=',str(good_trend_stock_list_code_transformed)]))

	if good_trend_stock_list!=[]:

		above_ma_stock_list=[goodsec for goodsec in good_trend_stock_list if Jqdata(goodsec).judge_above_ma(ma_days,end_day)==True]   #股票代码没有变化，按平台来的。

		above_ma_stock_list_code_transformed=[Jqdata(goodsec).transform_code(goodsec) for goodsec in above_ma_stock_list]  #股票代码发生变化，按腾讯来的。

	else:

		above_ma_stock_list_code_transformed=[]

	return above_ma_stock_list_code_transformed    #返回最终的代码列表


def find_stock_high_low_increase(total_money,ma_days,end_day):  #找出全市场中，过去三天低点和高点每天都上涨的股票，包括今天在内三天的数据。

	stockslist=get_all_sec(end_day)

	# 找出连涨2天的股票。

	good_trend_stock_list=[good_stock for good_stock in stockslist if Jqdata(good_stock).judge_stock_trend(end_day)==True]  #股票代码没有变化，按平台来的。

	good_trend_stock_list_code_transformed=[Jqdata(sec).transform_code(sec) for sec in good_trend_stock_list ]  #股票代码发生变化，按腾讯来的。

	above_ma_stock_list=[goodsec for goodsec in good_trend_stock_list if Jqdata(goodsec).judge_above_ma(ma_days,end_day)==True]   #股票代码没有变化，按平台来的。

	above_ma_stock_list_code_transformed=[Jqdata(goodsec).transform_code(goodsec) for goodsec in above_ma_stock_list]  #股票代码发生变化，按腾讯来的。

	if above_ma_stock_list_code_transformed!=[]:


		final_df=get_target_stocklist_closechange_data(above_ma_stock_list_code_transformed,total_money,end_day)  #将得到的高于5天线，连涨两天的股票的信息，列出来，作为参考。

		dataframe_ranking(final_df,'vlchg')   #以当天的成交量变化，来排序。

		final_df.reset_index(drop=True,inplace=True)   #又重新生成索引。

	else:

		final_df=pd.DataFrame()   #返回一个空的数据框架。

	return above_ma_stock_list_code_transformed,final_df    #返回最终的代码列表，和一个dataframe.



def find_stock_high_low_increase_with_codelist(ma_days,end_day):  #找出全市场中，过去三天低点和高点每天都上涨的股票，包括今天在内三天的数据。

	stockslist=get_all_sec(end_day)

	# 找出连涨2天的股票。

	good_trend_stock_list=[good_stock for good_stock in stockslist if Jqdata(good_stock).judge_stock_trend(end_day)==True]  #股票代码没有变化，按平台来的。

	if good_trend_stock_list!=[]:

		above_ma_stock_list=[goodsec for goodsec in good_trend_stock_list if Jqdata(goodsec).judge_above_ma(ma_days,end_day)==True]   #股票代码没有变化，按平台来的。

		above_ma_stock_list_code_transformed=[Jqdata(goodsec).transform_code(goodsec) for goodsec in above_ma_stock_list]  #股票代码发生变化，按腾讯来的。

	else:

		above_ma_stock_list_code_transformed=[]



	return above_ma_stock_list_code_transformed    #返回最终的代码列表，和一个dataframe.




def find_index_stock_high_low_increase(index_code,total_money,ma_days,end_day):  #找出某个指数的成分股中，过去三天低点和高点每天都上涨的股票，包括今天在内三天的数据。

	stockslist=Jqdata(index_code).get_index_securities()

	# 找出连涨2天的股票。

	good_trend_stock_list=[good_stock for good_stock in stockslist if Jqdata(good_stock).judge_stock_trend(end_day)==True]  #股票代码没有变化，按平台来的。

	good_trend_stock_list_code_transformed=[Jqdata(sec).transform_code(sec) for sec in good_trend_stock_list ]  #股票代码发生变化，按腾讯来的。

	above_ma_stock_list=[goodsec for goodsec in good_trend_stock_list if Jqdata(goodsec).judge_above_ma(ma_days,end_day)==True]   #股票代码没有变化，按平台来的。

	above_ma_stock_list_code_transformed=[Jqdata(goodsec).transform_code(goodsec) for goodsec in above_ma_stock_list]  #股票代码发生变化，按腾讯来的。

	if above_ma_stock_list_code_transformed !=[]:

		final_df=get_target_stocklist_closechange_data(above_ma_stock_list_code_transformed,total_money,end_day)  #将得到的高于5天线，连涨两天的股票的信息，列出来，作为参考。

		dataframe_ranking(final_df,'vlchg')   #以当天的成交量变化，来排序。

		final_df.reset_index(drop=True,inplace=True)   #又重新生成索引。

	else:

		final_df=pd.DataFrame()  #返回一个空的数据框架。

	#print(''.join(['above_ma_stock_with_closechange_list=',str(above_ma_stock_list_with_closechange)]))

	#print(''.join(['above_ma_stock_with_closechange_columns=',str(columns_above_ma_stock_list_with_closechange)]))	

	return above_ma_stock_list_code_transformed,final_df    #返回最终的代码列表，和一个dataframe.

def find_index_stock_high_low_increase_with_codelist(index_code,ma_days,end_day):  #找出某个指数的成分股中，过去三天低点和高点每天都上涨的股票，包括今天在内三天的数据。

	stockslist=Jqdata(index_code).get_index_securities()

	# 找出连涨2天的股票。

	good_trend_stock_list=[good_stock for good_stock in stockslist if Jqdata(good_stock).judge_stock_trend(end_day)==True]  #股票代码没有变化，按平台来的。

	if good_trend_stock_list!=[]:
	

		above_ma_stock_list=[goodsec for goodsec in good_trend_stock_list if Jqdata(goodsec).judge_above_ma(ma_days,end_day)==True]   #股票代码没有变化，按平台来的。

		above_ma_stock_list_code_transformed=[Jqdata(goodsec).transform_code(goodsec) for goodsec in above_ma_stock_list]  #股票代码发生变化，按腾讯来的。


	else:

		above_ma_stock_list_code_transformed=[]

	return above_ma_stock_list_code_transformed   #返回最终的代码列表

def find_stock_in_trend_after_trading(total_money,ma_days,end_day):  #获取全市场中，过去2天，每天高点低点，不停上升的股票。

	stockslist=get_all_sec(end_day)

	# 找出连涨2天的股票。

	good_trend_stock_list=[good_stock for good_stock in stockslist if Jqdata(good_stock).judge_stock_increase(end_day)==True]  #股票代码没有变化，按平台来的。

	good_trend_stock_list_code_transformed=[Jqdata(sec).transform_code(sec) for sec in good_trend_stock_list ]  #股票代码发生变化，按腾讯来的。

	good_trend_stock_df=get_target_stocklist_closechange_data(good_trend_stock_list_code_transformed,total_money,end_day)



	print(''.join(['stock_intrend=',str(good_trend_stock_list_code_transformed)]))

	print(''.join(['goodstock_date=',str(end_day)]))


	good_trend_stock_list_with_closechange,columns_good_trend_stock_list_with_closechange=dataframe2list(good_trend_stock_df)

	print(''.join(['good_trend_stock_list=',str(good_trend_stock_list_with_closechange)]))

	print(''.join(['good_trend_stock_columns=',str(columns_good_trend_stock_list_with_closechange)]))

	#获取3天连涨的股票列表中的，超越5天线的股票。

	#list_name=''.join(['above_ma_stock_list_',str(ma_days)])

	above_ma_stock_list=[goodsec for goodsec in good_trend_stock_list if Jqdata(goodsec).judge_above_ma(ma_days,end_day)==True]   #股票代码没有变化，按平台来的。

	above_ma_stock_list_code_transformed=[Jqdata(goodsec).transform_code(goodsec) for goodsec in above_ma_stock_list]  #股票代码发生变化，按腾讯来的。

	print(''.join(['above_ma_stock_list','=',str(above_ma_stock_list_code_transformed)]))

	final_df=get_target_stocklist_closechange_data(above_ma_stock_list_code_transformed,total_money,end_day)  #将得到的高于5天线，连涨两天的股票的信息，列出来，作为参考。

	above_ma_stock_list_with_closechange,columns_above_ma_stock_list_with_closechange=dataframe2list(final_df)

	#dataframe_ranking(final_df,'vlchg')   #以当天的成交量变化，来排序。

	#final_df.reset_index(drop=True,inplace=True)   #又重新生成索引。

	print(''.join(['above_ma_stock_list=',str(above_ma_stock_list_with_closechange)]))

	print(''.join(['above_ma_stock_columns=',str(columns_above_ma_stock_list_with_closechange)]))	

	#return above_ma_stock_list_code_transformed,final_df    #返回最终的代码列表，和一个dataframe.

	

def find_stock_in_bad_trend(ma_days,end_day):  #获取全市场中，过去3天，每天高点低点，不停上升的股票。

	stockslist=get_all_sec(end_day)

	# 找出连跌2天的股票。

	bad_trend_stock_list=[bad_stock for bad_stock in stockslist if Jqdata(bad_stock).judge_stock_down(end_day)==True]  #股票代码没有变化，按平台来的。

	bad_trend_stock_list_code_transformed=[Jqdata(sec).transform_code(sec) for sec in bad_trend_stock_list ]  #股票代码发生变化，按腾讯来的。

	print(''.join(['bad_stock_in_trend=',str(bad_trend_stock_list_code_transformed)]))

	print(''.join(['bad_stock_date=',str(end_day)]))

	#以下是将上面的代码，代码转换，获取股票名称和3日内的涨幅，排序，然后做成列表和列的数据，打印出来。改成了两日内的涨幅，因为选择实际也是选择连续两天上涨的股票，那么就是计算这两天上涨的涨幅，而不应该是计算三天。

	bad_trend_stock_list_with_closechange=[[Jqdata(bad_stock).transform_code(bad_stock),Jqdata(bad_stock).security_name(),Jqdata(bad_stock).get_close_change(end_day,2)] for bad_stock in bad_trend_stock_list]
	
	columns_bad_trend_stock_list_with_closechange=['code','name','closechange']
	
	bad_trend_stock_df=list2dataframe(bad_trend_stock_list_with_closechange,columns_bad_trend_stock_list_with_closechange)

	bad_trend_stock_df.drop(bad_trend_stock_df[pd.isna(bad_trend_stock_df['closechange'])].index,inplace=True)    #在本dataframe closechange 列中，分数为nan(这是一个非字符的float类型的数据，用pd.isna()方法判断这是一个nan数据类型的，确定后就去掉)
	
	dataframe_ranking(bad_trend_stock_df,'closechange')  #排序

	#print(len(bad_trend_stock_df))

	bad_trend_stock_list_with_closechange,columns_bad_trend_stock_list_with_closechange=dataframe2list(bad_trend_stock_df)

	print(''.join(['bad_trend_stock_with_closechange_list=',str(bad_trend_stock_list_with_closechange)]))

	print(''.join(['bad_trend_stock_with_closechange_columns=',str(columns_bad_trend_stock_list_with_closechange)]))

	#获取3天下跌的股票列表中的，低于5天线的股票。

	list_name=''.join(['below_ma_stock_list_',str(ma_days)])

	below_ma_stock_list=[badsec for badsec in bad_trend_stock_list if Jqdata(badsec).judge_above_ma(ma_days,end_day)==False]  #股票代码没有变化，按平台来的。

	below_ma_stock_list_code_transformed=[Jqdata(badsec).transform_code(badsec) for badsec in below_ma_stock_list]  #股票代码发生变化，按腾讯来的。

	print(''.join([list_name,'=',str(below_ma_stock_list_code_transformed)]))

	#以下是将上面超越5天线的的代码，代码转换，获取股票名称和3日内的涨幅，排序，然后做成列表和列的数据，打印出来。改成了两日内的涨幅，因为选择实际也是选择连续两天上涨的股票，那么就是计算这两天上涨的涨幅，而不应该是计算三天。

	below_ma_stock_list_with_closechange=[[Jqdata(bad_stock).transform_code(bad_stock),Jqdata(bad_stock).security_name(),Jqdata(bad_stock).get_close_change(end_day,2)] for bad_stock in below_ma_stock_list]
	
	columns_below_ma_stock_list_with_closechange=['code','name','closechange']
	
	below_ma_stock_df=list2dataframe(below_ma_stock_list_with_closechange,columns_below_ma_stock_list_with_closechange)

	below_ma_stock_df.drop(below_ma_stock_df[pd.isna(below_ma_stock_df['closechange'])].index,inplace=True)    #在本dataframe 中'closechange'列中，分数为nan(这是一个非字符的float类型的数据，用pd.isna()方法判断这是一个nan数据类型的，确定后就去掉)
	
	dataframe_ranking(below_ma_stock_df,'closechange')  #排序

	#print(len(below_ma_stock_df))

	below_ma_stock_list_with_closechange,columns_below_ma_stock_list_with_closechange=dataframe2list(below_ma_stock_df)

	print(''.join(['below_ma_stock_with_closechange_list=',str(below_ma_stock_list_with_closechange)]))

	print(''.join(['below_ma_stock_with_closechange_columns=',str(columns_below_ma_stock_list_with_closechange)]))	


def find_stock_in_bad_trend_onsite(end_day):  #获取全市场中，过去3天，每天高点低点，不停上升的股票。在平台，实时得观察连续跌得股票。

	stockslist=get_all_sec(end_day)

	# 找出连跌2天的股票。

	bad_trend_stock_list=[bad_stock for bad_stock in stockslist if Jqdata(bad_stock).judge_stock_down(end_day)==True]  #股票代码没有变化，按平台来的。

	bad_trend_stock_list_code_transformed=[Jqdata(sec).transform_code(sec) for sec in bad_trend_stock_list ]  #股票代码发生变化，按腾讯来的。


	return bad_trend_stock_list_code_transformed     #返回列表，告知当前，包括今天连续跌两天的股票名单，代码为腾讯标准代码。


def find_index_in_trend_with_stock(ma_days,end_day):  #获取全市场中，过去3天，连续两天都上涨的指数下面的股票数据信息。

	#以下这行不需要了，以后按照实时的数据，获取所有的指数列表。

	#index_list=['sh000688', 'sh000682','sz399965','sh000685','sh000852', 'sz399905', 'sh000016', 'sh000097', 'sz399006', 'sz399004', 'sz399005', 'sz399300', 'sz399330', 'sz399333', 'sz399612', 'sz399673', 'sz399903', 'sz399324', 'sh000814', 'sh000819', 'sh000827', 'sh000901', 'sh000928', 'sz399932', 'sh000933', 'sh000934', 'sh000935', 'sh000986', 'sh000987', 'sh000989', 'sh000991', 'sh000992', 'sh000993', 'sz399030', 'sz399060', 'sz399295', 'sz399296', 'sz399393', 'sz399394', 'sz399395', 'sz399396', 'sz399412', 'sz399417', 'sz399419', 'sz399420', 'sz399423', 'sz399432', 'sz399435', 'sz399438', 'sz399439', 'sz399440', 'sz399441', 'sz399608', 'sz399610', 'sz399646', 'sz399647', 'sz399653', 'sz399654', 'sz399669', 'sz399674', 'sz399675', 'sz399676', 'sz399677', 'sz399683', 'sz399684', 'sz399687', 'sz399693', 'sz399695', 'sz399699', 'sz399704', 'sz399705', 'sz399706', 'sz399804', 'sz399805', 'sz399806', 'sz399809', 'sz399807', 'sz399808', 'sz399810', 'sz399811', 'sz399812', 'sz399813', 'sz399814', 'sz399967', 'sz399970', 'sz399971', 'sz399973', 'sz399975', 'sz399976', 'sz399986', 'sz399987', 'sz399989', 'sz399993', 'sz399994', 'sz399995', 'sz399996', 'sz399997', 'sz399998', 'sz399966']

	# 找出连涨2天的指数。

	index_list=filter_allindex(end_day)

	good_trend_index_list=[good_index for good_index in index_list if Jqdata(good_index).judge_stock_increase(end_day)==True]  #股票代码没有变化，按平台来的。

	#good_trend_index_list_code_transformed=[Jqdata(index).transform_code(index) for index in good_trend_index_list ]  #股票代码发生变化，按腾讯来的。

	print(''.join(['index_in_trend=',str(good_trend_index_list)]))

	print(''.join(['good_index_date=',str(end_day)]))

	#以下是将上面的代码，代码转换，获取股票名称和2日内的涨幅，排序，然后做成列表和列的数据，打印出来。 改成了两日内的涨幅，因为选择实际也是选择连续两天上涨的股票，那么就是计算这两天上涨的涨幅，而不应该是计算三天。

	good_trend_index_list_with_closechange=[[good_index,Jqdata(good_index).security_name(),Jqdata(good_index).get_close_change(end_day,2)] for good_index in good_trend_index_list]
	
	columns_good_trend_index_list_with_closechange=['code','name','closechange']
	
	good_trend_index_df=list2dataframe(good_trend_index_list_with_closechange,columns_good_trend_index_list_with_closechange)

	good_trend_index_df.drop(good_trend_index_df[pd.isna(good_trend_index_df['closechange'])].index,inplace=True)    #在本dataframe closechange 列中，分数为nan(这是一个非字符的float类型的数据，用pd.isna()方法判断这是一个nan数据类型的，确定后就去掉)
	
	dataframe_ranking(good_trend_index_df,'closechange')  #排序

	#print(len(good_trend_index_df))

	good_trend_index_list_with_closechange,columns_good_trend_index_list_with_closechange=dataframe2list(good_trend_index_df)

	print(''.join(['good_trend_index_with_closechange_list=',str(good_trend_index_list_with_closechange)]))

	print(''.join(['good_trend_index_with_closechange_columns=',str(columns_good_trend_index_list_with_closechange)]))


	if good_trend_index_list!=[]:


		#如下将每个指数下面的连涨两天的股票找出来。

		limit_number=len(good_trend_index_list)

		i=0

		index_stock_dic={}

		while i < limit_number:

			stockslist=Jqdata(good_trend_index_list[i]).get_index_securities()

			good_trend_stock_list=[good_stock for good_stock in stockslist if Jqdata(good_stock).judge_stock_increase(end_day)==True]  #股票代码没有变化，按平台来的。

			good_trend_stock_list_code_transformed=[Jqdata(sec).transform_code(sec) for sec in good_trend_stock_list ]  #股票代码发生变化，按腾讯来的。


			#以下是将上面的代码，代码转换，获取股票名称和2日内的涨幅，排序，然后做成列表和列的数据，打印出来。 改成了两日内的涨幅，因为选择实际也是选择连续两天上涨的股票，那么就是计算这两天上涨的涨幅，而不应该是计算三天。

			good_trend_stock_list_with_closechange=[[Jqdata(good_stock).transform_code(good_stock),Jqdata(good_stock).security_name(),Jqdata(good_stock).get_close_change(end_day,2)] for good_stock in good_trend_stock_list]
			
			columns_good_trend_stock_list_with_closechange=['code','name','closechange']
			
			good_trend_stock_df=list2dataframe(good_trend_stock_list_with_closechange,columns_good_trend_stock_list_with_closechange)

			good_trend_stock_df.drop(good_trend_stock_df[pd.isna(good_trend_stock_df['closechange'])].index,inplace=True)    #在本dataframe closechange 列中，分数为nan(这是一个非字符的float类型的数据，用pd.isna()方法判断这是一个nan数据类型的，确定后就去掉)
			
			dataframe_ranking(good_trend_stock_df,'closechange')  #排序

			#print(len(good_trend_stock_df))

			good_trend_stock_list_with_closechange,columns_good_trend_stock_list_with_closechange=dataframe2list(good_trend_stock_df)

			#print(''.join(['good_trend_stock_with_closechange_list=',str(good_trend_stock_list_with_closechange)]))

			#print(''.join(['good_trend_stock_with_closechange_columns=',str(columns_good_trend_stock_list_with_closechange)]))

			index_stock_dic.setdefault(good_trend_index_list[i],good_trend_stock_list_with_closechange)

			i +=1



		print(''.join(['good_index_stock_dic=',str(index_stock_dic)]))

		print(''.join(['good_index_stock_dic_columns=',str(columns_good_trend_stock_list_with_closechange)]))

	else:

		print('没有表现好的指数，程序结束！')


def find_target_in_trend(target_list,end_day):  #获取全市场中，过去3天，连续两天都上涨的目标。

	# 找出连涨2天的指数。

	good_trend_target_list=[good_target for good_target in target_list if Jqdata(good_target).judge_stock_increase(end_day)==True]  #股票代码没有变化，按平台来的。

	if good_trend_target_list!=[]:

		print('过去三天每天收盘价都上涨的的指数。如下：')

		good_trend_target_list_with_closechange=[[good_target,Jqdata(good_target).security_name(),Jqdata(good_target).get_close_change(end_day,2)] for good_target in good_trend_target_list]

		columns_good_trend_target_list_with_closechange=['code','name','closechange']
		
		good_trend_target_df=list2dataframe(good_trend_target_list_with_closechange,columns_good_trend_target_list_with_closechange)

		good_trend_target_df.drop(good_trend_target_df[pd.isna(good_trend_target_df['closechange'])].index,inplace=True)    #在本dataframe closechange 列中，分数为nan(这是一个非字符的float类型的数据，用pd.isna()方法判断这是一个nan数据类型的，确定后就去掉)
		
		dataframe_ranking(good_trend_target_df,'closechange')  #排序

		#good_trend_target_df.reset_target(inplace=True)

		return good_trend_target_df

	else:

		print('没有表现好的指数，程序结束！')

		good_trend_target_df=pd.DataFrame()

		#return None    不能这么写，还是要返回 good_trend_target_df, 因为以后的程序要判断这个dataframe是否是空，用None不好与一个可能是dataframe的数据比较，会出现错误。TypeError: Could not compare [None] with block values

		return good_trend_target_df



def find_index_in_trend(end_day):#获取全市场中，过去3天，连续两天都上涨的指数.

	target_list=filter_allindex(end_day)

	good_trend_target_df=find_target_in_trend(target_list,end_day)

	good_trend_target_df['code']=good_trend_target_df['code'].apply(transform_to_normal_code)

	good_trend_target_df['name']=good_trend_target_df['name'].apply(remove_string_from_index)

	return good_trend_target_df



def find_index_high_low_increase(end_day):  #获取全市场中，过去3天，每天高点低点，不停上升的指数。

	#index_list=['sh000688','sh000682','sh000685','sh000852', 'sz399905', 'sh000016', 'sh000097', 'sz399006', 'sz399004', 'sz399005', 'sz399300', 'sz399330', 'sz399333', 'sz399612', 'sz399673', 'sz399903', 'sz399324', 'sh000814', 'sh000819', 'sh000827', 'sh000901', 'sh000928', 'sz399932', 'sh000933', 'sh000934', 'sh000935', 'sh000986', 'sh000987', 'sh000989', 'sh000991', 'sh000992', 'sh000993', 'sz399030', 'sz399060', 'sz399295', 'sz399296', 'sz399393', 'sz399394', 'sz399395', 'sz399396', 'sz399412', 'sz399417', 'sz399419', 'sz399420', 'sz399423', 'sz399432', 'sz399435', 'sz399438', 'sz399439', 'sz399440', 'sz399441', 'sz399608', 'sz399610', 'sz399646', 'sz399647', 'sz399653', 'sz399654', 'sz399669', 'sz399674', 'sz399675', 'sz399676', 'sz399677', 'sz399683', 'sz399684', 'sz399687', 'sz399693', 'sz399695', 'sz399699', 'sz399704', 'sz399705', 'sz399706', 'sz399804', 'sz399805', 'sz399806', 'sz399809', 'sz399807', 'sz399808', 'sz399810', 'sz399811', 'sz399812', 'sz399813', 'sz399814', 'sz399967', 'sz399970', 'sz399971', 'sz399973', 'sz399975', 'sz399976', 'sz399986', 'sz399987', 'sz399989', 'sz399993', 'sz399994', 'sz399995', 'sz399996', 'sz399997', 'sz399998', 'sz399966']


	# 找出连涨2天的指数。

	index_list=filter_allindex(end_day)

	good_trend_index_list=[good_index for good_index in index_list if Jqdata(good_index).judge_stock_trend(end_day)==True]  #股票代码没有变化，按平台来的。


	if good_trend_index_list!=[]:


		print('过去三天每天低点和高点都上升的指数。如下：')

		good_trend_index_list_with_closechange=[[good_index,Jqdata(good_index).security_name(),Jqdata(good_index).get_close_change(end_day,2)] for good_index in good_trend_index_list]
		
		columns_good_trend_index_list_with_closechange=['code','name','closechange']
		
		good_trend_index_df=list2dataframe(good_trend_index_list_with_closechange,columns_good_trend_index_list_with_closechange)

		good_trend_index_df.drop(good_trend_index_df[pd.isna(good_trend_index_df['closechange'])].index,inplace=True)    #在本dataframe closechange 列中，分数为nan(这是一个非字符的float类型的数据，用pd.isna()方法判断这是一个nan数据类型的，确定后就去掉)
		
		dataframe_ranking(good_trend_index_df,'closechange')  #排序

		good_trend_index_df.reset_index(inplace=True)

		return good_trend_index_df

	else:

		print('没有表现好的指数，程序结束！')

		good_trend_index_df=pd.DataFrame()

		#return None    不能这么写，还是要返回 good_trend_index_df, 因为以后的程序要判断这个dataframe是否是空，用None不好与一个可能是dataframe的数据比较，会出现错误。TypeError: Could not compare [None] with block values

		return good_trend_index_df


def remove_string_from_index(index_name):  #主要是为了从panda里面将指数名称里面带指数的两个字去掉。

	index_name_string=index_name

	index_name_string=index_name_string.replace('指数','') 

	return index_name_string


def make_column_int(column_value):  #将列的数据变成整数。

	column_value=int(column_value)

	return column_value




def find_good_index_sec(ma_days,end_day):  #连续两天上涨的指数中，连续两天上涨的股票信息。

	#index_list=['sh000688','sh000682','sh000685','sh000852', 'sz399905', 'sh000016', 'sh000097', 'sz399006', 'sz399004', 'sz399005', 'sz399300', 'sz399330', 'sz399333', 'sz399612', 'sz399673', 'sz399903', 'sz399324', 'sh000814', 'sh000819', 'sh000827', 'sh000901', 'sh000928', 'sz399932', 'sh000933', 'sh000934', 'sh000935', 'sh000986', 'sh000987', 'sh000989', 'sh000991', 'sh000992', 'sh000993', 'sz399030', 'sz399060', 'sz399295', 'sz399296', 'sz399393', 'sz399394', 'sz399395', 'sz399396', 'sz399412', 'sz399417', 'sz399419', 'sz399420', 'sz399423', 'sz399432', 'sz399435', 'sz399438', 'sz399439', 'sz399440', 'sz399441', 'sz399608', 'sz399610', 'sz399646', 'sz399647', 'sz399653', 'sz399654', 'sz399669', 'sz399674', 'sz399675', 'sz399676', 'sz399677', 'sz399683', 'sz399684', 'sz399687', 'sz399693', 'sz399695', 'sz399699', 'sz399704', 'sz399705', 'sz399706', 'sz399804', 'sz399805', 'sz399806', 'sz399809', 'sz399807', 'sz399808', 'sz399810', 'sz399811', 'sz399812', 'sz399813', 'sz399814', 'sz399967', 'sz399970', 'sz399971', 'sz399973', 'sz399975', 'sz399976', 'sz399986', 'sz399987', 'sz399989', 'sz399993', 'sz399994', 'sz399995', 'sz399996', 'sz399997', 'sz399998', 'sz399966']

	# 找出连涨2天的指数。

	index_list=filter_allindex(end_day)

	good_trend_index_list=[good_index for good_index in index_list if Jqdata(good_index).judge_stock_increase(end_day)==True]  #股票代码没有变化，按平台来的。

	good_trend_index_list_with_closechange=[[good_index,Jqdata(good_index).security_name(),Jqdata(good_index).get_close_change(end_day,2)] for good_index in good_trend_index_list] #可以将代码不使用腾讯的
	
	columns_good_trend_index_list_with_closechange=['code','name','cl_2']
	
	good_trend_index_df=list2dataframe(good_trend_index_list_with_closechange,columns_good_trend_index_list_with_closechange)

	good_trend_index_df.drop(good_trend_index_df[pd.isna(good_trend_index_df['cl_2'])].index,inplace=True)    #在本dataframe closechange 列中，分数为nan(这是一个非字符的float类型的数据，用pd.isna()方法判断这是一个nan数据类型的，确定后就去掉)
	
	dataframe_ranking(good_trend_index_df,'cl_2')  #排序


	#如下将每个指数下面的连涨两天的股票找出来。

	limit_number=len(good_trend_index_list)

	i=0

	index_stock_df_dic={}

	while i < limit_number:

		stockslist=Jqdata(good_trend_index_list[i]).get_index_securities()

		good_trend_stock_list=[good_stock for good_stock in stockslist if Jqdata(good_stock).judge_stock_increase(end_day)==True]  #股票代码没有变化，按平台来的。

		good_trend_stock_list_code_transformed=[Jqdata(sec).transform_code(sec) for sec in good_trend_stock_list ]  #股票代码发生变化，按腾讯来的。


		#以下是将上面的代码，代码转换，获取股票名称和2日内的涨幅，排序，然后做成列表和列的数据，打印出来。 改成了两日内的涨幅，因为选择实际也是选择连续两天上涨的股票，那么就是计算这两天上涨的涨幅，而不应该是计算三天。

		good_trend_stock_list_with_closechange=[[Jqdata(good_stock).transform_code(good_stock),Jqdata(good_stock).security_name(),Jqdata(good_stock).get_close_change(end_day,2)] for good_stock in good_trend_stock_list]
		
		columns_good_trend_stock_list_with_closechange=['code','name','cl_2']
		
		good_trend_stock_df=list2dataframe(good_trend_stock_list_with_closechange,columns_good_trend_stock_list_with_closechange)

		good_trend_stock_df.drop(good_trend_stock_df[pd.isna(good_trend_stock_df['cl_2'])].index,inplace=True)    #在本dataframe closechange 列中，分数为nan(这是一个非字符的float类型的数据，用pd.isna()方法判断这是一个nan数据类型的，确定后就去掉)
		
		dataframe_ranking(good_trend_stock_df,'cl_2')  #排序

		#print(len(good_trend_stock_df))

		good_trend_stock_list_with_closechange,columns_good_trend_stock_list_with_closechange=dataframe2list(good_trend_stock_df)

		#print(''.join(['good_trend_stock_with_closechange_list=',str(good_trend_stock_list_with_closechange)]))

		#print(''.join(['good_trend_stock_with_closechange_columns=',str(columns_good_trend_stock_list_with_closechange)]))

		index_stock_df_dic.setdefault(good_trend_index_list[i],good_trend_stock_df)

		i +=1



	#print(''.join(['good_index_stock_dic=',str(index_stock_dic)]))

	#print(''.join(['good_index_stock_dic_columns=',str(columns_good_trend_stock_list_with_closechange)]))


	return good_trend_index_df,index_stock_df_dic




def get_allsec_closechange_data_after_trading(total_money,end_day):  #这个函数包含了了股票市场和流通市值的数据。使用这个函数，必须是盘后，所以盘中是不适合使用的。

	stockslist=get_all_sec(end_day)

	#获取当天，2天，5天，10天，20天，30天，60天，和今年到现在为止，该股票的涨幅数据。

	df_list=[
	[Jqdata(stock).transform_code(stock),Jqdata(stock).security_name(),
	Jqdata(stock).security_industry(end_day),
	 Jqdata(stock).buy_single_stock_volume(total_money,end_day),
	round(Jqdata(stock).get_close_change(end_day,1),2),
	round(Jqdata(stock).get_close_change(end_day,2),2),    
	round(Jqdata(stock).get_close_change(end_day,5),2),
	round(Jqdata(stock).get_close_change(end_day,10),2),
	round(Jqdata(stock).get_close_change(end_day,20),2),
	round(Jqdata(stock).get_close_change(end_day,30),2),
	round(Jqdata(stock).get_close_change(end_day,60),2),
	round(Jqdata(stock).get_close_change_YTD(target_date()),2),
	Jqdata(stock).get_sec_market_size(end_day)[0],   #如下这两行是为了市值和流动市值准备的，但系统老出错，没有办法，放弃吧。
	Jqdata(stock).get_sec_market_size(end_day)[1],
	Jqdata(stock).get_volume_change(end_day),
	Jqdata(stock).get_volume_ranking(end_day),

	end_day,


	] 
	for stock in stockslist]

	df_columns=['code','name','industry','b_vl','cl_1','cl_2','cl_5','cl_10','cl_20','cl_30','cl_60','YTD','shizhi','ltzhi','vlchg','vlrk','tm',]

	#df_columns=['code','name','cl_1','cl_2','cl_5','cl_10','cl_20','cl_30','cl_60','YTD','Data_time',]


	df=list2dataframe(df_list,df_columns)

	#去掉数据为i 空的那些行


	df.drop(df[pd.isna(df['cl_1'])].index,inplace=True)
	df.drop(df[pd.isna(df['cl_2'])].index,inplace=True)
	df.drop(df[pd.isna(df['cl_5'])].index,inplace=True)
	df.drop(df[pd.isna(df['cl_10'])].index,inplace=True)
	df.drop(df[pd.isna(df['cl_20'])].index,inplace=True)
	df.drop(df[pd.isna(df['cl_30'])].index,inplace=True)
	df.drop(df[pd.isna(df['cl_60'])].index,inplace=True)
	df.drop(df[pd.isna(df['YTD'])].index,inplace=True)

	df_list,df_columns=dataframe2list(df)


	#print(df)

	print(''.join(['allsec_closechange_list=',str(df_list)]))

	print(''.join(['allsec_closechange_columns=',str(df_columns)]))



def get_single_index_allsec_closechange_data(target_index,total_money,end_day):  #这个函数用来查询某个指数下面所有股票的短期，长期表现，可以快速的查询出来他们的表现。

	stockslist=Jqdata(target_index).get_index_securities()

	#获取当天，2天，5天，10天，20天，30天，60天，和今年到现在为止，该股票的涨幅数据。

	df_list=[
	[Jqdata(stock).transform_code(stock),Jqdata(stock).security_name(),
	 Jqdata(stock).buy_single_stock_volume(total_money,end_day),
	round(Jqdata(stock).get_close_change(end_day,1),2),
	round(Jqdata(stock).get_close_change(end_day,2),2),    
	round(Jqdata(stock).get_close_change(end_day,5),2),
	round(Jqdata(stock).get_close_change(end_day,10),2),
	round(Jqdata(stock).get_close_change(end_day,20),2),
	round(Jqdata(stock).get_close_change(end_day,30),2),
	round(Jqdata(stock).get_close_change(end_day,60),2),
	round(Jqdata(stock).get_close_change_YTD(target_date()),2),
	#Jqdata(stock).get_sec_market_size(end_day)[0],   #如下这两行是为了市值和流动市值准备的，但系统老出错，没有办法，放弃吧。
	#Jqdata(stock).get_sec_market_size(end_day)[1],
	end_day,


	] 
	for stock in stockslist]

	df_columns=['code','name','buy_volume','cl_1','cl_2','cl_5','cl_10','cl_20','cl_30','cl_60','YTD','Data_time',]

	#df_columns=['code','name','cl_1','cl_2','cl_5','cl_10','cl_20','cl_30','cl_60','YTD','Data_time',]


	df=list2dataframe(df_list,df_columns)

	#去掉数据为i 空的那些行


	df.drop(df[pd.isna(df['cl_1'])].index,inplace=True)
	df.drop(df[pd.isna(df['cl_2'])].index,inplace=True)
	df.drop(df[pd.isna(df['cl_5'])].index,inplace=True)
	df.drop(df[pd.isna(df['cl_10'])].index,inplace=True)
	df.drop(df[pd.isna(df['cl_20'])].index,inplace=True)
	df.drop(df[pd.isna(df['cl_30'])].index,inplace=True)
	df.drop(df[pd.isna(df['cl_60'])].index,inplace=True)
	df.drop(df[pd.isna(df['YTD'])].index,inplace=True)

	df_list,df_columns=dataframe2list(df)


	#print(df)

	#print(''.join(['allsec_closechange_list=',str(df_list)]))

	#print(''.join(['allsec_closechange_columns=',str(df_columns)]))

	return df


def get_allsec_today_closechange(end_day):

	stockslist=get_all_sec(end_day)

	#获取当天，3天，5天，10天，20天，30天，60天，和今年到现在为止，该股票的涨幅数据。

	df_list=[
	[Jqdata(stock).transform_code(stock),Jqdata(stock).security_name(),
	Jqdata(stock).get_close_change(end_day,1),
	Jqdata(stock).get_close_change_YTD(target_date()),
	end_day,
	] 
	for stock in stockslist]

	df_columns=['code','name','cl_1','YTD','Data_time']

	df=list2dataframe(df_list,df_columns)

	#去掉数据为i 空的那些行


	df.drop(df[pd.isna(df['cl_1'])].index,inplace=True)

	df.drop(df[pd.isna(df['YTD'])].index,inplace=True)

	df_list,df_columns=dataframe2list(df)

	new_df=dataframe_ranking(df,'cl_1')

	new_df.reset_index(drop=True,inplace=True)

	allsec_number=len(df)

	new_df.loc[allsec_number,'code']='平均值'

	new_df.loc[allsec_number,'cl_1']=new_df['cl_1'].mean()

	new_df.loc[allsec_number,'YTD']=new_df['YTD'].mean()


	return df

	#print(''.join(['allsec_closechange_list=',str(df_list)]))

	#print(''.join(['allsec_closechange_columns=',str(df_columns)]))






def get_stocklist_today_closechange(stockslist,end_day):  #监测自己关注的股票的即时涨跌幅信息，节省时间。

	#stockslist=get_all_sec(end_day)

	#获取当天，3天，5天，10天，20天，30天，60天，和今年到现在为止，该股票的涨幅数据。

	df_list=[
	[stock,Jqdata(stock).security_name(),
	Jqdata(stock).get_close_change(end_day,1),
	Jqdata(stock).get_close_change_YTD(target_date()),
	Jqdata(stock).get_volume_change(end_day),
	Jqdata(stock).get_volume_ranking(end_day),
	end_day,
	] 
	for stock in stockslist]

	df_columns=['code','name','cl_1','YTD','vlchg','vlrk','Data_time']

	df=list2dataframe(df_list,df_columns)

	#去掉数据为i 空的那些行


	df.drop(df[pd.isna(df['cl_1'])].index,inplace=True)

	df.drop(df[pd.isna(df['YTD'])].index,inplace=True)

	df_list,df_columns=dataframe2list(df)

	dataframe_ranking(df,'cl_1')


	#print(df)

	return df

	#print(''.join(['allsec_closechange_list=',str(df_list)]))

	#print(''.join(['allsec_closechange_columns=',str(df_columns)]))



def daily_check_momentum_score_all_sec(total_money,end_date):

	stockslist=get_all_sec(end_date)

	stock_info_list=[]	

	limit_number=len(stockslist)

	i=0

	while i < limit_number:

		each_stock_info=[]

		stock_code=stockslist[i]

		my_jq=Jqdata(stock_code)

		score,score_days=my_jq.momentum_score(end_date)

		data_date=end_date

		#sec_name=my_jq.security_name()

		#stock_code=my_jq.transform_code(stock_code)

		stock_code,stock_name,stock_industry,ATR_value,ATR_days,risk_value,present_point,buying_stock_volume,buying_money=my_jq.buy_stocks_volume(total_money,end_date)


		stock_code=my_jq.transform_code(stock_code)  #这就是将聚宽认定的代码后四位去掉，只留下大家普遍使用的代码的格式。sz399905 sh000016,以便于后续的工作


		pct_money=(round(buying_money/total_money,4))*100
		

		close_change=my_jq.get_close_change(data_date,1)

		money_change=my_jq.get_money_change(data_date)

		volume_change=my_jq.get_volume_change(data_date)

		each_stock_info=[stock_code,
		                 stock_name,
		                 stock_industry,
		                 score,
		                 ATR_value,
		                 buying_stock_volume,
		                 present_point,
		                 buying_money,
		                 pct_money,
		                 close_change,
		                 volume_change,
		                 money_change,
		                 data_date,
		                 risk_value,
		                 (ATR_days-1),
		                 score_days,
		                ]

		stock_info_list.append(each_stock_info)

		i +=1

	stock_info_columns=['Code','Name','Industry','Score','ATR_value','buy_volume',
	            'price_now','buy_value','weight','closechange','volumechange','moneychange',
	            'Data_time','risk','ATR_d','Score_d',]

	#stock_info_columns=['Code', 'Name', 'Score', 'Score_d', 'Data_time']

	df=list2dataframe(stock_info_list,stock_info_columns)

	df.drop(df[pd.isna(df['Score'])].index,inplace=True)    #在本dataframe 中'Score'列中，分数为nan(这是一个非字符的float类型的数据，用pd.isna()方法判断这是一个nan数据类型的，确定后就去掉)

	df.drop(df[pd.isna(df['price_now'])].index,inplace=True) #没有现价数据 的列 也去掉

	df.sort_values(by=['Score'],ascending=False,inplace=True)

	stock_info_list,stock_info_columns=dataframe2list(df)

	print(''.join(['df_list_all_sec_score='+str(stock_info_list)]))

	print(''.join(['df_columns_score='+str(stock_info_columns)]))


	return stock_info_list,stock_info_columns


def daily_check_allsec_nofilter(total_money,end_date):  #跟上面的类似，但更多的是体现return 的功能

	stockslist=get_all_sec(end_date)

	stock_info_list=[]	

	limit_number=len(stockslist)

	i=0

	while i < limit_number:

		each_stock_info=[]

		stock_code=stockslist[i]

		my_jq=Jqdata(stock_code)

		score,score_days=my_jq.momentum_score(end_date)

		data_date=end_date

		#sec_name=my_jq.security_name()

		#stock_code=my_jq.transform_code(stock_code)

		stock_code,stock_name,stock_industry,ATR_value,ATR_days,risk_value,present_point,buying_stock_volume,buying_money=my_jq.buy_stocks_volume(total_money,end_date)


		stock_code=my_jq.transform_code(stock_code)  #这就是将聚宽认定的代码后四位去掉，只留下大家普遍使用的代码的格式。sz399905 sh000016,以便于后续的工作


		pct_money=(round(buying_money/total_money,4))*100
		

		close_change=my_jq.get_close_change(data_date,1)

		money_change=my_jq.get_money_change(data_date)

		volume_change=my_jq.get_volume_change(data_date)

		each_stock_info=[stock_code,
		                 stock_name,
		                 stock_industry,
		                 score,
		                 ATR_value,
		                 buying_stock_volume,
		                 present_point,
		                 buying_money,
		                 pct_money,
		                 close_change,
		                 volume_change,
		                 money_change,
		                 data_date,
		                 risk_value,
		                 (ATR_days-1),
		                 score_days,
		                ]

		stock_info_list.append(each_stock_info)

		i +=1

	stock_info_columns=['Code','Name','Industry','Score','ATR_value','buy_volume',
	            'price_now','buy_value','weight','closechange','volumechange','moneychange',
	            'Data_time','risk','ATR_d','Score_d',]

	#stock_info_columns=['Code', 'Name', 'Score', 'Score_d', 'Data_time']

	df=list2dataframe(stock_info_list,stock_info_columns)

	df.drop(df[pd.isna(df['Score'])].index,inplace=True)    #在本dataframe 中'Score'列中，分数为nan(这是一个非字符的float类型的数据，用pd.isna()方法判断这是一个nan数据类型的，确定后就去掉)

	df.drop(df[pd.isna(df['price_now'])].index,inplace=True) #没有现价数据 的列 也去掉

	df.sort_values(by=['Score'],ascending=False,inplace=True)

	df.reset_index(drop=True,inplace=True)

	#df['ranking']=df.index+1

	new_df=df.drop(df.iloc[:,12:-1],axis=1) #创建一个新的dataframe

	return new_df





def check_stock_indicator_stocklist_from_df(source_df,total_money,end_day):

	stocklist=get_codelist_from_df(source_df)

	resultdf=Jq_codelist(stocklist).check_stock_indicator_stocklist(total_money,end_day)

	return resultdf



def check_stock_indicator_single_index(target_index,total_money,end_day): #显示某个指数下面所有的股票，当日的相关指标，看看是否有什么问题！比如涨停打开，冲高回落，是否要卖出！

	stocklist=Jqdata(target_index).get_index_securities()

	df=Jq_codelist(stocklist).check_stock_indicator_stocklist(total_money,end_day)

	return df

def check_stock_indicator_allstock(total_money,end_day):

	stockslist=get_all_sec(end_day)

	stockslist=[Jqdata(stock).transform_code(stock) for stock in stockslist]

	df=Jq_codelist(stockslist).check_stock_indicator_stocklist(total_money,end_day)

	return df


def buy_sec_filtering(stocklist,total_money,end_day):

	sec_checking_df=Jq_codelist(stocklist).check_stock_indicator_stocklist(total_money,end_day)

	#stock_buying_df=stocklist_buy_number(source_df,stocklist)

	return sec_checking_df



warngap_value=8   #这个是设置一个指数或者其他的观察目标从最高点下跌的比例，作为警示之用，超过一定的下跌幅度，就清仓，避免风险。

def fall_warning(source_df,warngap_value):  #检查dataframe里面是否有危险的指数/股票从高点下跌超过一定的幅度，特别是针对冲高的目标保持警惕。

	name_list=[]

	try:

		for index, row in source_df.iterrows():

		    if row['MHg'] > warngap_value or row['MCg'] > warngap_value:

		        #print(f"Warning: {row['name']} MHg/MCg 的值大于 {warngap_value}，建议 清仓！")

		        name_list.append(row['code']+row['name'])


	except:

		print('目标dataframe里面没有MHg/MCg关键字，请检查你的数据源。')


	print(f"WARNING: 以下标的的 MHg/MCg 的值大于 {warngap_value}, 建议 清仓！")

	print(name_list)



def check_allindex_indicator_strategy(end_day):

	indexlist=filter_allindex(end_day)

	allindexdf=check_index_indicator_indexlist(indexlist,end_day)

	return allindexdf



def get_single_index_indicator_data(index_code,end_day):
	
	my_jq=Jqdata(index_code)

	#sec_ma=my_jq.get_ma(ma_days,end_day)

	security_name=my_jq.security_name()

	sec_closechange=round(my_jq.get_close_change(end_day,1),2)

	yesterday_close=my_jq.get_yesterday_close(end_day)

	present_point,present_open,present_high,present_low=my_jq.get_stock_point(end_day)

	max_high,max_close=my_jq.get_max_high_close(end_day)


	MHg_data=round((max_high - present_point)/max_high,4)*100  #15天内，现价距离最高点的回撤百分比


	MCg_data=round((max_close - present_point)/max_close,4)*100 #15天内，现价距离最高收盘价的回撤百分比

	high_price_percent=round((present_high - yesterday_close)/yesterday_close,4)*100  #当天最高价的涨幅

	high_low_gap=round((present_high - present_low)/yesterday_close,4)*100   #当天截止到目前的振幅   注意振幅是如何计算的，当天最高价-当天最低价 / 昨天收盘价

	
	ma5,ma10,ma20=my_jq.get_ma(5,end_day),my_jq.get_ma(10,end_day),my_jq.get_ma(20,end_day)

	gap_to_ma5=round(round((present_point - ma5)/ma5,4)*100,2)

	gap_to_ma10=round(round((present_point - ma10)/ma10,4)*100,2) 

	gap_to_ma20=round(round((present_point - ma20)/ma20,4)*100,2) 

	#gap_to_ma=round(round((present_point - sec_ma)/sec_ma,4)*100,2)   #距离几天均线还有多少个百分点,正常就是高于均线，负数就是低于均线。


	zhenfu_data=my_jq.get_zhenfu(end_day)

	zhenfu_ranking=my_jq.get_zhenfu_ranking(end_day)


	volume_change=my_jq.get_volume_change(end_day)

	volume_ranking=my_jq.get_volume_ranking(end_day)

	#sec_industry=Jqdata(indexlist[i]).security_industry(end_day)

	highest_price_time=my_jq.find_stock_highest_price_time(end_day)


	if highest_price_time < '10:00':

	#if highest_price_time < datetime.time(10,0,0):

		warningword='危'

	else:

		warningword=' '


	sec_info=[index_code,security_name,present_point,MHg_data,MCg_data,gap_to_ma5,gap_to_ma10,gap_to_ma20,sec_closechange,high_price_percent,present_open,present_high,present_low,high_low_gap,volume_change,volume_ranking,highest_price_time,warningword,zhenfu_data,zhenfu_ranking]


	return sec_info



def check_index_indicator_indexlist(indexlist,end_day):#获取低于5天/10线之类的指数/股票/ETF信息，以便及时发现，并警惕。


	index_info_list=[get_single_index_indicator_data(x,end_day) for x in indexlist]


	df_columns=['code','name','price','MHg','MCg','距5日线','距10日线','距20日线','涨幅','最高涨幅','open','high','low','振幅','vlchg','vlrk','h_p_t','warn','zf','zfr']  #尽可能压缩列的名字长度。


	df=list2dataframe(index_info_list,df_columns)

	df['code']=df['code'].apply(transform_to_normal_code)

	warning_row=df[df['open']==df['high']] #找出开盘价等于最高价的股票/指数，提出警告。

	warning_security_name_list=list(warning_row['name']) #列出这些不好的股票/指数的名单。

	df['name']=df['name'].apply(remove_string_from_index)

	#dataframe_ranking(df,'距均线')

	df.reset_index(drop=True,inplace=True)


	if end_day==datetime.datetime.now().strftime('%Y-%m-%d'):

		date_time=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  #如果要看的是当前日期的，就这么写，可以精确到具体时间。

		print('数据日期：{}'.format(date_time))

	else:

		print('数据日期：{}'.format(end_day))

	for index, row in df.iterrows():

	    if row['MHg'] > warngap_value or row['MCg'] > warngap_value:

	        print(f"Warning: {row['code']} {row['name']} MHg/MCg 的值大于 {warngap_value}，建议 清仓！")


	print('\n **********  开盘价等于最高价的指数有  {},  要当心！********* '.format(warning_security_name_list))


	return df


def sec_buy_number(source_df,code):

	sec_code=Jqdata(code).transform_code(Jqdata(code).jqcode)

	position_list=[]

	try:

		target_position=np.where(source_df['Code']==sec_code)[0][0]

		position_list=[target_position]

		new_df=source_df.iloc[position_list]

		#print(new_df)

	except:

		print('目前的股票池里面没有待查询的代码，这不正常！可能该股票上市没有超过90天！')



	return new_df




def generate_standard_df_for_site(df_list,df_columns,ranking_keyword):
    df=list2dataframe(df_list,df_columns)
    df['ranking']=df.index+1
    new_df=df.drop(df.iloc[:,12:-1],axis=1) #创建一个新的dataframe
    newdf=dataframe_ranking(new_df,ranking_keyword)
    #print(df)
    return newdf




index_list=['sh000001', 'sh000682','sh000685',
  'sh000688', 'sh000852', 'sz399905', 'sh000016', 'sh000097', 'sh000805', 'sz399006', 'sz399004', 'sz399005', 'sz399300', 'sz399330', 'sz399333', 'sz399612', 'sz399673', 'sz399903', 'sh000922',
 'sz399324', 'sh000812', 'sh000811', 'sh000810', 'sh000813', 'sh000814', 'sh000815', 'sh000816', 'sh000818', 'sh000819', 'sh000827', 'sh000901', 'sh000928', 'sh000929', 'sh000930', 'sh000931', 
 'sz399932', 'sh000933', 'sh000934', 'sh000935', 'sh000941', 'sh000977', 'sh000978', 'sh000979', 'sh000986', 'sh000987', 'sh000988', 'sh000989', 'sh000990', 'sh000991', 'sh000992', 'sh000993', 'sh000998', 
 'sz399030', 'sz399060', 'sz399295', 'sz399296', 'sz399393', 'sz399394', 'sz399395', 'sz399396', 'sz399412', 'sz399417', 'sz399419', 'sz399420', 'sz399423', 'sz399432', 'sz399435', 'sz399438', 'sz399439',
  'sz399440', 'sz399441', 'sz399608', 'sz399610', 'sz399646', 'sz399647', 'sz399653', 'sz399654', 'sz399669', 'sz399674', 'sz399675', 'sz399676', 'sz399677', 'sz399683', 'sz399684', 'sz399687', 'sz399693', 
  'sz399695', 'sz399699', 'sz399704', 'sz399705', 'sz399706', 'sz399804', 'sz399805', 'sz399806', 'sz399809', 'sz399807', 'sz399808', 'sz399810', 'sz399811', 'sz399812', 'sz399813', 'sz399814', 'sz399967', 
  'sz399970', 'sz399971', 'sz399973', 'sz399975', 'sz399976', 'sz399986', 'sz399987', 'sz399989', 'sz399993', 'sz399994', 'sz399995', 'sz399996', 'sz399997', 'sz399998','sz399966',]




def daily_check_momentum_score_all_index(index_list,total_money,end_date):

	#但是上面的步骤只是删除了不满足条件的指数，虽然在这个过程中，对所有指数的股票进行了100天线和15缺口的过滤，但下面的buy_listofstocks_volume_filter还会将这个过程重新做一遍
	#因为这次过滤的结果就是将符合条件的股票选择出来，保存起来，并进行数据清理和分析，

	code_df_list_dic={}

	limit_number=len(index_list)

	i=0

	while i < limit_number:

		index_code=index_list[i]

		#print(index_code)

		list_name=index_code

		my_jq=Jqdata(index_code)

		stockslist=my_jq.get_index_securities()

		limit_number_index=len(stockslist)

		i_index=0

		stock_info_list=[]

		while i_index < limit_number_index:

			each_stock_info=[]

			new_jq=Jqdata(stockslist[i_index])

			score,score_days=new_jq.momentum_score(end_date)

			data_date=end_date

			stock_code,stock_name,stock_industry,ATR_value,ATR_days,risk_value,present_point,buying_stock_volume,buying_money=new_jq.buy_stocks_volume(total_money,end_date)

			stock_code=new_jq.transform_code(stock_code)

			pct_money=(round(buying_money/total_money,4))*100

			close_change=new_jq.get_close_change(data_date,1)

			money_change=new_jq.get_money_change(data_date)

			volume_change=new_jq.get_volume_change(data_date)

			each_stock_info=[stock_code,
		                 stock_name,
		                 stock_industry,
		                 score,
		                 ATR_value,
		                 buying_stock_volume,
		                 present_point,
		                 buying_money,
		                 pct_money,
		                 close_change,
		                 volume_change,
		                 money_change,
		                 data_date,
		                 risk_value,
		                 (ATR_days-1),
		                 score_days,
		                ]

			stock_info_list.append(each_stock_info)

			i_index += 1

		stock_info_columns=['Code','Name','Industry','Score','ATR_value','buy_volume',
	            'price_now','buy_value','weight','closechange','volumechange','moneychange',
	            'Data_time','risk','ATR_d','Score_d',]

		df=list2dataframe(stock_info_list,stock_info_columns)

		df.drop(df[pd.isna(df['Score'])].index,inplace=True) 

		df.drop(df[pd.isna(df['price_now'])].index,inplace=True)

		df.sort_values(by=['Score'],ascending=False,inplace=True)

		stock_info_list,stock_info_columns=dataframe2list(df)

		code_df_list_dic.setdefault(list_name,stock_info_list)

		i += 1

	print(''.join(['df_list_score_for_all_index='+str(code_df_list_dic)]))

	#for key,value in code_df_list_dic.items():

		#new_list=[key,value]

		#print(new_list,',')

		#time.sleep(0.005)


	print(''.join(['df_columns_score_for_all_index='+str(stock_info_columns)]))

	return code_df_list_dic,stock_info_columns


def add_ranking_into_dataframe(df,column_name,insert_column_number):

	df_list,df_columns=dataframe2list(df)  #这一行和下一行必须要，这样的话，才能生成全新的df，并且index重新生成，否则是用旧的index.达不到获得ranking的目的。

	new_df=list2dataframe(df_list,df_columns)

	ranking_list=list(new_df.index)

	new_df.insert(insert_column_number,column_name,ranking_list)

	new_df['ranking']=new_df['ranking'] + 1

	return new_df


def check_all_sec_close_change(start_day,end_day):  #获取一段时间内，所有股票的涨跌幅，并按涨跌排序。
    
    stockslist=get_all_sec(end_day)
    
    sec_list=[]
    
    for stock in stockslist:
        
        each_sec_list=[]
        
        my_jq=Jqdata(stock)

        sec_name=my_jq.security_name()

        sec_code=my_jq.transform_code(stock) #要转换到腾讯认可的股票代码。

        price_change=my_jq.get_close_change_within_days(start_day,end_day)

        
        each_sec_list=[sec_code,sec_name,price_change,start_day,end_day]
        
        sec_list.append(each_sec_list)
    
    column_list=['code','name','close_change','start_d','end_d']
    
    df=list2dataframe(sec_list,column_list)

    df.drop(df[pd.isna(df['close_change'])].index,inplace=True) #将数据为nan，这个是数据行的nan的这一行去掉。不看它的结果。   
    
    df.sort_values(by=['close_change'],ascending=False,inplace=True)

    #df=df.reset_index()  #生成index从0开始的dataframe

    #df['ranking']=df.index+1  #因为已经排名好了,所以index + 1就是每行的排名。
    
    #print(df)
    
    df_list,df_columns=dataframe2list(df)

	#new_df=add_ranking_into_dataframe(df,'ranking',3)

	#df_list,df_columns=dataframe2list(new_df)

    print(df_list)  #必须要写这行和下一行，打印出来的越多，结果显示越完整，也容易拷贝。

    print(df_columns)

    return df



index_list=['sh000001','sh000682','sh000685', 
  'sh000688', 'sh000852', 'sz399905', 'sh000016', 'sh000097', 'sh000805', 'sz399006', 'sz399004', 'sz399005', 'sz399300', 'sz399330', 'sz399333', 'sz399612', 'sz399673', 'sz399903', 'sh000922',
 'sz399324', 'sh000812', 'sh000811', 'sh000810', 'sh000813', 'sh000814', 'sh000815', 'sh000816', 'sh000818', 'sh000819', 'sh000827', 'sh000901', 'sh000928', 'sh000929', 'sh000930', 'sh000931', 
 'sz399932', 'sh000933', 'sh000934', 'sh000935', 'sh000941', 'sh000977', 'sh000978', 'sh000979', 'sh000986', 'sh000987', 'sh000988', 'sh000989', 'sh000990', 'sh000991', 'sh000992', 'sh000993', 'sh000998', 
 'sz399030', 'sz399060', 'sz399295', 'sz399296', 'sz399393', 'sz399394', 'sz399395', 'sz399396', 'sz399412', 'sz399417', 'sz399419', 'sz399420', 'sz399423', 'sz399432', 'sz399435', 'sz399438', 'sz399439',
  'sz399440', 'sz399441', 'sz399608', 'sz399610', 'sz399646', 'sz399647', 'sz399653', 'sz399654', 'sz399669', 'sz399674', 'sz399675', 'sz399676', 'sz399677', 'sz399683', 'sz399684', 'sz399687', 'sz399693', 
  'sz399695', 'sz399699', 'sz399704', 'sz399705', 'sz399706', 'sz399804', 'sz399805', 'sz399806', 'sz399809', 'sz399807', 'sz399808', 'sz399810', 'sz399811', 'sz399812', 'sz399813', 'sz399814', 'sz399967', 
  'sz399970', 'sz399971', 'sz399973', 'sz399975', 'sz399976', 'sz399986', 'sz399987', 'sz399989', 'sz399993', 'sz399994', 'sz399995', 'sz399996', 'sz399997', 'sz399998','sz399966',]




index_list=checking_list_for_repeat(index_list)

def judge_index_status(code,end_day):

    my_jq=Jqdata(code)

    code=my_jq.jqcode

    
    df=get_price(code,frequency='daily',end_date=end_day,count=1,fields=['close'])

    bool_value=pd.isna(df.close).values[0]

    if bool_value==True:

    	return False

    else:

    	return True


def index_market_health(index_list,end_day,ma_days=200): #用来判断市场是否健康的标准，默认用200天的均线。

	

	bad_index_number=0

	good_index_list=[]

	no_data_index=[]


	for index_code in index_list:

		x=Jqdata(index_code).judge_index_point(end_day)  #True 代表有数据

		if not x:

			no_data_index.append(index_code)  #如果是无数据，就要收集起来。
		

	print(no_data_index)

	index_list=[x for x in index_list if x not in no_data_index]

	index_number=len(index_list)


	print('请注意以下提到的指数的风险，风险指数越多，说明市场越不好，越要谨慎！')

	print('**************************************************************************')

	for index in index_list:

		my_jq=Jqdata(index)

		bool_value=my_jq.judge_index_above_ma(ma_days,end_day)


		if bool_value==False:

			bad_index_number += 1

		else:

			good_index_list.append(index)


	risk_index=round(bad_index_number/index_number,4)*100

	print('**************************************************************************')


	print('监测总计有 {} 个指数， 提示风险指数有 {} ,占比为  {} %'.format(index_number,bad_index_number,risk_index))

	print('**************************************************************************')

	print('                             ')

	for index in good_index_list:

		my_jq=Jqdata(index)

		sec_code=my_jq.jqcode

		sec_name=my_jq.security_name()

		print('{} {} 在200天均线之上，是好指数，值得关注！'.format(sec_code,sec_name))


	good_index_number=len(good_index_list)

	good_index_pct=round(good_index_number/index_number,4)*100


	print('监测共有 {} 个是指数，好的指数有 {},占比为 {}%'.format(index_number,good_index_number,good_index_pct))

	print('好指数如下：\n')

	good_index_dic={}

	good_index_dic.setdefault(end_day,good_index_list)

	print('above_200_index='+str(good_index_dic))

	print('NO_data_index='+str(no_data_index))





etf_list=[['sz162411', '华宝油气LOF', 0.648, 0.47, 0.451, 5.71, nan, nan, 20.9, 42.11, 39.66], ['sz159930', '能源ETF', 1.064, 1.33, 0.25, 4.31, nan, nan, 17.31, 13.68, 17.57], ['sh515220', '煤炭ETF', 2.325, 1.35, 2.624, 3.84, nan, nan, 20.78, 16.08, 22.69], ['sz160416', '石油基金LOF', 1.222, 0.16, 0.022, 2.35, nan, nan, 14.21, 16.05, 14.96], ['sh510410', '资源ETF', 1.106, 0.45, 0.15, 1.37, nan, nan, 11.83, 4.24, 3.36], ['sz159938', '医药卫生ETF', 1.726, 0.76, 0.121, 1.11, nan, nan, 7.539999999999999, -11.58, -13.09], ['sh515180', '红利ETF易方达', 1.264, -0.32, 0.407, 0.72, nan, nan, 9.53, -0.16, -0.63], ['sh512200', '房地产ETF', 0.815, -2.16, 3.144, 0.49, nan, nan, 18.29, 2.0, 1.37], ['sh512800', '银行ETF', 1.126, 0.36, 1.685, 0.27, nan, nan, 7.55, -0.79, -0.53], ['sh518880', '黄金ETF', 3.849, -0.59, 9.629, -0.44, nan, nan, 0.44, 6.03, 5.65], ['sz159940', '金融ETF基金', 0.961, -0.1, 0.196, -0.62, nan, nan, 7.02, -7.68, -8.129999999999999], ['sh512660', '军工ETF', 1.132, 1.07, 5.954, -0.88, nan, nan, 2.17, -18.15, -22.31], ['sz165525', '基建工程LOF', 0.7659999999999999, -1.03, 0.045, -1.03, nan, nan, 5.510000000000001, -5.779999999999999, -6.239999999999999], ['sh510500', '中证500ETF', 7.021, -0.54, 8.859, -1.57, nan, nan, 5.04, -14.08, -14.98], ['sh512980', '传媒ETF', 0.6559999999999999, -1.35, 0.456, -1.65, nan, nan, 7.89, -19.61, -23.54], ['sh512400', '有色金属ETF', 1.256, 0.0, 1.038, -1.95, nan, nan, 7.26, -3.31, -6.890000000000001], ['sh510050', '上证50ETF', 2.823, -0.46, 9.911, -2.05, nan, nan, 5.06, -13.62, -13.3], ['sh512070', '证券保险ETF', 0.612, -0.97, 1.123, -2.24, nan, nan, 5.149999999999999, -17.07, -17.85], ['sh513090', '香港证券ETF', 1.087, -0.46, 1.879, -2.42, nan, nan, 9.8, -14.88, -16.13], ['sh515210', '钢铁ETF', 1.482, -0.74, 0.776, -2.5, nan, nan, 5.71, -8.459999999999999, -10.88], ['sh512910', '中证100ETF', 1.184, -0.34, 0.083, -2.55, nan, nan, 3.95, -16.03, -15.73], ['sh512880', '证券ETF', 0.93, -0.64, 6.64, -2.72, nan, nan, 2.31, -19.83, -20.92], ['sh515630', '证保ETF', 1.014, -1.07, 0.085, -2.78, nan, nan, 3.79, -17.22, -18.23], ['sz159993', '龙头券商ETF', 0.986, -1.1, 0.255, -2.86, nan, nan, 2.07, -20.16, -21.43], ['sh510300', '沪深300ETF', 4.126, -0.43, 9.577, -2.87, nan, nan, 3.85, -16.21, -16.33], ['sh515880', '通信ETF', 0.909, -1.73, 0.552, -3.5, nan, nan, 2.02, -17.89, -19.06], ['sz159901', '深证100ETF易方达', 3.091, -0.1, 0.687, -3.65, nan, nan, 2.96, -20.78, -20.85], ['sz159949', '创业板50ETF', 1.124, -0.09, 8.658, -3.85, nan, nan, 3.5, -22.54, -23.12], ['sz159902', '中小100ETF', 4.0, -0.2, 0.05, -3.89, nan, nan, 1.42, -18.13, -19.58], ['sz159996', '家电ETF', 0.961, -1.44, 0.417, -3.9, nan, nan, 3.11, -22.81, -24.33], ['sz159928', '消费ETF', 0.977, -0.31, 1.032, -4.12, nan, nan, 1.77, -21.21, -18.99], ['sh588000', '科创50ETF', 1.125, -0.97, 4.752, -4.74, nan, nan, -1.49, -20.27, -22.52], ['sh512580', '碳中和龙头ETF', 1.518, 0.07, 0.586, -4.89, nan, nan, -1.04, -13.11, -16.09], ['sh512720', '计算机ETF', 1.092, -1.44, 0.342, -5.29, nan, nan, -0.8200000000000001, -22.33, -23.37], ['sz159939', '信息技术ETF', 1.158, -1.78, 0.065, -5.62, nan, nan, -0.6, -22.75, -23.82], ['sh515030', '新能源车ETF', 1.914, 0.63, 2.815, -5.949999999999999, nan, nan, 1.97, -16.46, -18.93], ['sh515050', '5GETF', 0.905, -3.0, 1.964, -6.12, nan, nan, -1.2, -28.4, -28.63], ['sh512480', '半导体ETF', 1.027, -2.0, 4.185, -6.64, nan, nan, -1.34, -19.83, -21.0], ['sh512100', '1000ETF', nan, nan, nan, nan, nan, nan, nan, nan, nan]]

etf_columns=['code', 'name', 'close', 'cl_1', 'cl_2','cl_5','cl_10', 'cl_20', 'cl_30',  'cl_60', 'YTD']


def target_date():

	my_time=datetime.datetime.now().strftime('%Y%m%d')
	#year=str(int(my_time[0:4])-1)
	year=str(int(my_time[0:4])-1) #过两天改回来。
	target_date=str(year+'-12-30')     #这个数据应该每年都要改，就是看每年的最后一个交易日是几号，然后就改成这几号，这样的数据就是准的。

	return target_date


def etf_monitor(total_money,datenow):

	etf_df=pd.DataFrame(etf_list,columns=etf_columns)

	code_tuple=tuple(etf_df.code)

	limit_number=len(code_tuple)

	i=0

	while i < limit_number:
	    etf_code=code_tuple[i]
	    target_position=np.where(etf_df['code']==etf_code)[0][0]
	    my_jq=Jqdata(etf_code)

	    #etf_df.loc[target_position,'daily_change']=my_jq.get_close_change(datenow,1)
	    present_close,b,c,d=my_jq.get_point()


	    etf_df.loc[target_position,'close']=present_close


	    


	    etf_df.loc[target_position,'cl_1']=my_jq.get_close_change(datenow,1)
	    etf_df.loc[target_position,'cl_2']=my_jq.get_close_change(datenow,2)
	    etf_df.loc[target_position,'cl_5']=my_jq.get_close_change(datenow,5)
	    etf_df.loc[target_position,'cl_10']=my_jq.get_close_change(datenow,10)
	    etf_df.loc[target_position,'cl_20']=my_jq.get_close_change(datenow,20)
	    etf_df.loc[target_position,'cl_30']=my_jq.get_close_change(datenow,30)
	    etf_df.loc[target_position,'cl_60']=my_jq.get_close_change(datenow,60)

	    etf_df.loc[target_position,'YTD']=my_jq.get_close_change_YTD(target_date()) 


	    etf_df.loc[target_position,'money']=my_jq.get_present_money(datenow)

	    etf_buy_volume=my_jq.buy_single_etf_volume(total_money,datenow)

	    etf_df.loc[target_position,'b_vl']=etf_buy_volume

	    etf_df.loc[target_position,'data_time']=datenow
	    
	    i += 1

	etf_df.sort_values(by=['cl_5'],ascending=False,inplace=True)

	etfdf_list,etfdf_columns=dataframe2list(etf_df)

	print(''.join(['etfdf_list='+str(etfdf_list)]))

	print(''.join(['etfdf_columns='+str(etfdf_columns)]))

	#print(etf_df)


	return etfdf_list,etfdf_columns




def daily_etf_monitor(total_money,datenow,rankingkeyword):  #这是线上的实时的监测etf的数据。

	etf_df=pd.DataFrame(etf_list,columns=etf_columns)

	code_tuple=tuple(etf_df.code)

	limit_number=len(code_tuple)

	if datenow==datetime.datetime.now().strftime('%Y-%m-%d'):

		date_time=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  #如果要看的是当前日期的，就这么写，可以精确到具体时间。

	else:

		date_time=datenow

	print('数据日期：{}'.format(date_time))

	#print('数据日期: {}'.format(datenow))

	i=0

	while i < limit_number:
	    etf_code=code_tuple[i]
	    target_position=np.where(etf_df['code']==etf_code)[0][0]
	    my_jq=Jqdata(etf_code)

	    #etf_df.loc[target_position,'daily_change']=my_jq.get_close_change(datenow,1)
	    present_close,b,c,d=my_jq.get_point()

	    etf_df.loc[target_position,'b_vl']=my_jq.buy_single_etf_volume(total_money,datenow)
	    etf_df.loc[target_position,'close']=present_close

	    


	    etf_df.loc[target_position,'cl_1']=round(my_jq.get_close_change(datenow,1),2)
	    #etf_df['cl_1']=etf_df['cl_1'].apply(lambda x: round(x,2))
	    etf_df.loc[target_position,'cl_2']=round(my_jq.get_close_change(datenow,2),2)
	    etf_df.loc[target_position,'cl_5']=round(my_jq.get_close_change(datenow,5),2)
	    etf_df.loc[target_position,'cl_10']=round(my_jq.get_close_change(datenow,10),2)
	    etf_df.loc[target_position,'cl_20']=round(my_jq.get_close_change(datenow,20),2)
	    etf_df.loc[target_position,'cl_30']=round(my_jq.get_close_change(datenow,30),2)
	    etf_df.loc[target_position,'cl_60']=round(my_jq.get_close_change(datenow,60),2)

	    etf_df.loc[target_position,'YTD']=round(my_jq.get_close_change_YTD(target_date()),2)

	    etf_df.loc[target_position,'vlchg']=my_jq.get_volume_change(datenow)

	    etf_df.loc[target_position,'vlrk']=my_jq.get_volume_ranking(datenow)

	    etf_df.loc[target_position,'h_p_t']=my_jq.find_stock_highest_price_time(datenow)


	    #etf_df.loc[target_position,'money']=my_jq.get_present_money(datenow)

	    #etf_df.loc[target_position,'data_time']=datenow
	    
	    i += 1

	etf_df.sort_values(by=[rankingkeyword],ascending=False,inplace=True)

	#etf_df.loc[etf_df['h_p_t'] < datetime.time(10,0,0),'warn'] = '危'

	etf_df.loc[etf_df['h_p_t'] < '10:00','warn'] = '危'

	#etf_df.loc[etf_df['h_p_t'] > datetime.time(10,0,0),'warn'] = ' '

	etf_df.loc[etf_df['h_p_t'] > '10:00','warn'] = ' '

	etf_df['b_vl']=etf_df['b_vl'].round().astype(int)

	etf_df['vlrk']=etf_df['vlrk'].round().astype(int)

	#new_columns=['code', 'name', 'close', 'cl_1', 'cl_2', 'cl_5', 'cl_10', 'cl_20',
    #   'cl_30', 'cl_60', 'YTD', 'b_vl', 'vlchg', 'vlrk','h_p_t','warn','money',]

	#etf_df.reindex(columns=new_columns)

	filter_etf_with_tradingmoney(etf_df,'code',datenow)

	etf_df.reset_index(drop=True,inplace=True)   #排序很重要，就是每次排序以后，重新整理下索引比较好些，养成好习惯。

	#etfdf_list,etfdf_columns=dataframe2list(etf_df)

	#print(''.join(['etfdf_list='+str(etfdf_list)]))

	#print(''.join(['etfdf_columns='+str(etfdf_columns)]))

	#print(etf_df)

	#etf_df.set_index('code',inplace=True)   #把code列作为索引，但是现在可以冻结所有列了，就不这么做了！否则冻结列之后就出现了两行列，第一行是所有列，第二行是code列（索引），效果就不好了。


	return etf_df



etfdf_list=[['560050', 'MSCIA50', 'MSCI中国A50互联互通'], ['512380', 'MSCI中国', 'MSCI中国A股'], ['512360', 'MSCI国际', 'MSCI中国A股国际'], ['515160', 'MSCI招商', 'MSCI中国A股国际通'], ['159980', '有色ETF', '上期有色金属'], ['510180', '180ETF', '上证180'], ['510030', '价值ETF', '上证180价值'], ['510010', '治理ETF', '上证180公司治理'], ['510230', '金融ETF', '上证180金融股'], ['510290', '380ETF', '上证380'], ['510050', '50ETF', '上证50'], ['510060', '央企ETF', '上证中央企业50'], ['510130', '中盘ETF', '上证中盘'], ['510630', '消费30', '上证主要消费行业'], ['510660', '医药行业', '上证医药卫生行业'], ['510270', '国企ETF', '上证国有企业100'], ['510170', '商品ETF', '上证大宗商品股票'], ['510150', '消费ETF', '上证消费80'], ['513990', '港股通ETF', '上证港股通'], ['588000', '科创50', '上证科创板50成份'], ['588100', '科创信息', '上证科创板新一代信息技术'], ['588160', '新材科创', '上证科创板新材料'], ['588200', '科创芯片', '上证科创板芯片'], ['510880', '红利ETF', '上证红利'], ['510210', '综指ETF', '上证综合'], ['510410', '资源ETF', '上证自然资源'], ['510200', '上证券商', '上证证券行业'], ['510020', '超大ETF', '上证超级大盘'], ['510650', '金融行业', '上证金融地产行业'], ['159866', '日经ETF', '东京日经225'], ['513800', '东证ETF', '东京证券交易所股票价格'], ['512760', '芯片ETF', '中华交易服务半导体芯片行业'], ['513900', '港股100', '中华交易服务港股通精选100人民币'], ['512770', '战略新兴', '中国战略新兴产业成份'], ['159902', '中小100ETF', '中小企业100'], ['159907', '中小300ETF', '中小企业300'], ['159918', '中创400ETF', '中小创业企业400'], ['562000', '中证100ETF基金', '中证100'], ['512100', '1000ETF', '中证1000'], ['510990', 'ESG180', '中证180ESG'], ['512330', '信息ETF', '中证500信息技术'], ['515590', '500ETFEW', '中证500等权重'], ['512260', '500低波', '中证500行业中性低波动'], ['560500', '500质量', '中证500质量成长'], ['159811', '5G50ETF', '中证5G产业50'], ['515050', '5GETF', '中证5G通信主题'], ['515800', '800ETF', '中证800'], ['159707', '地产ETF', '中证800地产'], ['516110', '汽车ETF', '中证800汽车与零部件'], ['515630', '保险证券', '中证800证券保险'], ['159887', '银行ETF', '中证800银行'], ['516720', 'ESGETF', '中证ESG 120策略'], ['510810', '上海国企', '中证上海国企'], ['562990', '碳中和E', '中证上海环交所碳中和'], ['560080', '中药ETF', '中证中药'], ['159928', '消费ETF', '中证主要消费'], ['560660', '云50ETF', '中证云计算50'], ['516510', '云计算', '中证云计算与大数据主题'], ['159819', '人工智能ETF', '中证人工智能主题'], ['515980', '人工智能', '中证人工智能产业'], ['561500', '漂亮50', '中证企业核心竞争力50'], ['512980', '传媒ETF', '中证传媒'], ['159613', '信息安全ETF', '中证信息安全主题'], ['515790', '光伏ETF', '中证光伏产业'], ['560980', '光伏龙头', '中证光伏龙头30'], ['159939', '信息技术ETF', '中证全指信息技术'], ['159873', '医疗设备ETF', '中证全指医疗保健设备与服务'], ['159883', '医疗器械', '中证全指医疗器械'], ['159938', '医药卫生ETF', '中证全指医药卫生'], ['512480', '半导体', '中证全指半导体产品与设备'], ['159944', '材料ETF', '中证全指原材料'], ['159936', '可选消费ETF', '中证全指可选消费'], ['159996', '家电ETF', '中证全指家用电器'], ['159745', '建材ETF', '中证全指建筑材料'], ['512200', '地产ETF', '中证全指房地产'], ['159611', '电力ETF', '中证全指电力公用事业'], ['159945', '能源ETF基金', '中证全指能源'], ['512880', '证券ETF', '中证全指证券公司'], ['515230', '软件ETF', '中证全指软件'], ['515880', '通信ETF', '中证全指通信设备'], ['159940', '金融地产ETF', '中证全指金融地产'], ['513220', '全球互联', '中证全球中国互联网'], ['513360', '教育ETF', '中证全球中国教育主题'], ['516560', '养老ETF', '中证养老产业'], ['159790', '碳中和ETF', '中证内地低碳经济主题'], ['159768', '房地产ETF', '中证内地地产主题'], ['159752', '新能源50', '中证内地新能源主题'], ['561320', '交运ETF', '中证内地运输主题'], ['512660', '军工ETF', '中证军工'], ['512710', '军工龙头', '中证军工龙头'], ['159825', '农业ETF', '中证农业主题'], ['159616', '农牧ETF', '中证农牧主题'], ['159992', '创新药', '中证创新药产业'], ['159869', '游戏ETF', '中证动漫游戏'], ['512170', '医疗ETF', '中证医疗'], ['515950', '医药龙头', '中证医药50'], ['515960', '医药100', '中证医药健康100策略'], ['159929', '医药ETF', '中证医药卫生'], ['516820', '医疗创新', '中证医药及医疗器械创新'], ['510160', '产业升级', '中证南方小康产业'], ['515090', '可持续', '中证可持续发展100'], ['515150', '国企富国', '中证国企一带一路'], ['512040', '国信价值', '中证国信价值'], ['512670', '国防ETF', '中证国防'], ['516950', '基建ETF', '中证基建'], ['516970', '基建50', '中证基建工程'], ['516550', '农业ETF', '中证大农业'], ['515400', '大数据', '中证大数据产业'], ['515900', '央创ETF', '中证央企创新驱动'], ['512950', '央企改革', '中证央企结构调整'], ['510500', '500ETF', '中证小盘500'], ['159778', '工业互联ETF', '中证工业互联网主题'], ['159855', '影视ETF', '中证影视主题'], ['159623', '成渝经济圈ETF', '中证成渝地区双城经济圈成份'], ['560800', '数字ETF', '中证数字经济主题'], ['516190', '文娱ETF', '中证文娱传媒'], ['515860', '科技100', '中证新兴科技100策略'], ['561130', '国货ETF', '中证新华社民族品牌工程'], ['159761', '新材料50ETF', '中证新材料主题'], ['516160', '新能源', '中证新能源'], ['515030', '新汽车', '中证新能源汽车'], ['515700', '新能车', '中证新能源汽车产业'], ['159766', '旅游ETF', '中证旅游主题'], ['516800', '智能制造ETF', '中证智能制造主题'], ['515250', '智能汽车', '中证智能汽车主题'], ['515920', '智能消费', '中证智能消费主题'], ['516380', '智能电动车ETF', '中证智能电动汽车'], ['562530', '1000价值', '中证智选1000价值稳健策略'], ['562520', '1000成长', '中证智选1000成长创新策略'], ['159617', '500价值ETF', '中证智选500价值稳健策略'], ['159620', '500成长ETF', '中证智选500成长创新策略'], ['159881', '有色60', '中证有色金属'], ['561330', '矿业ETF', '中证有色金属矿业主题'], ['562500', '机器人', '中证机器人'], ['159667', '工业母机ETF', '中证机床'], ['512870', '杭州湾区', '中证杭州湾区'], ['517300', 'AH300ETF', '中证沪港深300'], ['517080', 'HGS500', '中证沪港深500'], ['517990', '医药基金', '中证沪港深500医药卫生全收益'], ['517390', '云计算AH', '中证沪港深云计算产业'], ['517050', '互联网50', '中证沪港深互联网'], ['517800', 'AI50', '中证沪港深人工智能50'], ['517120', '创新药50', '中证沪港深创新药产业'],  ['517880', '品牌消费', '中证沪港深品牌消费50'], ['517850', '张江ETF', '中证沪港深张江自主创新50'], ['517550', '消费主题', '中证沪港深消费龙头'], ['517770', '游戏传媒', '中证沪港深游戏及文化传媒'], ['517660', '物联网AH', '中证沪港深物联网主题'], ['517960', '科技AH', '中证沪港深科技100'], ['517350', '中概科技', '中证沪港深科技龙头'], ['159793', '线上消费ETF平安', '中证沪港深线上消费主题'], ['512190', '之江凤凰', '中证浙江凤凰行动50'], ['515760', '浙江国资', '中证浙江国资创新发展'], ['159605', '中概互联ETF', '中证海外中国互联网30'], ['513050', '中概互联', '中证海外中国互联网50'], ['515650', '消费50', '中证消费50'], ['516600', '消服ETF', '中证消费服务领先'], ['561100', '电子龙头', '中证消费电子主题'], ['516130', '消费龙头ETF', '中证消费龙头'], ['159711', '港股通ETF', '中证港股通50'], ['513550', '港股通50', '中证港股通50人民币'], ['159788', '港股通100ETF', '中证港股通中国100人民币'], ['159792', '港股通互联网ETF', '中证港股通互联网'], ['513700', '香港医药ETF', '中证港股通医药卫生综合'], ['159735', '港股消费ETF', '中证港股通消费主题'], ['513070', 'HK消费50', '中证港股通消费主题人民币'], ['513860', 'HKC科技', '中证港股通科技'], ['513980', '科技港股', '中证港股通科技人民币'], ['513530', '港股红利', '中证港股通高股息投资'], ['159743', '湖北ETF', '中证湖北新旧动能转换'], ['515220', '煤炭ETF', '中证煤炭'], ['516260', '物联网50', '中证物联网主题'], ['512580', '碳中和', '中证环保产业'], ['159861', '环保ETF', '中证环保产业50'], ['562900', '现代农业', '中证现代农业主题'], ['516910', '物流ETF', '中证现代物流'], ['512290', '生物医药', '中证生物医药'], ['159837', '生物科技ETF', '中证生物科技主题'], ['512400', '有色ETF', '中证申万有色金属'], ['516900', '食品50', '中证申万食品饮料'], ['159997', '电子ETF', '中证电子'], ['515260', '电子ETF', '中证电子50'], ['159796', '电池50ETF', '中证电池主题'], ['159865', '养殖ETF', '中证畜牧养殖'], ['561710', '疫苗生物ETF', '中证疫苗与生物技术'], ['159731', '石化ETF', '中证石化产业'], ['515200', '创新100', '中证研发创新100'], ['560960', '碳60ETF', '中证碳中和60'], ['159781', '双创50ETF', '中证科创创业50'], ['515580', '中证科技', '中证科技100'], ['159807', '科技ETF', '中证科技50'], ['515750', '科技50', '中证科技50策略'], ['512220', 'TMTETF', '中证科技传媒通信150'], ['560990', '科技先锋', '中证科技先锋'], ['515000', '科技ETF', '中证科技龙头'], ['516150', '稀土基金', '中证稀土产业'], ['562800', '稀有金属', '中证稀有金属主题'], ['512970', '湾区ETF', '中证粤港澳大湾区发展主题'], ['515080', '中证红利', '中证红利'], ['512890', '红利LV', '中证红利低波动'], ['515100', '红利100', '中证红利低波动100'], ['159758', '红利50ETF', '中证红利质量'], ['159725', '线上消费', '中证线上消费主题'], ['159870', '化工ETF', '中证细分化工产业主题'], ['512120', '医药50', '中证细分医药产业主题'], ['516650', '有色50', '中证细分有色金属产业主题'], ['516960', '机械ETF', '中证细分机械设备产业主题'], ['515170', '食品饮料', '中证细分食品饮料产业主题'], ['562010', '绿色能源', '中证绿色能源'], ['159930', '能源ETF', '中证能源'], ['516640', '芯片龙头', '中证芯片产业'], ['159786', 'VRETF', '中证虚拟现实主题'], ['562910', '高端制造', '中证装备产业'], ['159998', '计算机', '中证计算机'], ['516730', '证券公司', '中证证券公司30'], ['516980', '证券先锋', '中证证券公司先锋策略'], ['159872', '智能网联汽车ETF', '中证车联网主题'], ['159852', '软件ETF', '中证软件服务'], ['512690', '酒ETF', '中证酒'], ['512640', '金融地产', '中证金融地产'], ['159851', '金科ETF', '中证金融科技主题'], ['515210', '钢铁ETF', '中证钢铁'], ['512800', '银行ETF', '中证银行'], ['517900', '银行优选', '中证银行AH价格优选'], ['512750', '基本面50', '中证锐联基本面50'], ['512650', '长三角', '中证长三角一体化发展主题'], ['517160', '长江ETF', '中证长江保护主题'], ['513310', '中韩芯片', '中证韩交所中韩半导体'], ['159736', '饮食ETF', '中证食品饮料'], ['513140', '港股金融', '中证香港300金融服务人民币'], ['513120', 'HK创新药', '中证香港创新药'], ['159747', '香港科技ETF', '中证香港科技'], ['513090', '香港证券', '中证香港证券投资主题人民币'], ['159638', '高端装备ETF', '中证高端装备细分50'], ['159915', '创业板ETF', '创业板'], ['159836', '创300ETF', '创业板300'], ['159949', '创业板50', '创业板50'], ['159966', '创业板价值ETF', '创业板低波价值'], ['159967', '创业板成长ETF', '创业板动量成长'], ['159814', '创业大盘', '创业板大盘'], ['159773', '创科技ETF', '创业板科技'], ['159628', '国证2000ETF', '国证2000'], ['159717', 'ESGETF', '国证ESG300'], ['159662', '交运ETF', '国证交通运输行业'], ['159760', '泰康公卫健康ETF', '国证公共卫生与健康'], ['159804', '创精选88', '国证创业板中盘精选88'], ['159995', '芯片ETF', '国证半导体芯片'], ['159728', '在线消费ETF', '国证在线消费'], ['159755', '电池ETF', '国证新能源车电池'], ['159880', '有色ETF基金', '国证有色金属行业'], ['159973', '民企领先100ETF', '国证民企领先100'], ['159732', '消费电子ETF', '国证消费电子主题'], ['159636', '港股通科技30ETF', '国证港股通科技'], ['159859', '生物药30', '国证生物医药'], ['159643', '疫苗ETF', '国证疫苗与生物科技'], ['159625', '绿色电力ETF', '国证绿色电力'], ['159993', '龙头券商', '国证证券龙头'], ['159843', '食品饮料ETF', '国证食品饮料行业'], ['159730', '龙头家电ETF', '国证龙头家电'], ['159965', '央视50', '央视财经50'], ['512550', 'A50基金', '富时中国A50'], ['517090', '共赢ETF', '富时中国国企开放共赢'], ['159990', '小盘价值', '巨潮小盘价值'], ['513030', '德国ETF', '德国法兰克福DAX'], ['510900', 'H股ETF', '恒生中国企业'], ['159726', '恒生红利ETF', '恒生中国内地企业高股息率'], ['513330', '恒生互联', '恒生互联网科技业'], ['513060', '恒生医疗', '恒生医疗保健'], ['517380', '创新药AH', '恒生沪深港创新药精选50'], ['513160', '科技30', '恒生港股通中国科技'], ['513320', 'HK新经济', '恒生港股通新经济'], ['513690', '恒生股息', '恒生港股通高股息率'], ['513180', '恒指科技', '恒生科技'], ['159892', '恒生医药ETF', '恒生香港上市生物科技'], ['159981', '能源化工', '易盛郑商所能源化工A'], ['513500', '标普500', '标准普尔500'], ['510300', '300ETF', '沪深300'], ['516830', '300ESG', '沪深300ESG基准'], ['562320', '300价值A', '沪深300价值'], ['512010', '医药ETF', '沪深300医药卫生'], ['562310', '300成长', '沪深300成长'], ['512530', '300红利', '沪深300红利'], ['515300', '红利300', '沪深300红利低波动'], ['159933', '国投金融地产ETF', '沪深300金融地产'], ['512070', '证券保险', '沪深300非银行金融'], ['513080', '法国ETF', '法国巴黎CAC40'], ['159901', '深证100ETF', '深证100'], ['159912', '深300ETF', '深证300'], ['159913', '深价值', '深证300价值'], ['159716', '深创100ETF', '深证创新100'], ['159910', '基本面120ETF', '深证基本面120'], ['159916', '深F60ETF', '深证基本面60'], ['159903', '深成ETF', '深证成份'], ['159906', '深成长龙头ETF', '深证成长40'], ['159709', '物联基金', '深证物联网50'], ['159909', 'TMT50ETF', '深证电子信息传媒产业50'], ['159905', '深红利ETF', '深证红利'], ['159976', '湾创ETF', '粤港澳大湾区创新100'], ['159941', '纳指ETF', '纳斯达克100'], ['159920', '恒生ETF', '香港恒生'], ['165525', '基建工程LOF', '基建工程'], ['160416', '石油基金LOF', '国际油气QDII'], ['162411', '华宝油气LOF', '境外油气QDII'], ['560180', 'ESG300ETF', '300ESG'], ['562330', '中证500价值ETF', '500价值'], ['159606', '中证500成长ETF', '500质量'], ['159896', '物联网龙头ETF', 'CS物联网'], ['561160', '锂电池ETF', 'CS电池'], ['512720', '计算机ETF', 'CS计算机'], ['159653', 'ESG300ETF', 'ESG300'], ['510770', 'G60创新ETF', 'G60成指'], ['512390', '平安MSCI低波ETF', 'MSCI中国A股低波'], ['515910', '质量ETF', 'MSCI中国A股国际质量'], ['512090', 'MSCIA股ETF易方达', 'MSCI中国A股国际通'], ['561190', '双碳ETF', 'SEEE碳中和'], ['159738', '云计算ETF沪港深', 'SHS云计算人民币'], ['159856', '互联网龙头ETF', 'SHS互联网人民币'], ['159622', '创新药ETF沪港深', 'SHS创新药人民币'], ['159723', '科技龙头ETF', 'SHS科技龙头'], ['517280', '网购ETF', 'SHS线上消费人民币'], ['561180', 'A100ETF', '中证100'], ['159845', '中证1000ETF', '中证1000'], ['159647', '中药ETF', '中证中药'], ['516620', '影视ETF', '中证影视'], ['560650', '核心50ETF', '企业核心竞争力50'], ['562920', '信息安全ETF', '信息安全'], ['159808', '创业板ETF融通', '创业板R'], ['515110', '一带一路国企ETF', '国企一带一路'], ['159697', '油气ETF', '国证油气'], ['159507', '电信ETF', '国证通信'], ['561580', '央企红利ETF', '央企红利'], ['560700', '央企红利50ETF', '央企股东回报'], ['517180', '中国国企ETF', '富时中国国企开放共赢'], ['560860', '工业有色ETF', '工业有色'], ['159658', '数字经济ETF', '数字经济'], ['516270', '新能源50ETF', '新能源'], ['516590', '智能汽车50ETF', '智能电车'], ['515450', '红利低波50ETF', '标普中国A股大盘红利低波50'], ['159721', '深创龙头ETF', '深创100'], ['159943', '深证成指ETF', '深证成指'],  ['159983', '粤港澳大湾区ETF', '湾创100'],['515320', '电子50ETF', '电子50'], ['562860', '疫苗ETF', '疫苗生物'], ['516570', '化工行业ETF', '石化产业'], ['588010', '科创新材料ETF', '科创材料'], ['516220', '化工龙头ETF', '细分化工全收益'], ['159652', '有色50ETF', '细分有色'], ['159886', '机械ETF', '细分机械'], ['561170', '绿电50ETF', '绿色电力'], ['159692', '证券ETF东财', '证券公司30'], ['510090', 'ESGETF基金', '责任'], ['159666', '交通运输ETF', '运输']]
etfdf_columns=['code', 'name', 'index']


def find_target_etf_in_trend(stockslist,total_money,ma_days,end_day):  #获取全市场中，过去2天，收盘价都上升的etf基金。并且在5天线上的ETF基金。


	# 找出连涨2天的股票。

	good_trend_stock_list=[good_stock for good_stock in stockslist if Jqdata(good_stock).judge_stock_increase(end_day)==True]  #股票代码没有变化，按平台来的。

	#good_trend_stock_list_code_transformed=[Jqdata(sec).transform_code(sec) for sec in good_trend_stock_list ]  #股票代码发生变化，按腾讯来的。

	if good_trend_stock_list!=[]:


		above_ma_stock_list=[goodsec for goodsec in good_trend_stock_list if Jqdata(goodsec).judge_above_ma(ma_days,end_day)==True]   #股票代码没有变化，按平台来的。

	#above_ma_stock_list_code_transformed=[Jqdata(goodsec).transform_code(goodsec) for goodsec in above_ma_stock_list]  #股票代码发生变化，按腾讯来的。

	else:

		above_ma_stock_list=[]  #列表为空。


	if above_ma_stock_list!=[]:

	#final_df=get_target_stocklist_closechange_data(above_ma_stock_list_code_transformed,total_money,end_day)  #将得到的高于5天线，连涨两天的股票的信息，列出来，作为参考。

		final_df=get_target_etflist_closechange_data(above_ma_stock_list,total_money,end_day)

		dataframe_ranking(final_df,'vlchg')   #以当天的成交量变化，来排序。


	else:

		final_df=pd.DataFrame()  #返回一个空的。

	#final_df.reset_index(drop=True,inplace=True)   #又重新生成索引。

	#final_df.reset_index(inplace=True)


	return final_df    #返回最终的代码列表，和一个dataframe.


def find_etf_in_trend(total_money,ma_days,end_day):

	alletf_df=list2dataframe(etfdf_list,etfdf_columns)

	alletf_list=list(alletf_df['code'])

	stockslist=alletf_list

	final_df=find_target_etf_in_trend(stockslist,total_money,ma_days,end_day)

	return final_df


def find_etf_from_jq_in_trend(total_money,ma_days,end_day):

	etflist=get_all_etf(end_day)

	etflist=[transform_to_normal_code(x) for x in etflist]

	final_df=find_target_etf_in_trend(etflist,total_money,ma_days,end_day)

	return final_df



#与上面的函数的区别是这个函数返回的是一个代码列表，可以在某些场景下不需要运行get_target_etflist_closechange_data函数，节约时间。

def find_etf_in_trend_with_codelist(total_money,ma_days,end_day):  #获取全市场中，过去2天，收盘价都上升的etf基金。并且在5天线上的ETF基金。


	alletf_df=list2dataframe(etfdf_list,etfdf_columns)

	alletf_list=list(alletf_df['code'])

	stockslist=alletf_list

	# 找出连涨2天的股票。

	good_trend_stock_list=[good_stock for good_stock in stockslist if Jqdata(good_stock).judge_stock_increase(end_day)==True]  #股票代码没有变化，按平台来的。

	#good_trend_stock_list_code_transformed=[Jqdata(sec).transform_code(sec) for sec in good_trend_stock_list ]  #股票代码发生变化，按腾讯来的。

	if good_trend_stock_list!=[]:


		above_ma_stock_list=[goodsec for goodsec in good_trend_stock_list if Jqdata(goodsec).judge_above_ma(ma_days,end_day)==True]   #股票代码没有变化，按平台来的。

	#above_ma_stock_list_code_transformed=[Jqdata(goodsec).transform_code(goodsec) for goodsec in above_ma_stock_list]  #股票代码发生变化，按腾讯来的。

	else:

		above_ma_stock_list=[]  #列表为空。


	return above_ma_stock_list  #返回最终的代码列表





def find_etf_high_low_increase(total_money,ma_days,end_day):  #获取全市场中，过去2天，每天高点低点，不停上升的etf基金。并且在5天线上的ETF基金。


	alletf_df=list2dataframe(etfdf_list,etfdf_columns)

	alletf_list=list(alletf_df['code'])

	stockslist=alletf_list

	# 找出连涨2天的股票。

	good_trend_stock_list=[good_stock for good_stock in stockslist if Jqdata(good_stock).judge_stock_trend(end_day)==True]  #股票代码没有变化，按平台来的。

	#good_trend_stock_list_code_transformed=[Jqdata(sec).transform_code(sec) for sec in good_trend_stock_list ]  #股票代码发生变化，按腾讯来的。

	if good_trend_stock_list !=[]:

		above_ma_stock_list=[goodsec for goodsec in good_trend_stock_list if Jqdata(goodsec).judge_above_ma(ma_days,end_day)==True]   #股票代码没有变化，按平台来的。

	#above_ma_stock_list_code_transformed=[Jqdata(goodsec).transform_code(goodsec) for goodsec in above_ma_stock_list]  #股票代码发生变化，按腾讯来的。

	else:

		above_ma_stock_list=[]

	#final_df=get_target_stocklist_closechange_data(above_ma_stock_list_code_transformed,total_money,end_day)  #将得到的高于5天线，连涨两天的股票的信息，列出来，作为参考。

	if above_ma_stock_list!=[]:

		final_df=get_target_etflist_closechange_data(above_ma_stock_list,total_money,end_day)

		dataframe_ranking(final_df,'vlchg')   #以当天的成交量变化，来排序。

	else:

		final_df=pd.DataFrame()

	#final_df.reset_index(drop=True,inplace=True)   #又重新生成索引。

	#final_df.reset_index(inplace=True)


	return final_df    #返回最终的代码列表，和一个dataframe.



#和上面函数的区别就是返回的是一个代码列表，不用运行get_target_etflist_closechange_data函数，节约时间。

def find_etf_high_low_increase_with_codelist(total_money,ma_days,end_day):  #获取全市场中，过去2天，每天高点低点，不停上升的etf基金。并且在5天线上的ETF基金。


	alletf_df=list2dataframe(etfdf_list,etfdf_columns)

	alletf_list=list(alletf_df['code'])

	stockslist=alletf_list

	# 找出连涨2天的股票。

	good_trend_stock_list=[good_stock for good_stock in stockslist if Jqdata(good_stock).judge_stock_trend(end_day)==True]  #股票代码没有变化，按平台来的。

	#good_trend_stock_list_code_transformed=[Jqdata(sec).transform_code(sec) for sec in good_trend_stock_list ]  #股票代码发生变化，按腾讯来的。

	if good_trend_stock_list !=[]:

		above_ma_stock_list=[goodsec for goodsec in good_trend_stock_list if Jqdata(goodsec).judge_above_ma(ma_days,end_day)==True]   #股票代码没有变化，按平台来的。

	#above_ma_stock_list_code_transformed=[Jqdata(goodsec).transform_code(goodsec) for goodsec in above_ma_stock_list]  #股票代码发生变化，按腾讯来的。

	else:

		above_ma_stock_list=[]

	#final_df=get_target_stocklist_closechange_data(above_ma_stock_list_code_transformed,total_money,end_day)  #将得到的高于5天线，连涨两天的股票的信息，列出来，作为参考。


	return above_ma_stock_list   #返回最终的代码列表

'''
def find_etf_in_trend(total_money,ma_days,end_day):

	alletf_df=list2dataframe(etfdf_list,etfdf_columns)

	alletf_list=list(alletf_df['code'])

	index_code_list,final_etf_df=find_stock_in_trend_onsite(total_money,ma_days,end_day)

	return final_etf_df
'''


def get_single_etf_data(etf_code,total_money,datenow):

	my_jq=Jqdata(etf_code)

	etf_b_vl=my_jq.buy_single_etf_volume(total_money,datenow)

	max_high,max_close=my_jq.get_max_high_close(datenow)

	present_max_high_change,present_max_low_change=my_jq.get_present_high_low_change(datenow)

	#etf_df.loc[target_position,'daily_change']=my_jq.get_close_change(datenow,1)
	present_close,b,c,d=my_jq.get_point()


	etf_MHg=round((max_high - present_close)/max_high,4)*100  #15天内，现价距离最高点的回撤百分比

	etf_MCg=round((max_close - present_close)/max_close,4)*100 #15天内，现价距离最高收盘价的回撤百分比

	etf_cl_changes = [round(my_jq.get_close_change(datenow, i), 2) for i in [1, 2, 5, 10, 20, 30, 60]]

	etf_cl_YTD=round(my_jq.get_close_change_YTD(target_date()),2)

	etf_vlchg=my_jq.get_volume_change(datenow)

	etf_vlrk=my_jq.get_volume_ranking(datenow)

	etf_h_p_t=my_jq.find_stock_highest_price_time(datenow)

	etf_warn=my_jq.judge_stock_highest_price_time_risk(datenow)

	etf_zf=my_jq.get_zhenfu(datenow)

	etf_zfr=my_jq.get_zhenfu_ranking(datenow)

	etf_info_list = [etf_b_vl, present_close,present_max_high_change,present_max_low_change,etf_MHg, etf_MCg] + etf_cl_changes + [etf_cl_YTD, etf_vlchg, etf_vlrk, etf_h_p_t, etf_warn, etf_zf, etf_zfr]

	#etf_info_list=[etf_code,etf_b_vl,etf_MHg,etf_MCg,etf_cl_1,etf_cl_2,etf_cl_5,etf_cl_10,etf_cl_20,etf_cl_30,etf_cl_60,etf_cl_YTD,etf_vlchg,etf_vlrk,etf_h_p_t,etf_warn,etf_zf,etf_zfr]

	#return etf_b_vl,etf_MHg,etf_MCg,etf_cl_1,etf_cl_2,etf_cl_5,etf_cl_10,etf_cl_20,etf_cl_30,etf_cl_60,etf_cl_YTD,etf_vlchg,etf_vlrk,etf_h_p_t,etf_warn,etf_zf,etf_zfr
	return etf_info_list


def daily_all_etf_monitor(total_money,datenow):  #这是线上的实时的监测etf的数据。监测的是所有的etf的市场表现。会把成交量低的ETF干掉。


	new_etfdf_list=[single_etf_list+get_single_etf_data(single_etf_list[0],total_money,datenow) for single_etf_list in etfdf_list]

	etfdf_columns=['code', 'name', 'index', 'b_vl', 'close','Lup','Ldown','MHg', 'MCg', 'cl_1', 'cl_2', 'cl_5', 'cl_10', 'cl_20', 'cl_30', 'cl_60', 'YTD', 'vlchg', 'vlrk', 'h_p_t', 'warn', 'zf', 'zfr']

	etf_df=list2dataframe(new_etfdf_list,etfdf_columns)

	etf_df = etf_df.drop(etf_df[etf_df['b_vl'] == 1000000].index)

	#etf_df.sort_values(by=[rankingkeyword],ascending=False,inplace=True)

	selectedrows=filter_etf_with_tradingmoney(etf_df,'code',datenow)  #会把成交量低的ETF干掉。

	return selectedrows


def get_target_etflist_closechange_data(etflist,total_money,datenow):

	source_df=daily_all_jqetf_monitor(total_money,datenow)  #会把成交量低的etf删除掉！

	selected_df = source_df.loc[source_df['code'].isin(etflist)]  #找出符合etflist里面的代码的所有行。

	rows = selected_df.copy()

	rows.reset_index(drop=True,inplace=True)

	return rows

def get_special_etf_df(source_df,spe_number): #获取特殊的etf,比如513开始的就是T + 0的etf

	target_code_list=list(source_df['code'])

	final_code=[code for code in target_code_list if code.startswith(str(spe_number))]

	final_df=source_df.loc[source_df['code'].isin(final_code)]

	return final_df


def get_target_etflist_buying_data(etflist,total_money,datenow):

	source_df=daily_all_etf_monitor(total_money,datenow)

	selected_df = source_df.loc[source_df['code'].isin(etflist)]  #找出符合etflist里面的代码的所有行。

	rows = selected_df.copy()

	#取消下面这一行，因为可以冻结所有列了。

	#rows.set_index('code',inplace=True)

	rows['byvaue'] = rows['b_vl']*rows['close']*100

	rows['b_vl'] = rows['b_vl'].astype(int)

	#以下这两行暂时不用。


	df_number=len(rows)

	rows.loc[df_number,'byvaue']=rows['byvaue'].sum()  #总计购买金额


	#new_df=newdf.sort_values(by=[rankingkeyword],ascending=False,inplace=False)  #不改变原来的df,重新生成一个新的。

	#dataframe_ranking(new_df,'vlchg')   #SettingWithCopyWarning 这个错误出现，所以不用。

	#new_df.reset_index(drop=True,inplace=True)

	return rows


def dataframe_ranking(df,column): #排序的同时，替换了原有的datafrmae,inplace=True

	df.sort_values(by=[column],ascending=False,inplace=True)

	#df.reset_index(inplace=True)   #每次排序后，将索引重新生成。

	return df


def dataframe2list(source_df): #将dataframe转化成list的数据格式，一遍本地重新生成dataframe

	df_list=source_df.values.tolist()

	column_value=source_df.columns.tolist()

	columns_list=list(column_value)

	return df_list,column_value   #得到了dataframe的数据和列的数据，都是list的形式。


def list2dataframe(df_list,columns_list): #将list形式的数据和列转换成dataframe

	df=pd.DataFrame(df_list,columns=columns_list)

	return df





index_list=[['sh000688', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000682', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan],['sh000685', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan],['sh000852', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399905', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000016', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000097', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000805', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399006', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399004', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399005', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399300', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399330', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399333', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399612', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399673', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399903', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000922', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399324', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000812', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000811', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000810', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000809', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000813', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000814', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000815', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000816', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000818', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000819', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000827', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000901', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000928', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000929', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000930', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000931', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399932', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000933', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000934', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000935', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000941', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000977', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000978', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000979', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000986', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000987', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000988', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000989', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000990', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000991', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000992', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000993', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000998', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399030', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399060', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399295', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399296', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399393', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399394', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399395', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399396', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399412', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399417', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399419', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399420', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399423', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399432', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399435', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399438', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399439', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399440', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399441', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399608', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399610', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399646', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399647', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399653', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399654', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399669', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399674', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399675', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399676', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399677', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399683', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399684', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399687', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399693', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399695', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399699', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399704', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399705', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399706', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399804', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399805', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399806', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399809', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399807', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399808', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399810', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399811', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399812', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399813', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399814', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399967', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399970', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399971', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399973', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399975', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399976', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399986', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399987', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399989', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399993', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399994', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399995', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399996', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399997', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399998', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399966', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan]]
index_columns=['code', 'name', 'close', 'cl_1', 'cl_2','cl_5','cl_10', 'cl_20', 'cl_30',  'cl_60', 'YTD']


def index_monitor(datenow):

	index_df=pd.DataFrame(index_list,columns=index_columns)

	code_tuple=tuple(index_df.code)

	limit_number=len(code_tuple)

	i=0

	while i < limit_number:
	    index_code=code_tuple[i]
	    target_position=np.where(index_df['code']==index_code)[0][0]
	    my_jq=Jqdata(index_code)
	    
	    present_close,b,c,d=my_jq.get_point()
	    index_df.loc[target_position,'name']=my_jq.security_name()
	    index_df.loc[target_position,'close']=present_close

	    

	    index_df.loc[target_position,'cl_1']=my_jq.get_close_change(datenow,1)
	    index_df.loc[target_position,'cl_2']=my_jq.get_close_change(datenow,2)
	    index_df.loc[target_position,'cl_5']=my_jq.get_close_change(datenow,5)
	    index_df.loc[target_position,'cl_10']=my_jq.get_close_change(datenow,10)
	    index_df.loc[target_position,'cl_20']=my_jq.get_close_change(datenow,20)
	    index_df.loc[target_position,'cl_30']=my_jq.get_close_change(datenow,30)
	    index_df.loc[target_position,'cl_60']=my_jq.get_close_change(datenow,60)
	    index_df.loc[target_position,'YTD']=my_jq.get_close_change_YTD(target_date()) 

	    index_df.loc[target_position,'money']=my_jq.get_present_money(datenow)
	    index_df.loc[target_position,'data_time']=datenow
	    
	    i += 1

	index_df.sort_values(by=['cl_5'],ascending=False,inplace=True)

	index_df.drop(index_df[pd.isna(index_df['close'])].index,inplace=True) 

	indexdf_list,indexdf_columns=dataframe2list(index_df)



	print(''.join(['indexdf_list='+str(indexdf_list)]))

	print(''.join(['indexdf_columns='+str(indexdf_columns)]))

	return indexdf_list,indexdf_columns



def freeze_columns(source_df):   #这个函数就是冻结所有的列，让我们方便的阅读数据。

	source_df.reset_index(drop=True,inplace=True)  #先重设索引。


	final_df=source_df.style.set_table_styles(
	    [{'selector': 'thead th','props': [
	          ('position', 'sticky'),('top', '0'),('background-color', 'yellow'),
	               ]}]
	    )

	

	return final_df


#以下是上面这个函数的不同pandas版本的代码解决同样的问题。


'''
https://github.com/pandas-dev/pandas/issues/29072
 
the new 1.3.0 input format for release June 2021, which is more CSS friendly.


pd.DataFrame(np.random.randn(200, 50)).style.set_table_styles([
    {'selector': 'thead th', 'props': 'position: sticky; top:0; background-color:red;'},
    {'selector': 'tbody th', 'props': 'position: sticky; left:0; background-color:green;'}  
])
'''

def target_index_watch(target_index_list,datenow):

	if datenow==datetime.datetime.now().strftime('%Y-%m-%d'):

		date_time=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  #如果要看的是当前日期的，就这么写，可以精确到具体时间。

	else:

		date_time=datenow

	print('数据日期：{}'.format(date_time))



	index_performace_info=[get_single_index_data(index_code,datenow) for index_code in target_index_list]

	index_performace_columns=['code','name','close','MHg','MCg','cl_1','cl_2','cl_5','cl_10','cl_20','cl_30','cl_60','YTD','vlchg','vlrk','h_p_t','warn','zf','zfr']


	index_df=list2dataframe(index_performace_info,index_performace_columns)

	index_df.drop(index_df[pd.isna(index_df['close'])].index,inplace=True) 

	index_df.sort_values(by=['cl_5'],ascending=False,inplace=True)

	index_df.reset_index(drop=True,inplace=True)

	index_df['name']=index_df['name'].apply(remove_string_from_index)


	return index_df





#以下这个函数其实也只能大概是用作来动态调整仓位，主要是针对etf的，因为keyword根据cl_10,cl_20,等数据来的，这些数据都是在变化中的，所以可以根据数据的动态变化来调整仓位，目前我是选择cl_10,因为谁也看不了这么远。

def index_buying_quota(target_index_list,datenow,totalmoney,keyword): #根据每个指数的涨幅占到总涨幅的比重来决定指数买多少金额。指数的涨幅选择根据keyword来定。

	index_df=target_index_watch(target_index_list,datenow)

	index_number=len(target_index_list)

	total_YTD=index_df[keyword].sum()   #这个keyword 可以选择cl_1,cl_5,cl_10,YTD的涨幅来选择每个指数应该买多少份额，

	index_df['weight']=round(index_df[keyword]/total_YTD,4)*100  #看看每个指数的涨幅占比是多少百分比

	index_df['by_money']=index_df['weight']*totalmoney/100  #百分比要除掉100.

	index_df['by_money']=index_df['by_money'].apply(make_column_int)

	index_df['name']=index_df['name'].apply(remove_string_from_index)

	#取消下面这行，因为冻结所有列可以做到了。

	#index_df.set_index('code',inplace=True)


	return index_df


def operation_on_holding_etf(target_etf_list, totalmoney,datenow,keyword): #根据目前持有的etf的市场表现，跟etf_buying_quota类似，来重新计算每个etf应该买的金额。

	holdingstock_df=check_holding_sec(sec_list,sec_columns,totalmoney,datenow)   #这几个就是常量了，在前面会有设置。

	allmoney=float(holdingstock_df['value'].nlargest(1))   #把这个数据值提出来，而不是从dataframe里面选择。找出一列中最大的值，提取出来。

	etf_quota=etf_buying_quota(target_etf_list,allmoney,datenow,keyword)   #allmoney,是目前持有的总市值，根据这个总市值，来重新分配每个etf的配额。

	etf_quota.reset_index(inplace=True)   #如下两行是将索引值变成 一个列。

	etf_quota.rename(columns={'code': 'code'}, inplace=True)   

	newdf=etf_quota.iloc[:,[0,1,19]]  #只剩下几列需要的数据。

	print(newdf)

	#底下这行效果不太好。会出错。

	#merged_df = pd.merge(newdf, holdingstock_df, on='code', how='left')  #将具有相同的列的值的行的两个pd,合并，根据code 合并，这样，再去去掉不要的数据。

	merged_df=get_same_rows_from_dataframe(newdf,holdingstock_df,'code')

	print(merged_df)

	if not merged_df.empty:

		finaletf_df=merged_df.iloc[:,[0,1,2,10]]

		etf_df=finaletf_df.fillna(0, inplace=False)  #将没有数据的地方，换成0.

		print(etf_df)

		etf_df['operation']=etf_df['by_money']-etf_df['value']

		etf_df['operation']=etf_df['operation'].apply(make_column_int)

	else:

		etf_df=pd.DataFrame()



	#selected_df = source_df.loc[source_df['code'].isin(etflist)]



	return etf_df



index_list=[['sh000688', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan],['sz399965', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan],['sh000685', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan],['sh000682', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000852', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399905', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000016', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000097', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000805', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399006', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399004', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399005', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399300', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399330', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399333', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399612', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399673', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399903', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000922', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399324', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000812', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000811', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000810', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000809', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000813', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000814', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000815', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000816', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000818', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000819', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000827', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000901', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000928', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000929', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000930', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000931', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399932', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000933', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000934', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000935', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000941', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000977', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000978', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000979', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000986', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000987', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000988', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000989', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000990', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000991', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000992', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000993', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sh000998', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399030', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399060', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399295', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399296', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399393', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399394', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399395', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399396', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399412', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399417', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399419', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399420', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399423', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399432', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399435', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399438', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399439', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399440', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399441', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399608', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399610', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399646', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399647', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399653', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399654', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399669', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399674', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399675', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399676', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399677', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399683', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399684', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399687', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399693', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399695', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399699', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399704', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399705', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399706', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399804', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399805', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399806', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399809', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399807', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399808', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399810', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399811', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399812', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399813', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399814', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399967', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399970', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399971', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399973', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399975', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399976', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399986', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399987', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399989', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399993', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399994', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399995', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399996', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399997', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399998', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan], ['sz399966', nan, nan, nan, nan, nan, nan, nan, nan, nan, nan]]
index_columns=['code', 'name', 'close', 'cl_1', 'cl_2','cl_5','cl_10', 'cl_20', 'cl_30',  'cl_60', 'YTD']


def get_single_index_data(index_code,datenow):

	my_jq=Jqdata(index_code)

	index_name=my_jq.security_name()

	max_high,max_close=my_jq.get_max_high_close(datenow)

	present_close,b,c,d=my_jq.get_point()

	present_max_high_change,present_max_low_change=my_jq.get_present_high_low_change(datenow)

	index_MHg=round((max_high - present_close)/max_high,4)*100  #15天内，现价距离最高点的回撤百分比

	index_MCg=round((max_close - present_close)/max_close,4)*100 #15天内，现价距离最高收盘价的回撤百分比

	index_cl_changes = [round(my_jq.get_close_change(datenow, i), 2) for i in [1, 2, 5, 10, 20, 30, 60]]

	index_cl_YTD=round(my_jq.get_close_change_YTD(target_date()),2)

	index_vlchg=my_jq.get_volume_change(datenow)

	index_vlrk=my_jq.get_volume_ranking(datenow)

	index_h_p_t=my_jq.find_stock_highest_price_time(datenow)

	index_warn=my_jq.judge_stock_highest_price_time_risk(datenow)

	index_zf=my_jq.get_zhenfu(datenow)

	index_zfr=my_jq.get_zhenfu_ranking(datenow)

	if my_jq.is_doji(datenow):

		index_cross='yes'

	else:

		index_cross=''

	if my_jq.judge_stock_up_trend(datenow):

		duotou='yes'

	else:

		duotou=''

	index_info_list = [index_code,index_name,present_close,present_max_high_change,present_max_low_change,index_MHg, index_MCg] + index_cl_changes + [index_cl_YTD, index_vlchg, index_vlrk, index_h_p_t, index_warn, index_zf, index_zfr,index_cross,duotou]


	return index_info_list


def filter_allindex(end_day): #获取所有除了债相关的指数。

	allindex=get_all_index(end_day)

	allindex_list=[x for x in allindex if '债' not in Jqdata(x).security_name()]

	final_allindex=[transform_code_tecent(x) for x in allindex_list]  #因为指数的代码如果没有前缀和后缀的话，容易跟股票代码混乱，所以，不要用transform_to_normal_code变化。

	#以下是为了排除里面重复的指数代码，比如399905,000905 是同一个指数，应该排除一个，考虑到这种情况只发生在几个指数上，并且用以下的功能，会排除掉指数名称相同，实际不是同一指数的情况，所以还是不做筛选了。


	#index_df_list=[[x,Jqdata(x).security_name()] for x in final_allindex]

	#index_df_columns=['indexcode','indexname']

	#indexdf=list2dataframe(index_df_list,index_df_columns)

	#indexdf.drop_duplicates(subset='indexname', keep='first', inplace=True)

	#all_index_list=list(indexdf['indexcode'])

	return final_allindex


def indexlist_monitor(indexlist,datenow,rankingkeyword):

	
	new_indexdf_list=[get_single_index_data(index_code,datenow) for index_code in indexlist]


	indexdf_columns=['code', 'name','close','Lup','Ldown','MHg', 'MCg', 'cl_1', 'cl_2', 'cl_5', 'cl_10', 'cl_20', 'cl_30', 'cl_60', 'YTD', 'vlchg', 'vlrk', 'h_p_t', 'warn', 'zf', 'zfr','cross','duot']

	index_df=list2dataframe(new_indexdf_list,indexdf_columns)

	index_df['code']=index_df['code'].apply(transform_to_normal_code)

	index_df['name']=index_df['name'].apply(remove_string_from_index)

	index_df.sort_values(by=[rankingkeyword],ascending=False,inplace=True)

	return index_df


def daily_index_monitor(datenow,rankingkeyword):  #这是线上的实时的监测index的数据。监测的是所有的index的市场表现。会把成交量低的index干掉。

	indexlist=filter_allindex(datenow)

	indexdf=indexlist_monitor(indexlist,datenow,rankingkeyword)

	return indexdf

def get_target_high_low_info(targetlist,start_day,end_day):

	new_targetdf_list=[]

	for target_code in targetlist:

		myjq=Jqdata(target_code)

		target_name=myjq.security_name()

		presentclose=myjq.get_present_close(end_day)

		close_change=myjq.get_close_change(end_day,1) #当天的涨跌幅。

		present_low=myjq.get_present_low(end_day)  #获取当天最低点的价格。

		max_high,min_low,highest_price_date,lowest_price_date=myjq.get_max_high_min_low(start_day,end_day)

		target_public_date=myjq.get_security_public_date()

		highest_gap = round((max_high - min_low) / max_high, 4) * 100

		present_gap = round((max_high - present_low) / max_high, 4) * 100

		two_gap_difference=highest_gap - present_gap

		target_code=transform_to_normal_code(target_code)

		possible_down=round((presentclose - min_low)/presentclose, 4) * 100

		single_target_info=[target_code,target_name,presentclose,close_change,target_public_date,max_high,highest_price_date,min_low,lowest_price_date,present_low,highest_gap,present_gap,two_gap_difference,possible_down]

		new_targetdf_list.append(single_target_info)


	targetdf_columns=['code','name','现价','cl_1','上市日','high','最高日','low','最低日','当日最低','最大跌幅','当日最低跌幅','跌幅差距','可能下跌']

	target_df=list2dataframe(new_targetdf_list,targetdf_columns)


	return target_df


def get_index_high_low_data(start_day,end_day):

	indexlist=filter_allindex(end_day)

	target_df=get_target_high_low_info(indexlist,start_day,end_day)

	target_df['name']=target_df['name'].apply(remove_string_from_index)  #只针对指数的名字做如此操作。

	return target_df


def get_index_stock_high_low_data(index_code,start_day,end_day):

    stocklist=Jqdata(index_code).get_index_securities() 

    target_df=get_target_high_low_info(stocklist,start_day,end_day)

    return target_df

def get_allstock_high_low_data(start_day,end_day):

	stocklist=get_all_sec(end_day)

	target_df=get_target_high_low_info(stocklist,start_day,end_day)

	return target_df


def get_etf_high_low_data_from_jq(start_day,end_day):

	etflist=get_all_etf(end_day)

	target_df=get_target_high_low_info(etflist,start_day,end_day)

	selectedrows=filter_etf_with_tradingmoney(target_df,'code',present_day) 

	return selectedrows


def get_etf_high_low_data(start_day,end_day):

	new_etfdf_list=[]

	for single_etf_list in etfdf_list:

		etf_code=single_etf_list[0]

		myjq=Jqdata(etf_code)

		presentclose=myjq.get_present_close(end_day)

		present_close_change=myjq.get_close_change(end_day,1)

		present_low=myjq.get_present_low(end_day)

		max_high,min_low,highest_price_date,lowest_price_date=myjq.get_max_high_min_low(start_day,end_day)

		etf_public_date=myjq.get_security_public_date()

		highest_gap = round((max_high - min_low) / max_high, 4) * 100

		present_gap = round((max_high - present_low) / max_high, 4) * 100

		two_gap_difference=highest_gap - present_gap

		possible_down=round((presentclose - min_low)/presentclose, 4) * 100

		single_etf_info=single_etf_list+[presentclose,present_close_change,etf_public_date,max_high,highest_price_date,min_low,lowest_price_date,present_low,highest_gap,present_gap,two_gap_difference,possible_down]

		new_etfdf_list.append(single_etf_info)

	etfdf_columns=['code','name','index','现价','cl_1','上市日','high','最高日','low','最低日','当日最低','最大跌幅','当日最低跌幅','跌幅差距','可能下跌']

	etf_df=list2dataframe(new_etfdf_list,etfdf_columns)

	selectedrows=filter_etf_with_tradingmoney(etf_df,'code',present_day) 

	return selectedrows

def select_big_rows(source_df,big_number): #选择最大跌幅超过比如70的行

    selected_df=source_df[source_df['最大跌幅']>big_number]

    return selected_df

def select_worst_rows(source_df,small_number): #再从上面里面选择最大跌幅和目前跌幅差距小于比如15的行。

    selected_df=source_df[source_df['跌幅差距']<small_number]

    return selected_df


def select_vip_target(source_df,big_number,small_number):

	selected_df=select_big_rows(source_df,big_number)

	selected_rows=select_worst_rows(selected_df,small_number)

	return selected_rows


def get_single_stock_data(stockcode,total_money,datenow):

	yesterday_date=get_yesterday_date(datenow)

	my_jq=Jqdata(stockcode)

	stock_name=my_jq.security_name()

	stock_industry=my_jq.security_industry(datenow)

	max_high,max_close=my_jq.get_max_high_close(datenow)

	present_close,b,c,d=my_jq.get_point()

	stock_MHg=round((max_high - present_close)/max_high,4)*100  #15天内，现价距离最高点的回撤百分比

	stock_MCg=round((max_close - present_close)/max_close,4)*100 #15天内，现价距离最高收盘价的回撤百分比

	stock_b_vl=my_jq.buy_single_stock_volume(total_money,datenow)

	stock_cl_changes=[round(my_jq.get_close_change(datenow, i), 2) for i in [1, 2, 5, 10, 20, 30, 60]]

	stock_cl_YTD=round(my_jq.get_close_change_YTD(target_date()),2)

	stock_vlchg=my_jq.get_volume_change(datenow)

	stock_vlrk=my_jq.get_volume_ranking(datenow)

	stock_h_p_t=my_jq.find_stock_highest_price_time(datenow)

	stock_warn=my_jq.judge_stock_highest_price_time_risk(datenow)

	stock_zf=my_jq.get_zhenfu(datenow)

	stock_zfr=my_jq.get_zhenfu_ranking(datenow)

	stock_marketsize=get_sec_marketsize(stockcode,yesterday_date)

	stock_liudong_size=get_sec_liutong_size(stockcode,yesterday_date)

	stockcode=transform_to_normal_code(stockcode)   #这里是做了代码正规化处理，变成了6位。

	if my_jq.is_doji(datenow):

		cross='yes'  #表明是十字星.

	else:

		cross=''

	if my_jq.judge_stock_up_trend(datenow):

		duotou='yes'

	else:

		duotou=''


	stock_info_list = [stockcode,stock_name,stock_industry,stock_b_vl,stock_MHg, stock_MCg] + stock_cl_changes + [stock_cl_YTD, stock_vlchg, stock_vlrk, stock_h_p_t, stock_warn, stock_zf, stock_zfr,stock_marketsize,stock_liudong_size,cross,duotou]


	return stock_info_list



def get_target_stocklist_closechange_data(stockslist,total_money,end_day):   #这个函数是放弃了股票市场和流通市值的数据。为了实时的监测自己想要观察的股票数据。

    #stockslist=get_all_sec(end_day)

    #获取当天，2天，5天，10天，20天，30天，60天，和今年到现在为止，该股票的涨幅数据。

    #这里的不同点就是代码不用变化了，因为输入的就是sh,sz开始的代码。


    stockdf_list=[get_single_stock_data(stock_code,total_money,end_day) for stock_code in stockslist]


    df_columns=['code', 'name', 'industry', 'b_vl','MHg','MCg', 'cl_1', 'cl_2', 'cl_5', 'cl_10', 'cl_20', 'cl_30', 'cl_60', 'YTD', 'vlchg', 'vlrk', 'h_p_t', 'warn', 'zf', 'zfr','m_size','ld_size','cross','duot']


    df=list2dataframe(stockdf_list,df_columns)

    #去掉数据为i 空的那些行


    df.drop(df[pd.isna(df['cl_1'])].index,inplace=True)
    df.drop(df[pd.isna(df['cl_2'])].index,inplace=True)
    df.drop(df[pd.isna(df['cl_5'])].index,inplace=True)
    df.drop(df[pd.isna(df['cl_10'])].index,inplace=True)
    df.drop(df[pd.isna(df['cl_20'])].index,inplace=True)
    df.drop(df[pd.isna(df['cl_30'])].index,inplace=True)
    df.drop(df[pd.isna(df['cl_60'])].index,inplace=True)
    df.drop(df[pd.isna(df['YTD'])].index,inplace=True)

    #df.reset_index(drop=True,inplace=True)

    #etf_df.loc[etf_df['h_p_t'] < datetime.time(10,0,0),'warn'] = '危'

    #df.reset_index(drop=True,inplace=True)


    #df_list,df_columns=dataframe2list(df)

    dataframe_ranking(df,'vlchg')

    #df=filter_stock_with_tradingmoney(df,'code',end_day)  #过去20日成交量低于1500W的排除掉。

    df.reset_index(drop=True,inplace=True)

    stockinfo_list=[[Jqdata(x).security_name()]+list(Jqdata(x).get_stock_point(end_day)) for x in stockslist]

    stockinfo_columns=['name','present','open','high','low']

    stockinfo_df=list2dataframe(stockinfo_list,stockinfo_columns)

    warningrows=stockinfo_df[stockinfo_df['open']==stockinfo_df['high']]

    open_equal_high=list(warningrows['name'])


    #print(df)

    if end_day==datetime.datetime.now().strftime('%Y-%m-%d'):

        date_time=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  #如果要看的是当前日期的，就这么写，可以精确到具体时间。

    else:

        date_time=end_day

    print('数据日期：{}'.format(date_time))

    print('\n *****  开盘价等于最高价的股票有 当心，考虑 减仓 清仓！*********   {} '.format(open_equal_high))

    return df


def get_codelist_from_df(source_df):

	codelist=list(source_df['code'])

	if nan in codelist:

		codelist.remove(nan)

	codelist=list(set(codelist))  #去掉里面重复的代码。

	return codelist


def get_index_stock_closechange_data(index_code,total_money,end_day):

	indexstock_codelist=Jqdata(index_code).get_index_securities()

	index_stock_codelist=[transform_to_normal_code(x) for x in indexstock_codelist]

	indexstock_closechange_df=get_target_stocklist_closechange_data(index_stock_codelist,total_money,end_day)

	return indexstock_closechange_df


def get_allsec_closechange_data(total_money,end_day):   #实时的监测全市场所有的股票数据。因为要实时，所以市值和流动市值，是不能算进去的，因为不能马上算出来。就是把这个函数get_target_stocklist_closechange_data改动改动.

	allsec_list=get_all_sec(end_day)

	allsecdf=get_target_stocklist_closechange_data(allsec_list,total_money,end_day)

	return allsecdf


def get_single_sec_closedata_indicator(sec_code,total_money,end_day):  #快速查一个单独股票的相关信息。

	code_list=[sec_code]  #因为下面两个函数只针对list发生作用。

	closedata_df=get_target_stocklist_closechange_data(code_list,total_money,end_day)

	indicator_df=Jq_codelist(code_list).check_stock_indicator_stocklist(total_money,end_day)  #均线考察5天线。

	return closedata_df,indicator_df



def get_sec_marketsize(seccode,end_day): #一般不能获取当天的市值，因为老在变，并且出错。所以要获取前一天的。

	code=Jqdata(seccode).jqcode

	q = query (valuation).filter(valuation.code == code)

	df = get_fundamentals(q,end_day)


	market_size=int(df['market_cap'][0])

	#liutong_size=df['circulating_market_cap'][0]

	#exchange_rate=df['turnover_ratio'][0]

	return market_size


def get_sec_liutong_size(seccode,end_day): #一般不能获取当天的市值，因为老在变，并且出错。所以要获取前一天的。

	code=Jqdata(seccode).jqcode

	q = query (valuation).filter(valuation.code == code)

	df = get_fundamentals(q,end_day)


	#market_size=df['market_cap'][0]

	liutong_size=int(df['circulating_market_cap'][0])

	#exchange_rate=df['turnover_ratio'][0]

	return liutong_size


index_list=['sh000001', 'sh000685','sh000682','sz399965',
  'sh000688', 'sh000852', 'sz399905', 'sh000016', 'sh000097', 'sh000805', 'sz399006', 'sz399004', 'sz399005', 'sz399300', 'sz399330', 'sz399333', 'sz399612', 'sz399673', 'sz399903', 'sh000922',
 'sz399324', 'sh000812', 'sh000811', 'sh000810', 'sh000813', 'sh000814', 'sh000815', 'sh000816', 'sh000818', 'sh000819', 'sh000827', 'sh000901', 'sh000928', 'sh000929', 'sh000930', 'sh000931', 
 'sz399932', 'sh000933', 'sh000934', 'sh000935', 'sh000941', 'sh000977', 'sh000978', 'sh000979', 'sh000986', 'sh000987', 'sh000988', 'sh000989', 'sh000990', 'sh000991', 'sh000992', 'sh000993', 'sh000998', 
 'sz399030', 'sz399060', 'sz399295', 'sz399296', 'sz399393', 'sz399394', 'sz399395', 'sz399396', 'sz399412', 'sz399417', 'sz399419', 'sz399420', 'sz399423', 'sz399432', 'sz399435', 'sz399438', 'sz399439',
  'sz399440', 'sz399441', 'sz399608', 'sz399610', 'sz399646', 'sz399647', 'sz399653', 'sz399654', 'sz399669', 'sz399674', 'sz399675', 'sz399676', 'sz399677', 'sz399683', 'sz399684', 'sz399687', 'sz399693', 
  'sz399695', 'sz399699', 'sz399704', 'sz399705', 'sz399706', 'sz399804', 'sz399805', 'sz399806', 'sz399809', 'sz399807', 'sz399808', 'sz399810', 'sz399811', 'sz399812', 'sz399813', 'sz399814', 'sz399967', 
  'sz399970', 'sz399971', 'sz399973', 'sz399975', 'sz399976', 'sz399986', 'sz399987', 'sz399989', 'sz399993', 'sz399994', 'sz399995', 'sz399996', 'sz399997', 'sz399998','sz399966',]






def get_index_weights_df(index_list):

	code_df_list_dic={}

	limit_number=len(index_list)

	#print(index_list)

	#print(len(index_list))

	i=0

	while i < limit_number:

		index_code=index_list[i]

		my_jq=Jqdata(index_code)

		#print(my_jq.jqcode)

		index_weight_df=get_index_weights(my_jq.jqcode)

		index_weight_df['code']=index_weight_df.index

		index_weight_df['code']=index_weight_df['code'].apply(transform_code)

		dataframe_ranking(index_weight_df,'weight')

		#print(index_weight_df)

		stock_info_list,stock_info_columns=dataframe2list(index_weight_df)

		code_df_list_dic.setdefault(index_code,stock_info_list)

		i += 1

	print(''.join(['index_weight_df_list='+str(code_df_list_dic)]))

	#for key,value in code_df_list_dic.items():

		#new_list=[key,value]

		#print(new_list,',')

		#time.sleep(0.005)


	print(''.join(['index_weight_df_columns='+str(stock_info_columns)]))

	return code_df_list_dic,stock_info_columns


def find_top_target_from_df(source_df,target_percent): #决定短期看5，10天，中期看20，30天，长期不看了！

	df=source_df

	target_sec_number=int(len(df)*target_percent)     #根据比例选择出来，比较强势的股票。

	#newdf_2=dataframe_ranking(df,'cl_2').head(n=target_sec_number)     #通过2，5，10天涨幅排名，选择出短期强势的股票。的代码

	#codelist_2=list(newdf_2['code'])

	newdf_5=dataframe_ranking(df,'cl_5').head(n=target_sec_number)

	codelist_5=list(newdf_5['code'])

	newdf_10=dataframe_ranking(df,'cl_10').head(n=target_sec_number)

	codelist_10=list(newdf_10['code'])

	#short_strong_code_list=[code for code in codelist_2 if code in codelist_5 and code in codelist_10]

	short_strong_code_list=[code for code in codelist_5 if code in codelist_10]

	newdf_20=dataframe_ranking(df,'cl_20').head(n=target_sec_number)     #通过20，30，60天涨幅排名，选择出长期期强势的股票。的代码

	codelist_20=list(newdf_20['code'])

	newdf_30=dataframe_ranking(df,'cl_30').head(n=target_sec_number)

	codelist_30=list(newdf_30['code'])

	#newdf_60=dataframe_ranking(df,'cl_60').head(n=target_sec_number)

	#codelist_60=list(newdf_60['code'])

	#long_strong_code_list=[code for code in codelist_20 if code in codelist_30 and code in codelist_60]

	long_strong_code_list=[code for code in codelist_20 if code in codelist_30]

	best_code_list=[code for code in short_strong_code_list if code in long_strong_code_list]

	return short_strong_code_list,long_strong_code_list,best_code_list


def find_top_sec_from_past_performance(total_money,ma_days,end_day,target_percent):  #source_df是连涨两天的股票名单，当然是基于截止日那天的数据。根据比较选择出来一定比例的股票。这里设置比例。 根据短期2，5，10天，和长期20，30，60的标准选择出来短期，长期，和短期长期都好的股票列表。

	
	above_ma_stock_list,source_df=find_stock_in_trend_onsite(total_money,ma_days,end_day)

	df=source_df

	short_strong_code_list,long_strong_code_list,best_code_list=find_top_target_from_df(df,target_percent)

	return short_strong_code_list,long_strong_code_list,best_code_list


#找出某个指数下面比较强的股票名单。跟上面比较相似，只是用的是指数成分股，作为原始股票池。

def find_top_sec_single_index(index_code,total_money,end_day,target_percent):  #找出某个指数下面的最好的股票，通过2，5，10，和 20，30，60长期筛选出来。

	
	index_stock_code_list=Jqdata(index_code).get_index_securities() #得到目标指数代码的成分股。

	index_stock_code_list_transformed=[Jqdata(code).transform_code(code) for code in index_stock_code_list]  #转换成腾讯格式代码。

	df=get_target_stocklist_closechange_data(index_stock_code_list_transformed,total_money,end_day)

	short_strong_code_list,long_strong_code_list,best_code_list=find_top_target_from_df(df,target_percent)

	return short_strong_code_list,long_strong_code_list,best_code_list



def find_top_sec_from_df(source_df,target_percent):  #source_df必须是包含各种市场表现数据的，否则，就弄不出来下面这个选择过程。选择出最优的。

	
	#above_ma_stock_list,source_df=find_stock_in_trend_onsite(total_money,ma_days,end_day)

	df=source_df

	short_strong_code_list,long_strong_code_list,best_code_list=find_top_target_from_df(df,target_percent)

	selected_df = df.loc[df['code'].isin(best_code_list)] 

	return selected_df


def transform_code_tecent(code_transformed):  #把代码换成腾讯财经认可的代码。sh,sz开头的。
	codeRegex=re.compile(r'(\d{6}).(\w{4})')
	code_part=codeRegex.search(code_transformed)[1]
	character_part=codeRegex.search(code_transformed)[2]

	normal_code=None

	if character_part =='XSHG':
		normal_code='sh'+code_part
	else:
		normal_code='sz'+code_part

	return normal_code


def transform_tecent_code(code):  #将腾讯财经的代码转换为正常的6位代码。

	codeRegex=re.compile(r'(\w{2})(\d{6})')

	code_part=codeRegex.search(code)[2]

	return code_part


def transform_jqdata_code(code):

	sec_code=Jqdata(code).transform_code(code)

	return sec_code


def transform_to_normal_code(code):

	if len(code)==11:

		code=transform_jqdata_code(code)  #这种情况下就是jq的11位码

	

	elif len(code)==8:  #sh,sz开头或的情况

		code=transform_tecent_code(code)

	else:

		code=code


	return code


def get_sec_buying_volume(stockcode,total_money,end_day):

	my_jq=Jqdata(stockcode)

	if stockcode.startswith('6') or stockcode.startswith('0') or stockcode.startswith('3'):

		buying_volume=my_jq.buy_single_stock_volume(total_money,end_day)

	else:

		buying_volume=my_jq.buy_single_etf_volume(total_money,end_day)

	return buying_volume


def get_sec_industry(stockcode,end_day):

	my_jq=Jqdata(stockcode)

	if stockcode.startswith('6') or stockcode.startswith('0') or stockcode.startswith('3'):

		stock_industry=my_jq.security_industry(end_day)

	else:

		stock_industry='None'

	return stock_industry


def get_sec_name(stockcode):

	sec_name=Jqdata(stockcode).security_name()

	return sec_name


def get_present_stock_price(stockcode,end_day):

	stock_price=Jqdata(stockcode).get_present_close(end_day)

	return stock_price


def check_holding_sec(sec_list,sec_columns,total_money,present_day):

	my_stock_df=list2dataframe(sec_list,sec_columns)

	my_stock_df['code']=my_stock_df['code'].apply(transform_tecent_code) #将df里面的code都转换成正常的。

	my_stock_df['volume']=my_stock_df['volume'].apply(lambda x: int(round(x/100)))

	del my_stock_df['break']

	my_stock_df['name']=my_stock_df['code'].apply(get_sec_name)

	my_stock_df['industry']=my_stock_df['code'].apply(get_sec_industry,end_day=present_day)

	my_stock_df['b_vl']=my_stock_df['code'].apply(get_sec_buying_volume,args=(total_money,present_day))

	my_stock_df['more_vlme']=my_stock_df['b_vl'] - my_stock_df['volume']

	my_stock_df['price']=my_stock_df['code'].apply(get_present_stock_price,end_day=present_day)

	my_stock_df['value']=my_stock_df['volume']*my_stock_df['price']*100  #因为前面为了简化，将数量后面去掉了100.

	total_value=my_stock_df['value'].sum()

	my_stock_df['weight']=round(my_stock_df['value']/total_value,4)*100

	my_stock_df['profit']=round(my_stock_df['volume']*(my_stock_df['price'] - my_stock_df['cost']),2)*100  #因为前面为了简化，将数量后面去掉了100.

	my_stock_df['prft_rate']=round((my_stock_df['price']-my_stock_df['cost'])/my_stock_df['cost'],4)*100

	row_number=len(my_stock_df)

	#print(row_number)

	my_stock_df.loc[row_number,'price']='总计:'

	my_stock_df.loc[row_number,'value']=total_value

	my_stock_df.loc[row_number,'profit']=my_stock_df['profit'].sum()

	new_columns=['code','name','cost','industry','volume','b_vl','more_vlme','price','value','weight','profit','prft_rate']

	new_df=my_stock_df.reindex(columns=new_columns)

	warning_df=my_stock_df[my_stock_df['prft_rate'] < -10]   #亏损超过10的要报警。

	warning_list=list(warning_df['name'])

	print('亏损超过10%的有 {} 个  {}\n'.format(len(warning_list),warning_list))

	caution_df=my_stock_df[my_stock_df['prft_rate'] < -5 ] #亏损大于5的有。

	caution_list=list(caution_df['name'])

	caution_list=[name for name in caution_list if name not in warning_list]  #这样就把亏损大于10的去掉了。

	print('亏损超过5%   小于10%的有 {} 个 {}\n'.format(len(caution_list), caution_list))

	return new_df


def get_industry_analysis_from_df(source_df):


	# 计算各行业数量
	industry_counts = source_df['industry'].value_counts() 

	# 求总数量
	total_count = len(source_df)

	# 计算比例  
	industry_ratios = (industry_counts / total_count * 100).round(4)

	# 构建结果DataFrame
	industry_df = pd.DataFrame({'industry': industry_counts.index, 
	                            'ratio': industry_ratios})  

	industry_df.reset_index(drop=True,inplace=True)

	return industry_df



def get_holding_stock_etf_data(holdingstockdf,total_money,present_day,select_ma):

	holding_code_list=get_codelist_from_df(holdingstockdf)

	yesterday_date=get_yesterday_date(present_day)

	holding_etf_df=get_target_etflist_closechange_data(holding_code_list,total_money,present_day) #可能会把持有的一部分ETF漏掉，因为ETF有个成交额删选的功能。

	holdingetf_code_list=list(holding_etf_df['code'])  #得到只有ETF代码的列表。

	#print(holdingetf_code_list)

	#找出目前持有的股票情况，看看今年的市场表现。。
	#holdingstock_code_list=list(holdingstockdf['code'])
	#holdingstock_code_list.remove(nan)
	#holdingstock_code_list=[x for x in holdingstock_code_list if x.startswith('6') or x.startswith('0') or x.startswith('3')] #获取股票的代码
	#holdingstock_code_list=[x for x in holding_code_list if x not in holdingetf_code_list]  #获取所有的股票代码。这么做会把一部分ETF搞进来。

	holdingstock_code_list=[x for x in holding_code_list if x.startswith('3') or x.startswith('0') or x.startswith('6')]

	#print(holdingstock_code_list)

	holding_stock_df=get_target_stocklist_closechange_data(holdingstock_code_list,total_money,present_day)

	holding_stock_df['marketsize']=holding_stock_df['code'].apply(get_sec_marketsize,end_day=yesterday_date)

	holding_stock_df['liutong_size']=holding_stock_df['code'].apply(get_sec_liutong_size,end_day=yesterday_date)

	holding_thing_indicator=Jq_codelist(holding_code_list).check_stock_indicator_stocklist(total_money,present_day)


	holding_stock_indicator=Jq_codelist(holdingstock_code_list).check_stock_indicator_stocklist(total_money,present_day)

	holding_etf_indicator=Jq_codelist(holdingetf_code_list).check_stock_indicator_stocklist(total_money,present_day)

	holdingstock_strategy=Trading(holdingstock_code_list).trading_strategy_for_stock(total_money,present_day,select_ma)

	holdingetf_strategy=Trading(holdingetf_code_list).trading_strategy_for_stock(holdingetf_code_list,total_money,present_day,select_ma)

	#获取持有股票的行业分类比例的相关数据。

	industry_df=get_industry_analysis_from_df(holding_stock_df)

	return industry_df,holding_etf_df,holding_stock_df,holding_thing_indicator,holding_stock_indicator,holding_etf_indicator,holdingstock_strategy,holdingetf_strategy

def index_stock_industry_analysis(index_code,present_day):

	indexstockdf=get_target_index_weights(index_code,present_day)

	indexstockdf['industry']=indexstockdf['code'].apply(get_sec_industry,end_day=present_day)

	newcolumns=['code', 'display_name', 'industry','weight', 'date', 'marketsize', 'liutong_size']

	indexstockdf=indexstockdf.reindex(columns=newcolumns)

	industry_df=get_industry_analysis_from_df(indexstockdf)

	return indexstockdf,industry_df



def find_stock_kind(holdingstockdf,keyword_list): #为了找出持仓的某一列含有一类关键词的行，比如证券公司，和证券行业成分股的公司，如下面的例子，东方财富，中信建投，中金公司，列在列表里，就很容易归纳出来。
	

	hhh=holdingstockdf[~holdingstockdf['price'].isin(['总计:'])] 

	keyword_df_list=[hhh[hhh['name'].str.contains(x)] for x in keyword_list]

	final_key_word_df=pd.concat(keyword_df_list)
	'''
	ggg=hhh[hhh['name'].str.contains('证券') ]
	xxx=hhh[hhh['name'].str.contains('中金公司')]
	yyy=hhh[hhh['name'].str.contains('建投')]
	zzz=hhh[hhh['name'].str.contains('龙')]
	bbb=hhh[hhh['name'].str.contains('财富')]
	ccc=hhh[hhh['name'].str.contains('华鑫')]
	finalxxx=pd.concat([ggg,xxx,yyy,zzz,bbb,ccc])
	'''

	return final_key_word_df



def find_good_bad_in_stocklist(stocklist,end_day):  #算上截止日连续两天上涨和下跌的股票名单。

	warning_stock_list=[code for code in stocklist if Jqdata(code).judge_stock_down(end_day)==True]

	warning_stock_name_list=[Jqdata(code).security_name() for code in warning_stock_list]

	goodtrend_stock_list=[code for code in stocklist if Jqdata(code).judge_stock_increase(end_day)==True]

	goodtrend_stock_name_list=[Jqdata(code).security_name() for code in goodtrend_stock_list]

	print('连续两天  下跌  的股票： {}\n'.format(warning_stock_name_list))

	print('连续两天  上涨  的股票：{}\n'.format(goodtrend_stock_name_list))

	return warning_stock_list,goodtrend_stock_list


def get_target_index_weights(target_index_code,end_day):  #获取指数成分股的数据。

	yesterday_date=get_yesterday_date(end_day)

	df=get_index_weights(Jqdata(target_index_code).jqcode,yesterday_date)

	df=df.reset_index() #将以代码为索引的变成列,重新建立索引.

	df=df.reindex(columns=['code','display_name','weight','date'])

	df['code']=df['code'].apply(transform_jqdata_code)

	df['marketsize']=df['code'].apply(get_sec_marketsize,end_day=yesterday_date)

	df['liutong_size']=df['code'].apply(get_sec_liutong_size,end_day=yesterday_date)

	dataframe_ranking(df,'weight')

	df.reset_index(drop=True,inplace=True)

	return df



def get_new_stock_for_index(target_index_code,change_date_1,change_date_2):#寻找两个不同日期指数成分股调整后的股票数据。

	df_1=get_target_index_weights(target_index_code,change_date_1)

	df_1_codelist=list(df_1['code'])

	df_2=get_target_index_weights(target_index_code,change_date_2)

	df_2_codelist=list(df_2['code'])

	new_added_stock_codelist=[x for x in df_2_codelist if x not in df_1_codelist]  #新增的股票代码列表。

	old_removed_stock_codelist=[x for x in df_1_codelist if x not in df_2_codelist]  #删去的旧的股票代码列表。

	same_stock_codelist=[x for x in df_2_codelist if x in df_1_codelist]  #两次调整成分股的时候都保留下来的共同的成分股

	return new_added_stock_codelist,old_removed_stock_codelist,same_stock_codelist


def find_same_stock_in_two_index(index1,index2):  #主要用来查询两个指数共同的股票，来看看两个指数的相关性。如果相关性太高，就每个只买一点，因为相关性太高！

	index1_stocks = get_index_stocks(Jqdata(index1).jqcode)

	index1_name=Jqdata(index1).security_name()

	index2_stocks = get_index_stocks(Jqdata(index2).jqcode)

	index2_name=Jqdata(index2).security_name()

	print('{} {} 有 {} 只成分股！'.format(index1,index1_name,len(index1_stocks)))

	print('{} {} 有 {} 只成分股！'.format(index2,index2_name,len(index2_stocks)))

	# 获取两个指数中相同的股票代码列表
	same_stocks = list(set(index1_stocks) & set(index2_stocks))

	print('两个指数共同有的成分股有 {} 个，这些股票代码是 {}'.format(len(same_stocks),same_stocks))

	return same_stocks



#以下这个函数是为了给某列数据涂红，好让你非常清楚得看到某列的数据，将背景涂红。

def highlight_column(s): #use the apply method of the style property to apply a custom function that highlights the cells in a specific column.I get this from chatgpt of new bing.


    return ['background-color: red' if s.name == target_column else '' for v in s]


# Define a function to highlight every other row  隔一行涂一行颜色。
def highlight_every_other_row(s):

    return ['background-color: green' if s.name % 2 == 0 else '' for _ in s]

# Apply the function to the dataframe
#df.style.apply(highlight_every_other_row, axis=1)


#如下这个函数跟上面功能相同，不同之处在于，上面的需要索引值，如果没有索引就不灵，而下面这个不需要索引，就可以隔行着色。

# Define a function to highlight every other row
def highlight_every_other_row(s):
    return ['background-color: lightgreen' if i % 2 == 0 else '' for i in range(len(s))]

# Select every other row using iloc and apply the function
#df.iloc[::2].style.apply(highlight_every_other_row, axis=1)
#df.style.apply(highlight_every_other_row, axis=0)


#某只股票的K线图。
def sec_k_line_days(security, start_date, end_date):

	security=Jqdata(security).jqcode

	end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
	stock_info = get_security_info(security)
	df = get_price(security, start_date=start_date, end_date=end_date, frequency='daily')

	fig, ax1 = plt.subplots(figsize=(15, 10))
	ax2 = ax1.twinx()  # 创建一个共享X轴的新坐标轴

	# 画K线图
	candlestick_data = []
	for i in range(len(df)):
	    date = mdates.date2num(df.index[i].to_pydatetime())
	    open_price = df['open'][i]
	    close_price = df['close'][i]
	    high_price = df['high'][i]
	    low_price = df['low'][i]
	    if close_price > open_price:
	        color = 'red'
	    else:
	        color = 'green'
	    candlestick_data.append((date, open_price, close_price, high_price, low_price, color))

	ax1.xaxis_date()
	ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
	ax1.autoscale_view()

	for date, open_price, close_price, high_price, low_price, color in candlestick_data:
	    ax1.plot([date, date], [low_price, high_price], color='black')
	    ax1.plot([date, date], [open_price, close_price], color=color, linewidth=8)

	# 画成交量
	ax2.bar(df.index, df['volume'], color='gray', alpha=0.3)

	# 计算移动平均线
	df['ma5'] = df['close'].rolling(window=5).mean()
	df['ma10'] = df['close'].rolling(window=10).mean()
	df['ma20'] = df['close'].rolling(window=20).mean()

	# 画移动平均线
	ax1.plot(df.index, df['ma5'], label='MA5')
	ax1.plot(df.index, df['ma10'], label='MA10')
	ax1.plot(df.index, df['ma20'], label='MA20')

	ax1.set_title(f'{stock_info.display_name}')
	ax1.set_xlabel('Date')
	ax1.set_ylabel('Price')
	ax2.set_ylabel('Volume')  # 设置成交量的Y轴标签
	ax1.grid(True)
	ax1.legend()
	ax1.text(df.index[-1], df['close'][-1], security, ha='right', va='bottom', fontsize=12)

	plt.show()



#某只股票分时图。
def sec_k_line_datenow(security, end_date): 

	security=Jqdata(security).jqcode

	stock_info = get_security_info(security)  
	end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()  

	# 获取股票分时数据和成交量数据  
	stock_data = get_price(security, start_date=end_date, end_date=end_date + datetime.timedelta(days=1), frequency='1m', fields=['open', 'high', 'low', 'close', 'volume'])  

	# 删除休市时间的数据  
	stock_data = stock_data.between_time('09:30', '11:30').append(stock_data.between_time('13:00', '15:00'))  

	# 检查是否有数据  
	if not stock_data.empty:  
	    # 绘制分时图和成交量图  
	    fig, ax1 = plt.subplots(figsize=(12, 6))  
	    ax2 = ax1.twinx()  
	    ax1.plot(stock_data.index, stock_data['close'], label=stock_info.display_name)  
	    ax2.bar(stock_data.index, stock_data['volume'], alpha=0.3, width=0.0005, color='red')  
	    ax1.set_xlabel('Time')  
	    ax1.set_ylabel('Price')  
	    ax2.set_ylabel('Volume')  
	    ax1.set_title(f'{stock_info.display_name} 分时图 ({end_date})')  
	    ax1.set_ylim(stock_data['close'].min() * 0.99, stock_data['close'].max() * 1.01)  
	    ax2.set_ylim(stock_data['volume'].min() * 0.99, stock_data['volume'].max() * 4)  
	    ax1.legend()  
	    ax1.grid()  

	    # 删除休市时间段的数据  
	    idx = stock_data.between_time('11:30', '13:00').index  
	    if not idx.empty:  
	        ax1.plot(idx, stock_data.reindex(idx).dropna()['close'], color='white')  

	else:  
	    print("今天没有数据，请检查股票代码和日期。")  

	plt.show()  # 显示图形  


#某只股票的K线图和分时图
def sec_k_line(security, start_date, end_date):

	#某只股票的K线图。

	sec_k_line_days(security,start_date,end_date)

	#某只股票的分时图。

	sec_k_line_datenow(security, end_date)

#指数成分股按YTD排序得到的代码列表。
def  get_index_sec_by_YTD(index_code):

	stockslist=Jqdata(index_code).get_index_securities()


	df_list=[
	[stock,Jqdata(stock).security_name(),
	round(Jqdata(stock).get_close_change_YTD(target_date()),2),

	] 
	for stock in stockslist]

	df_columns=['code','name','YTD']

	df=list2dataframe(df_list,df_columns)

	dataframe_ranking(df,'YTD')


	code_sorted_list=list(df['code'])


	return code_sorted_list

#指数成分股所有股票的K线图和分时图。
def get_index_sec_k_line(index_code,start_date,end_date):

	code_sorted_list=get_index_sec_by_YTD(index_code)

	for code in code_sorted_list:

		sec_k_line(code,start_date,end_date)


	print('done')

#指数成分股的K线图。
def get_index_sec_k_line_days(index_code,start_date,end_date):

	code_sorted_list=get_index_sec_by_YTD(index_code)

	for code in code_sorted_list:

		sec_k_line_days(code,start_date,end_date)


	print('done')


#指数成分股的分时图。
def get_index_sec_k_line_datenow(index_code,end_date):

	code_sorted_list=get_index_sec_by_YTD(index_code)

	for code in code_sorted_list:

		sec_k_line_datenow(code,end_date)


	print('done')



#股票列表中的所有股票代码，按YTD排序得到的。
def  get_stocklist_by_YTD(stockslist):

	#stockslist=Jqdata(index_code).get_index_securities()


	df_list=[
	[Jqdata(stock).jqcode,Jqdata(stock).security_name(),
	round(Jqdata(stock).get_close_change_YTD(target_date()),2),

	] 
	for stock in stockslist]

	df_columns=['code','name','YTD']

	df=list2dataframe(df_list,df_columns)

	dataframe_ranking(df,'YTD')


	code_sorted_list=list(df['code'])


	return code_sorted_list

#股票列表中的所有股票K线图和分时图。
def get_stocklist_k_line(stockslist,start_date,end_date):

	code_sorted_list=get_stocklist_by_YTD(stockslist)

	for code in code_sorted_list:

		sec_k_line(code,start_date,end_date)


	print('done')

#股票列表中的所有股票K线图。
def get_stocklist_k_line_days(stockslist,start_date,end_date):

	code_sorted_list=get_stocklist_by_YTD(stockslist)

	for code in code_sorted_list:

		sec_k_line_days(code,start_date,end_date)


	print('done')


#股票列表中的所有股票分时图。
def get_stocklist_k_line_datenow(stockslist,end_date):

	code_sorted_list=get_stocklist_by_YTD(stockslist)

	for code in code_sorted_list:

		sec_k_line_datenow(code,end_date)


	print('done')


def show_index_weekly_volume_chart(index_code,end_date): #查看一个指数过去20周的周成交量的变化。以此观察市场。


	index_code=Jqdata(index_code).jqcode



	# Define the index name in Chinese characters
	index_name = Jqdata(index_code).security_name()

	# Get the daily volume data for the past 20 weeks
	daily_volume = get_price(index_code, end_date=end_date, count=20*7, frequency='1d')['volume']

	# Resample the daily volume data to weekly frequency
	weekly_volume = daily_volume.resample('1w').sum()

	# Calculate the weekly trading volume increase from the previous week as a percentage
	weekly_volume_increase = weekly_volume.pct_change() * 100

	# Set the chart size and color
	plt.figure(figsize=(16, 8))
	plt.rcParams['axes.facecolor'] = '#f5f5f5'

	# Create a bar chart of the weekly volume data with different color and bigger size
	plt.bar(weekly_volume.index, weekly_volume.values, color='#1f77b4', width=5)

	# Set the chart title and axis labels
	plt.title('Weekly Volume for ' + index_name, fontsize=20)
	plt.xlabel('Week')
	plt.ylabel('Volume')

	# Add the monthday labels below each bar with rotation=45
	ax = plt.gca()
	ax.xaxis.set_major_locator(mdates.WeekdayLocator(byweekday=mdates.FRIDAY))
	ax.xaxis.set_major_formatter(mdates.DateFormatter('%m%d'))
	plt.xticks(rotation=45)

	# Add the weekly trading volume increase as a line chart
	ax2 = ax.twinx()
	ax2.plot(weekly_volume.index, weekly_volume_increase.values, color='#ff7f0e', linewidth=2)
	ax2.set_ylabel('Volume Increase (%)')

	# Display the chart
	plt.show()

	# Add the weekly trading volume increase to the weekly volume data
	weekly_volume_df = pd.DataFrame({'Volume': weekly_volume, 'Volume Increase (%)': weekly_volume_increase})

	# Print the weekly trading volume data in pandas format with time series
	print(weekly_volume_df)




def show_index_daily_volume_chart(index_code,end_day):

	index_code=Jqdata(index_code).jqcode

	# 定义指数代码和名称

	index_name = Jqdata(index_code).security_name()

	# 过去30个交易日的成交量数据
	daily_volume = get_price(index_code, end_date=end_day, count=30, frequency='1d')['volume']

	# 计算每日成交量增幅
	daily_volume_increase = daily_volume.pct_change() * 100

	# 创建一个新的DataFrame，包含每日成交量和每日成交量增幅
	volume_data = pd.DataFrame({'Daily Volume': daily_volume, 'Daily Volume Increase (%)': daily_volume_increase})

	# 判断连续两日成交量长是否超过20%
	if ((daily_volume_increase[-2] > 20) & (daily_volume_increase[-1] > 20)):
	    print('{}  连续两日成交量增长超过20%，考虑  减仓清仓').format(index_name)

	# 判断连续三日成交量增长是否超过15%
	if ((daily_volume_increase[-3] > 15) & (daily_volume_increase[-2] > 15) & (daily_volume_increase[-1] > 15)):
	    print('{} 连续三日成交量增长超过15%，注意风险').format(index_name)

	# 绘制每日成交量的柱形图和每日成交量增幅形成的点线图
	fig, ax1 = plt.subplots(figsize=(16, 8))
	ax2 = ax1.twinx()
	ax1.bar(daily_volume.index, daily_volume.values, color='#1f77b4', width=0.5)

	# 设置图表标题和轴标签
	ax1.set_title('Daily Volume and Volume Increase for ' + index_name, fontsize=20)
	ax1.set_xlabel('Date')
	ax1.set_ylabel('Volume')
	ax2.set_ylabel('Volume Increase (%)')

	# 将日期以monthday的格式显示在图表下方
	ax1.xaxis.set_major_locator(mdates.DayLocator(interval=1))
	ax1.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
	plt.xticks(rotation=45)

	# 绘制每日成交量增幅形成的点线图
	ax2.plot(daily_volume_increase.index, daily_volume_increase.values, color='#ff7f0e', linewidth=2, marker='o')

	# 使每个日期标签都显示，并且以45度角倾斜
	for label in ax1.get_xticklabels():
	    label.set_visible(True)
	    label.set_rotation(45)

	plt.show()

	print(volume_data)


index_list=['sh000001', 'sh000682','sh000685','sz399965',
  'sh000688', 'sh000852', 'sz399905', 'sh000016', 'sh000097', 'sh000805', 'sz399006', 'sz399004', 'sz399005', 'sz399300', 'sz399330', 'sz399333', 'sz399612', 'sz399673', 'sz399903', 'sh000922',
 'sz399324', 'sh000812', 'sh000811', 'sh000810', 'sh000813', 'sh000814', 'sh000815', 'sh000816', 'sh000818', 'sh000819', 'sh000827', 'sh000901', 'sh000928', 'sh000929', 'sh000930', 'sh000931', 
 'sz399932', 'sh000933', 'sh000934', 'sh000935', 'sh000941', 'sh000977', 'sh000978', 'sh000979', 'sh000986', 'sh000987', 'sh000988', 'sh000989', 'sh000990', 'sh000991', 'sh000992', 'sh000993', 'sh000998', 
 'sz399030', 'sz399060', 'sz399295', 'sz399296', 'sz399393', 'sz399394', 'sz399395', 'sz399396', 'sz399412', 'sz399417', 'sz399419', 'sz399420', 'sz399423', 'sz399432', 'sz399435', 'sz399438', 'sz399439',
  'sz399440', 'sz399441', 'sz399608', 'sz399610', 'sz399646', 'sz399647', 'sz399653', 'sz399654', 'sz399669', 'sz399674', 'sz399675', 'sz399676', 'sz399677', 'sz399683', 'sz399684', 'sz399687', 'sz399693', 
  'sz399695', 'sz399699', 'sz399704', 'sz399705', 'sz399706', 'sz399804', 'sz399805', 'sz399806', 'sz399809', 'sz399807', 'sz399808', 'sz399810', 'sz399811', 'sz399812', 'sz399813', 'sz399814', 'sz399967', 
  'sz399970', 'sz399971', 'sz399973', 'sz399975', 'sz399976', 'sz399986', 'sz399987', 'sz399989', 'sz399993', 'sz399994', 'sz399995', 'sz399996', 'sz399997', 'sz399998','sz399966',]



def check_danger_index(end_day):

    danger_index_list = []

    indexcode_list=filter_allindex(end_day)
    #danger_index_set = set()
    
    for index_code in indexcode_list:
        #if index_code in danger_index_set:
        #    continue
            
        index_code=Jqdata(index_code).jqcode
        index_name = Jqdata(index_code).security_name()
        
        # 过去start_days个交易日的成交量数据
        daily_volume = get_price(index_code, end_date=end_day, count=5, frequency='1d')['volume']  
        
        # 计算每日成交量增幅
        daily_volume_increase = daily_volume.pct_change() * 100 
        
        # 判断一个交易日交易量增长超过30%
        if daily_volume_increase[-1] > 30:
            danger_index_list.append([index_code, index_name,  
                                      daily_volume_increase[-1],
                                      daily_volume_increase[-2],
                                      daily_volume_increase[-3]])
            #danger_index_set.add(index_code)
                                     
        # 判断连续两日成交量长是否超过20%
        if ((daily_volume_increase[-2] > 20) & (daily_volume_increase[-1] > 20)):
            danger_index_list.append([index_code, index_name,            
                                      daily_volume_increase[-1],
                                      daily_volume_increase[-2],
                                      daily_volume_increase[-3]])               
            #danger_index_set.add(index_code)  
                                      
        # 判断连续两日成交量长是否超过20%
        if ((daily_volume_increase[-2] < -20) & (daily_volume_increase[-1] < -20)):    
            danger_index_list.append([index_code, index_name,
                                      daily_volume_increase[-1],
                                      daily_volume_increase[-2],
                                      daily_volume_increase[-3]])
            #danger_index_set.add(index_code)
        
        # 判断连续三日成交量增长是否超过15% 
        if ((daily_volume_increase[-3] > 15) & 
            (daily_volume_increase[-2] > 15) & 
            (daily_volume_increase[-1] > 15)):
            danger_index_list.append([index_code, index_name, 
                                      daily_volume_increase[-1],
                                      daily_volume_increase[-2],
                                      daily_volume_increase[-3]])
            #danger_index_set.add(index_code)
                                      
        # 判断连续三日成交量增长是否超过15%
        if ((daily_volume_increase[-3] < -15) & 
            (daily_volume_increase[-2] < -15) & 
            (daily_volume_increase[-1] < -15)):
            danger_index_list.append([index_code, index_name,
                                      daily_volume_increase[-1],
                                      daily_volume_increase[-2],
                                      daily_volume_increase[-3]])
            #danger_index_set.add(index_code)   

    
    # 将 danger_index_list 转为 DataFrame 添加至 danger_index_set                  
    danger_df = pd.DataFrame(danger_index_list, columns=['index_code', 'index_name',  
                                                        'vlchg_1',
                                                        'vlchg_2',
                                                        'vlchg_3']) 

    #以上这行代码，之前写了两个 vlchg_1,导致有两列的名称一样，导致在使用style功能的时候，出现ValueError: style is not supported for non-unique indicies.，这种错误
    #一般由有相同的列名称，或者相同的索引值造成的，特此记录，以免以后出错。https://stackoverflow.com/questions/55430318/pandas-style-tag-give-valueerror-style-is-not-supported-for-non-unique-indices

    danger_df=danger_df.drop_duplicates()

    danger_df['close_chg']=danger_df['index_code'].apply(get_closechange,end_day=end_day)

    danger_df.reset_index(drop=True,inplace=True)

    if end_day==datetime.datetime.now().strftime('%Y-%m-%d'):

    	date_time=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  #如果要看的是当前日期的，就这么写，可以精确到具体时间。

    	print('数据日期：{}'.format(date_time))

    else:

    	print('数据日期：{}'.format(end_day))
    
    return danger_df



def get_trading_money(code,end_day):

	trading_money=Jqdata(code).get_present_money(end_day)

	return trading_money


def get_average_trading_money(code,end_day):

	average_money=Jqdata(code).get_average_money(end_day)

	return average_money


def get_closechange(code,end_day):

	present_close_change=Jqdata(code).get_present_close_change(end_day)

	return present_close_change




def check_danger_etf(end_day):


    danger_index_list = []
    #danger_index_set = set()

    etf_df=list2dataframe(etfdf_list,etfdf_columns)

    indexcode_list=tuple(etf_df.code)


    
    for index_code in indexcode_list:
        #if index_code in danger_index_set:
        #    continue
            
        index_code=Jqdata(index_code).jqcode
        index_name = Jqdata(index_code).security_name()
        
        # 过去start_days个交易日的成交量数据
        daily_volume = get_price(index_code, end_date=end_day, count=5, frequency='1d')['volume']  
        
        # 计算每日成交量增幅
        daily_volume_increase = daily_volume.pct_change() * 100 
        
        # 判断一个交易日交易量增长超过30%
        if daily_volume_increase[-1] > 30:
            danger_index_list.append([index_code, index_name,  
                                      daily_volume_increase[-1],
                                      daily_volume_increase[-2],
                                      daily_volume_increase[-3]])
            #danger_index_set.add(index_code)
                                     
        # 判断连续两日成交量长是否超过20%
        if ((daily_volume_increase[-2] > 20) & (daily_volume_increase[-1] > 20)):
            danger_index_list.append([index_code, index_name,            
                                      daily_volume_increase[-1],
                                      daily_volume_increase[-2],
                                      daily_volume_increase[-3]])               
            #danger_index_set.add(index_code)  
                                      
        # 判断连续两日成交量长是否超过20%
        if ((daily_volume_increase[-2] < -20) & (daily_volume_increase[-1] < -20)):    
            danger_index_list.append([index_code, index_name,
                                      daily_volume_increase[-1],
                                      daily_volume_increase[-2],
                                      daily_volume_increase[-3]])
            #danger_index_set.add(index_code)
        
        # 判断连续三日成交量增长是否超过15% 
        if ((daily_volume_increase[-3] > 15) & 
            (daily_volume_increase[-2] > 15) & 
            (daily_volume_increase[-1] > 15)):
            danger_index_list.append([index_code, index_name, 
                                      daily_volume_increase[-1],
                                      daily_volume_increase[-2],
                                      daily_volume_increase[-3]])
            #danger_index_set.add(index_code)
                                      
        # 判断连续三日成交量增长是否超过15%
        if ((daily_volume_increase[-3] < -15) & 
            (daily_volume_increase[-2] < -15) & 
            (daily_volume_increase[-1] < -15)):
            danger_index_list.append([index_code, index_name,
                                      daily_volume_increase[-1],
                                      daily_volume_increase[-2],
                                      daily_volume_increase[-3]])
            #danger_index_set.add(index_code)    
    
    # 将 danger_index_list 转为 DataFrame 添加至 danger_index_set                  
    danger_df = pd.DataFrame(danger_index_list, columns=['index_code', 'index_name',  
                                                        'vlchg_1',
                                                        'vlchg_2',
                                                        'vlchg_3']) 
    danger_df = danger_df.drop_duplicates()

    
    danger_df['pre_m']=danger_df['index_code'].apply(get_trading_money,end_day=end_day)


    danger_df['avrg_m']=danger_df['index_code'].apply(get_average_trading_money,end_day=end_day)

    filter=danger_df['avrg_m']<20000000  #排除掉低于2000W的ETF.

    remove_code_list=list(danger_df[filter]['index_code'])

    selected_df = danger_df.loc[~danger_df['index_code'].isin(remove_code_list)].copy()

    selected_df['avrg_m']=round(selected_df['avrg_m']/100000000,3)

    selected_df['close_chg']=selected_df['index_code'].apply(get_closechange,end_day=end_day)

    selected_df.reset_index(drop=True,inplace=True)

    if end_day==datetime.datetime.now().strftime('%Y-%m-%d'):

    	date_time=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  #如果要看的是当前日期的，就这么写，可以精确到具体时间。

    	print('数据日期：{}'.format(date_time))

    else:

    	print('数据日期：{}'.format(end_day))
    
    return selected_df



def filter_etf_with_tradingmoney(df,column_name,end_day):  #针对不管那种情况得到的etf df数据结果，用成交额进行过滤，去掉成交量低的不要。

    # Apply the trading money functions directly for filtering
    df['avrg_m'] = df[column_name].apply(get_average_trading_money, end_day=end_day)

    # Filter out stocks with average trading money less than 30,000,000   #日成交额小于3000W的绝对排除。成交量太低了，还是要放弃吧。
    selected_df = df[df['avrg_m'] >= 0.2].copy()     #0.3亿就是3000W。因为前面计算的就是以亿为单位的。

    # Drop unnecessary columns
    #selected_df = selected_df.drop(columns=['avrg_m'])  #决定对这个进行保留，以便观察成交额高的ETF。

    # Reset index
    selected_df.reset_index(drop=True, inplace=True)

    return selected_df


def filter_etflist_with_tradingmoney(target_etflist,column_name,end_day): #过滤掉成交额低的ETF.

	targetetf_list=[[x,Jqdata(x).security_name()] for x in target_etflist]

	targetetf_columns=['code','name']

	targetetf_df=list2dataframe(targetetf_list,targetetf_columns)

	final_etf_df=filter_etf_with_tradingmoney(targetetf_df,'code',end_day)

	final_etf_code_list=list(final_etf_df['code'])

	return final_etf_code_list


def display_top_sec_from_single_index(index_code,total_money,present_day,target_percent,select_ma):#找出某个指数下最强的股票，并且展示这些股票的策略数据和实际市场表现数据。这个是根据短期排名靠前和长期排名靠前的数据选择出来强的股票。

	a,b,index_stock_codelist=find_top_sec_single_index(index_code,total_money,present_day,target_percent)

	best_stock_codelist=[x for x in index_stock_codelist if not Jqdata(x).judge_recent_public_stock(present_day)]

	#best_stock_codelist=filter_stocklist_with_tradingmoney(index_stock_codelist,present_day) #先过滤掉成交额小的股票。

	topsec_strategy_df=Trading(best_stock_codelist).trading_strategy_buy(total_money,present_day,select_ma)

	greatstockdf=topsec_strategy_df[(topsec_strategy_df['buy']=='yes')]

	#greatstockdf=filter_stock_with_tradingmoney(greatstockdf,'code',present_day) #将每日成交额低于3000的除掉。 已经做过这个处理了。

	greatstockdf.reset_index(drop=True,inplace=True)

	final_best_stock_codelist=list(greatstockdf['code'])

	topsec_market_df=get_target_stocklist_closechange_data(final_best_stock_codelist,total_money,present_day)

	return greatstockdf,topsec_market_df


def find_stock_better_than_index(index_code,total_money,end_day,select_ma): #单个指数中的表现强于指数的股票数据信息，分为长期强于指数（5，10，20，30，60）和短期强于指数(5,10)两种情况。

	index_info_list=get_single_index_data(index_code,end_day)

	#index_columns=indexdf_columns=['code', 'name','close', 'MHg', 'MCg', 'cl_1', 'cl_2', 'cl_5', 'cl_10', 'cl_20', 'cl_30', 'cl_60', 'YTD', 'vlchg', 'vlrk', 'h_p_t', 'warn', 'zf', 'zfr']

	#index_df=list2dataframe(index_info_list,indexdf_columns)


	index_stock_code_list=Jqdata(index_code).get_index_securities() #得到目标指数代码的成分股。

	index_stock_code_list_transformed=[transform_jqdata_code(code) for code in index_stock_code_list]  #转换成腾讯格式代码。

	df=get_target_stocklist_closechange_data(index_stock_code_list_transformed,total_money,end_day)


	df['cl_5_index']=index_info_list[7]

	df['cl_10_index']=index_info_list[8]

	df['cl_20_index']=index_info_list[9]

	df['cl_30_index']=index_info_list[10]

	df['cl_60_index']=index_info_list[11]

	df['cl_YTD_index']=index_info_list[12]


	conditionlist=[
		         ((df['cl_5'] > df['cl_5_index']) & (df['cl_10'] > df['cl_10_index'])
	               & (df['cl_20']> df['cl_20_index']) & (df['cl_30'] > df['cl_30_index']) & (df['cl_60'] > df['cl_60_index']) & (df['YTD'] > df['cl_YTD_index'])),
		         ((df['cl_5'] > df['cl_5_index']) & (df['cl_10'] > df['cl_10_index']))
		              ]

	best_sec='best'  #股票的5，10，20，30，60以及年初到现在的表现都比指数好的股票，就是best,

	bad_sec='good'   #股票的5，10天的表现强于指数的就是 good.


	choicelist=[best_sec,bad_sec]

	df['mark']=np.select(conditionlist,choicelist,default='NO')   #this block of code is very new to me, to compare the different column, and then make a new column after action.

	df=df.drop(['cl_5_index','cl_10_index','cl_20_index','cl_30_index','cl_60_index'],axis=1)

	best_df=df[df['mark']=='best']

	good_df=df[df['mark']=='good']

	best_stock_codelist=list(best_df['code'])

	good_stock_codelist=list(good_df['code'])

	best_stock_strategy=Trading(best_stock_codelist).trading_strategy_buy(total_money,end_day,select_ma)

	good_stock_strategy=Trading(good_stock_codelist).trading_strategy_buy(total_money,end_day,select_ma)

	print('best 股票指的是 股票的5，10，20，30，60以及年初到现在的表现都比指数强 \n ')

	print('good 股票指的是 股票的5，10天的表现都强于指数\n')


	return best_stock_strategy,best_df,good_stock_strategy,good_df


def find_not_in_holdingstock(source_df,holdingstock_df):  #这是找出目标股票数据里面有，而持有的股票里面没有的股票。

	source_df_codelist=list(source_df['code'])

	holdingstock_df_codelist=list(holdingstock_df['code'])

	not_in_holdingstock=[x for x in source_df_codelist if x not in holdingstock_df_codelist]

	selected_df = source_df.loc[source_df['code'].isin(not_in_holdingstock)]  #找出符合etflist里面的代码的所有行。

	rows = selected_df.copy()

	return rows


def get_different_rows_from_dataframe(df_a,df_b,column_name): #从a中获取b中根据column_name的值没有的数据。

	if not df_a.empty and not df_b.empty: #必须要比较的两个任何一个不能为空的。

		selected_rows = df_a[~df_a[column_name].isin(df_b[column_name])].copy()

	else:

		selected_rows=pd.DataFrame()

	return selected_rows


def get_same_rows_from_dataframe(df_a,df_b,column_name):#从a中获取b中根据column_name的值也有的数据。

	if not df_a.empty and not df_b.empty: #必须要比较的两个任何一个不能为空的。

		selected_rows = df_a[df_a[column_name].isin(df_b[column_name])]

	else:

		selected_rows=pd.DataFrame()


	return selected_rows


#如何长期短期都需要买入的ETF市场表现数据。

def get_buying_etf_data(source_df,total_money,end_day):

	if not source_df.empty:

	    etf_code_list=list(source_df['code'])

	    etf_code_list.remove(nan)

	    market_df=get_target_etflist_closechange_data(etf_code_list,total_money,end_day)

	else:
 
	    market_df=pd.DataFrame()

	#finaldf=freeze_columns(market_df)

	return market_df


#如何长期短期都需要买入的ETF市场表现数据。

def get_buying_index_data(source_df,end_day):

	if not source_df.empty:

	    index_code_list=list(source_df['code'])

	    index_market_df=target_index_watch(index_code_list,end_day)

	else:
 
	    index_market_df=pd.DataFrame()

	#final_index_df=freeze_columns(index_market_df)

	return index_market_df


def get_buying_stock_data(source_df,total_money,end_day):

	if not source_df.empty:

	    stock_code_list=list(source_df['code'])

	    stock_market_df=get_target_stocklist_closechange_data(stock_code_list,total_money,end_day)

	else:
 
	    stock_market_df=pd.DataFrame()

	#final_stock_df=freeze_columns(stock_market_df)

	return stock_market_df


def judge_stock_zhangting(stock_code,end_day): #判断目标股票在至今的20个交易日里有没有涨停， 作为将来是早盘买还是收盘前买的依据。

	stock_closechange_df=Jqdata(stock_code).get_stock_closechange_in_days(21,end_day)  #获取前20个交易的日数据。因为要计算百分比，所以必须要21天的数据，这样前20天的收盘价，才能去除以前20天的收盘价，得到涨幅变化。

	closechange_list=list(stock_closechange_df['close'])

	if stock_code.startswith('sz30') or stock_code.startswith('sh688') or stock_code.startswith('30') or stock_code.startswith('688'):

		highest_increase=19.9  #涨停价的幅度

	else:

		highest_increase=9.9  #涨停价的幅度


	judge_value_list=[x for x in closechange_list if x > highest_increase]  #如果列表不为空，说明之前有涨停的时候。

	return bool(judge_value_list)


class Jq_codelist():
	'''create a class for jqdata with stocklist as objective'''


	def __init__(self,stockcodelist):

		self.stockcodelist=stockcodelist  #以stockcodelist为目标对象。


	def judge_stocklist_above_ma(self,ma_days,present_day):  #提出不满足条件的股票列表中的股票。

		del_items_list=[x for x in self.stockcodelist if not Jqdata(x).judge_above_ma(ma_days,present_day)]

		stocklist=[x for x in self.stockcodelist if x not in del_items_list]

		return stocklist,ma_days



	def check_stock_indicator_stocklist(self,total_money,end_day):#获取低于5天/10线之类的股票信息，以便及时发现，并警惕。

		codelist=checking_list_for_repeat(self.stockcodelist)  #先去掉表里面重复的元素。

		stocklist=[transform_to_normal_code(code) for code in codelist]  #一定要把它变成正规的code,而不是平台的code,否则容易出错。

		open_equal_high=[]  #开盘等于最高价的列表。

		limit_number=len(stocklist)

		i=0

		below_ma_list=[]

		while i < limit_number:

			#bool_value=Jqdata(stocklist[i]).judge_above_ma(ma_days,end_day)

			myjq=Jqdata(stocklist[i])

			#sec_ma=myjq.get_ma(ma_days,end_day)

			security_name=myjq.security_name()

			yesterday_close=myjq.get_yesterday_close(end_day)

			sec_closechange=round(myjq.get_close_change(end_day,1),2)  #今天相对于昨天的涨幅

			present_point,present_open,present_high,present_low=myjq.get_stock_point(end_day)  #获取现价,开盘价，最高，最低价

			if present_open==present_high:

				open_equal_high.append(security_name)


			if stocklist[i].startswith('sz30') or stocklist[i].startswith('sh688') or stocklist[i].startswith('30') or stocklist[i].startswith('688'):

				highest_increase=18  #涨停价的幅度

			else:

				highest_increase=8   #涨停价的幅度


			high_low_gap=round((present_high - present_low)/yesterday_close,4)*100  #当天截止到目前的振幅   注意振幅是如何计算的，当天最高价-当天最低价 / 昨天收盘价

			high_price_percent=round((present_high - yesterday_close)/yesterday_close,4)*100  #当天最高价的涨幅



			if high_price_percent >= highest_increase:
			    risk_word = '可能危险'
			    very_risk_word = '非常危险!' if high_price_percent > sec_closechange else ''
			else:
			    risk_word = ''
			    very_risk_word = ''

			ma5,ma10,ma20=myjq.get_ma(5,end_day),myjq.get_ma(10,end_day),myjq.get_ma(20,end_day)

			gap_to_ma5=round(round((present_point - ma5)/ma5,4)*100,2)

			gap_to_ma10=round(round((present_point - ma10)/ma10,4)*100,2) 

			gap_to_ma20=round(round((present_point - ma20)/ma20,4)*100,2) 

			#gap_to_ma=round(round((present_point - sec_ma)/sec_ma,4)*100,2)  #距离几天均线还有多少个百分点,正常就是高于均线，负数就是低于均线。

			#if bool_value==False:

			#	ma_word=''

			#else:

				#ma_word='高'

			

			buy_volume=myjq.buy_single_stock_volume(total_money,end_day)  #这里设置的是200，0000的买入总额。

			volume_change=myjq.get_volume_change(end_day)

			volume_ranking=myjq.get_volume_ranking(end_day)

			if stocklist[i].startswith('sz3') or stocklist[i].startswith('sh6') or stocklist[i].startswith('sz0') or stocklist[i].startswith('3') or stocklist[i].startswith('6') or stocklist[i].startswith('0'):


				sec_industry=myjq.security_industry(end_day)   #为股票设计的。

			else:

				sec_industry='None'  #为etf设计的。

			highest_price_time=myjq.find_stock_highest_price_time(end_day)

			if highest_price_time < '10:00':

			#if highest_price_time < datetime.time(10,0,0):

				warningword='危'

			else: 

				warningword=' '


			zhenfu_data=myjq.get_zhenfu(end_day)

			zhenfu_ranking=myjq.get_zhenfu_ranking(end_day)

			#sec_info=[stocklist[i],security_name,buy_volume,present_point,sec_ma,ma_word,gap_to_ma,sec_closechange,present_open,present_high,present_low,high_low_gap,high_price_percent,highest_increase,risk_word,very_risk_word,volume_change,volume_ranking]

			sec_info=[stocklist[i],security_name,sec_industry,buy_volume,present_point,gap_to_ma5,gap_to_ma10,gap_to_ma20,sec_closechange,present_open,present_high,present_low,high_low_gap,high_price_percent,highest_increase,risk_word,very_risk_word,volume_change,volume_ranking,highest_price_time,warningword,zhenfu_data,zhenfu_ranking]

			below_ma_list.append(sec_info)

			i += 1

		ma_message='\n参考的是距离 {}天均线的 差距!\n'.format(ma_days)

		print(ma_message)

		df_columns=['code','name','industry','b_vl','price','距5日线','距10日线','距20日线','涨幅','open','high','low','振幅','最高','涨停','risk','涨停开','vlchg','vlrk','h_p_t','warn','zf','zfr']  #尽可能压缩列的名字长度。


		df=list2dataframe(below_ma_list,df_columns)

		df.reset_index(drop=True, inplace=True)

		dataframe_ranking(df,'vlchg')

		df.reset_index(drop=True,inplace=True)

		monitoring_sec_number=len(df)

		if end_day==datetime.datetime.now().strftime('%Y-%m-%d'):

			date_time=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  #如果要看的是当前日期的，就这么写，可以精确到具体时间。

			print('数据日期：{}'.format(date_time))

		else:

			print('数据日期：{}'.format(end_day))



		#print('数据日期:{}'.format(end_day))

		print('\n **********  开盘价等于最高价的股票有  {},  要当心！********* '.format(open_equal_high))


		#print('低于 {} 天均线的股票有 {} 个，代码如下:'.format(ma_days,monitoring_sec_number))

		return df



	def get_below_ma_stocklist(self,total_money,ma_days,end_day):#获取低于5天/10线之类的股票信息，以便及时发现，并警惕。

		targetdf=self.check_stock_indicator_stocklist(total_money,end_day)

		key_column='距'+str(ma_days)+'日线'

		below_ma_df=targetdf[targetdf[key_column]<0]

		below_ma_number=len(below_ma_df)

		print('低于 {} 天均线的股票有 {} 个，代码如下:'.format(ma_days,below_ma_number))

		return below_ma_df


	def get_above_ma_stocklist(self,total_money,ma_days,end_day):#获取高于5天/10线之类的股票信息，以便及时发现，并警惕。

		targetdf=self.check_stock_indicator_stocklist(total_money,end_day)

		key_column='距'+str(ma_days)+'日线'

		above_ma_df=targetdf[targetdf[key_column]>=0]

		above_ma_number=len(above_ma_df)

		print('高于 {} 天均线的股票有 {} 个，代码有:'.format(ma_days,above_ma_number))

		return above_ma_df



#应该是 stock ,etf都可以使用以下的类和方法。

class Trading():

	'''create class for trading strategy for stock,index,etf'''

	def __init__(self,target_code_list):

		self.target_code_list=target_code_list
	 

	def get_trading_data(self,total_money,end_day,select_ma):

		targetcodelist=self.target_code_list

		targetcode_list_info = []

		n1, n2, n3, n4, n5 = 5,10,20,100,250    #这个是我的观察时间线。

		targetcode_list=[transform_to_normal_code(x) for x in targetcodelist]

		for targetcode in targetcode_list:
		    targetjq = Jqdata(targetcode)
		    targetcode_name = targetjq.security_name()
		    presentclose = targetjq.get_present_close(end_day)
		    max_high, max_close = targetjq.get_max_high_close(end_day)
		    if judge_stock_zhangting(targetcode, end_day):
		    	zhangting_info='yes'

		    else:

		    	zhangting_info=''

		    highest_gap = round((max_high - presentclose) / max_high, 4) * 100  #20天内从最高点回撤幅度大于10%就是可以考虑卖出的条件之一。


		    #targetcode_industry = targetjq.security_industry(end_day)

		    targetcode_industry = targetjq.security_industry(end_day) if targetcode.startswith('6') or targetcode.startswith('0') or targetcode.startswith('3') else 'None'

		    by_vl=targetjq.buy_single_stock_volume(total_money, end_day) if targetcode.startswith('6') or targetcode.startswith('0') or targetcode.startswith('3') else targetjq.buy_single_etf_volume(total_money, end_day)

		    #by_vl = targetjq.buy_single_stock_volume(total_money, end_day)
		    by_value = int(presentclose * by_vl * 100)


		    ma1,ma2,ma3,ma4,ma5 = targetjq.get_ma(n1, end_day),targetjq.get_ma(n2, end_day),targetjq.get_ma(n3, end_day),targetjq.get_ma(n4, end_day),targetjq.get_ma(n5, end_day)

		    buyin = ''
		    buy_time = ''
		    sellout=''


		    close_up, close_down, high_up, high_down, low_up, low_down = targetjq.judge_stock_data(end_day)

		    standard_ma_day = targetjq.get_ma(select_ma, end_day)


		    targetcode_info = [targetcode, targetcode_name, targetcode_industry, presentclose, by_vl, by_value, zhangting_info,buyin,
		                       buy_time,sellout,standard_ma_day,ma1, ma2, ma3, ma4, ma5, close_up, close_down, high_up, high_down,
		                       low_up, low_down, highest_gap]

		    if not pd.isna(presentclose):  # Instead of dropping rows with missing values using df.drop, you can filter the rows during the loop itself to avoid creating unnecessary rows in the DataFrame.

		        targetcode_list_info.append(targetcode_info)

		targetcode_columns = ['code', 'name', 'industry', 'close', 'b_vl', 'by_value', 'zhangting','buy', '买点','sell','standard_ma',
		                      "MA{}".format(n1), "MA{}".format(n2), "MA{}".format(n3), "MA{}".format(n4),
		                      "MA{}".format(n5), '连涨', '连跌', '高连升', '高连降', '低连升', '低连降', '高点跌幅']

		df = pd.DataFrame(targetcode_list_info, columns=targetcode_columns)
		df.reset_index(drop=True, inplace=True)

		return df


	def trading_strategy_buy(self,total_money, end_day, select_ma):  #只为股票服务。大于均线的先选择出来。

		target_code_list=self.target_code_list

		targetcode_list=[x for x in target_code_list if Jqdata(x).judge_above_ma(select_ma,end_day)]  #选择出符合大于选择的移动平均线的股票。比如先选择出大于10天线的股票。

		df=Trading(targetcode_list).get_trading_data(total_money,end_day,select_ma)  #这里必须开始把这个Trading 调用筛选后的targetcode_list

		df.loc[df['zhangting'] == 'yes', '买点'] ='早'

		conditions = [
		    (df['连涨'] == 'yes'),
		    (df['高连升'] == 'yes')
		]

		df['buy'] = np.where(np.all(conditions, axis=0), 'yes', '')

		df.drop(['zhangting','standard_ma'], axis=1, inplace=True)  #删除不需要的列。


		message = '本次选择的买卖标准是 **** {}  天线！**** \n'.format(select_ma)

		print(message)

		print('buy 代表现价大于 {} 天线，股价连续三天收盘价上涨，最高价每天都上涨。\n'.format(select_ma))

		print('sell 代表现价低于 {} 天线，股价连续三天收盘价下跌，最低价每天都下跌，并且过去20天内股价从最高价到现价已经下跌10%了！ \n'.format(select_ma))


		return df


	def trading_strategy_for_stock(self,total_money, end_day, select_ma):  #只为股票服务。大于均线的先选择出来。这个主要是为了笼统的查询一堆股票的情况，通过以下的交易策略进行标记数据。

		df=self.get_trading_data(total_money,end_day,select_ma)  #这里必须开始把这个Trading 调用筛选后的targetcode_list

		df.loc[df['zhangting'] == 'yes', '买点'] ='早'

		conditions = [
		    (df['close'] > df['standard_ma']),
		    (df['连涨'] == 'yes'),
		    (df['高连升'] == 'yes')
		]

		df['buy'] = np.where(np.all(conditions, axis=0), 'yes', '')


		conditions_2=[
		    (df['close'] < df['standard_ma']),
		    (df['连跌']=='yes'),
		     (df['低连降']=='yes'),
		     (df['高点跌幅']>10)
		]

		df['sell'] = np.where(np.all(conditions_2, axis=0), 'yes', '')


		df.drop(['zhangting','standard_ma'], axis=1, inplace=True)  #删除不需要的列。


		message = '本次选择的买卖标准是 **** {}  天线！**** \n'.format(select_ma)

		print(message)

		print('buy 代表现价大于 {} 天线，股价连续三天收盘价上涨，最高价每天都上涨。\n'.format(select_ma))

		print('sell 代表现价低于 {} 天线，股价连续三天收盘价下跌，最低价每天都下跌，并且过去20天内股价从最高价到现价已经下跌10%了！ \n'.format(select_ma))


		return df


	def trading_target_stock_list(self,total_money,end_day,select_ma):  #主要是选择出大于10天线，符合buy的股票。

		targetcode_list=[transform_to_normal_code(x) for x in self.target_code_list]

		stockcode_list=[x for x in targetcode_list if not Jqdata(x).judge_recent_public_stock(end_day)]  #去掉上市低于1年的股票。

		#stockcode_list=filter_stocklist_with_tradingmoney(targetcode_list,end_day) #先过滤掉不符合成交额需要的股票。

		final_df=Trading(stockcode_list).trading_strategy_buy(total_money,end_day,select_ma)  #选择出大于10天线（选择标准）的股票列表。

		buystock_df=final_df[final_df['buy']=='yes']

		if not buystock_df.empty:

			final_buystock_df=dataframe_ranking(buystock_df.copy(),'低连升')

			buystock_code=list(buystock_df['code'])

			#buystock_market_df=get_target_stocklist_closechange_data(buystock_code,total_money,end_day)

			#final_buystock_market_df=dataframe_ranking(buystock_market_df.copy(),'cl_5')

			#buystock_indicator_df=check_stock_indicator_stocklist_from_df(buystock_df,total_money,ma_days,end_day)

			#final_buystock_indicator_df=dataframe_ranking(buystock_indicator_df.copy(),'最高')

		else:

			final_buystock_df=pd.DataFrame()

			#final_buystock_market_df=pd.DataFrame()

			#final_buystock_indicator_df=pd.DataFrame()

			buystock_code=[]


		#return final_buystock_df,final_buystock_market_df,final_buystock_indicator_df,final_df
		return final_buystock_df,buystock_code



def get_previous_days(end_day,countdays=5):

	previous_trade_days = get_trade_days(count=countdays, end_date=end_day)

	previous_trade_days=[x.strftime('%Y-%m-%d') for x in previous_trade_days]

	return previous_trade_days


def get_yesterday_date(end_day):

	yesterday_date=get_previous_days(end_day,2)[0]

	return yesterday_date


def remove_sell_stock(source_df): #这个函数是为了把标记卖出的股票删除掉。

	mvp_df=source_df[source_df['sell']==''].copy()    #把选择出来的不用卖的留下来。

	#mvp_df=no_sell_df[no_sell_df['买点']=='早']   #把曾经有涨停的单独列出来。最有价值的留下来。 这个带有歧视的意思，没有涨停的也有潜力。

	#mvp_codelist=get_codelist_from_df(mvp_df)  #获取最后留下来的代码。

	return mvp_df


#主要是为了获取交易日正在交易的当天，前一天为基准的各种历史数据。得到各种股票代码为了交易日当天做准备。
def best_trading_stock_list(total_money,ma_days,end_day,select_ma): #选取包括制定日期在内的5个交易日的表现最好的股票，然后把他们全部集中起来参考他们近期的各种数据和表现。

	#previous_trade_days = get_previous_days(end_day,5) #获取指定交易日前5个交易日（包括指定交易日当天）的日期数据

	#bull_stock_list=[]

	#for trade_day in previous_trade_days:

		#yesterdaydate=get_yesterday_date(trade_day) #获取制定交易日前一日的日期数据。得到两个数据，前一天和指定日期两天的日期。

		#potential_bull_stock=get_best_stock(trade_day,select_ma)  #之前这里使用select_bull_stock现在改了。

	bull_stock_list=get_best_stock(end_day,select_ma) 

	#bull_stock_list += potential_bull_stock  #这个必须是表加表，用append就是表中表，就会出错！


	bull_stock_list=list(set(bull_stock_list))  #这是把重复的代码去掉，因为里面有重复的代码。

	print(''.join(['bullstock_codelist=',str(bull_stock_list)]))


	return bull_stock_list  #仅仅获取代码的数据，简化流程。


def make_stock_df_in_one(not_sell_mvp,present_mvp_indicator,present_mvp_closechange): #将几个dataframe合并，适用于股票。

	simple_mvp_df=not_sell_mvp.drop(['close','by_value','MA5','MA10','MA20','MA100','MA250',],axis=1,inplace=False)

	simple_mvp_closechange=present_mvp_closechange.drop(['name','industry','b_vl','zf','vlchg','vlrk','h_p_t','warn','zfr','zf'],axis=1,inplace=False)

	simple_mvp_indicator=present_mvp_indicator.drop(['name','b_vl','industry','price','open','high','low','zf'],axis=1,inplace=False)

	merged_df=simple_mvp_df.merge(simple_mvp_indicator, on='code').merge(simple_mvp_closechange, on='code')


	new_columns=['code', 'name', 'industry', 'b_vl', 'buy', '买点', 'sell', '连涨', '连跌',
	     '高连升', '高连降', '低连升', '低连降', '高点跌幅', '距5日线','距10日线','距20日线', '涨幅', '振幅','zfr', '最高',
	      '涨停', 'risk', '涨停开', 'vlchg', 'vlrk', 'h_p_t', 'warn', 'MHg',
	      'MCg', 'cl_1', 'cl_2', 'cl_5', 'cl_10', 'cl_20', 'cl_30', 'cl_60',
	      'YTD', 'm_size', 'ld_size','cross','duot']

	merged_df=merged_df.reindex(columns=new_columns)

	return merged_df


def make_index_df_in_one(not_sell_mvp,present_mvp_indicator,present_mvp_closechange): #将几个dataframe合并，适用于指数和etf.

	simple_mvp_df=not_sell_mvp.drop(['close','standard_ma','MA5','MA10','MA20','MA100','MA250',],axis=1,inplace=False)

	simple_mvp_closechange=present_mvp_closechange.drop(['name','zf','vlchg','vlrk','h_p_t','warn','zfr','zf'],axis=1,inplace=False)

	simple_mvp_indicator=present_mvp_indicator.drop(['name','price','open','high','low','zf','MHg','MCg'],axis=1,inplace=False)

	merged_df=simple_mvp_df.merge(simple_mvp_indicator, on='code').merge(simple_mvp_closechange, on='code')


	new_columns=['code', 'name', 'buy', 'sell', '连涨', '连跌',
	     '高连升', '高连降', '低连升', '低连降', '高点跌幅', '距5日线','距10日线','距20日线', '涨幅', '振幅','zfr',
	      'vlchg', 'vlrk', 'h_p_t', 'warn', 'MHg',
	      'MCg', 'cl_1', 'cl_2', 'cl_5', 'cl_10', 'cl_20', 'cl_30', 'cl_60',
	      'YTD','cross','duot']

	merged_df=merged_df.reindex(columns=new_columns)

	return merged_df

def get_target_all_data(targetlist,present_day,select_ma):

	trading_target_df=Trading_index(targetlist).get_index_list_data(present_day,select_ma)

	closechange_df=indexlist_monitor(targetlist,present_day,'cl_5')

	indicator_df=check_index_indicator_indexlist(targetlist,present_day)

	final_df=make_index_df_in_one(trading_target_df,indicator_df,closechange_df)

	return final_df


def get_best_index_alldata(present_day,select_ma):  #就是把上面的函数和其他的函数直接运行在一起。

	best_index=get_best_index(present_day,select_ma)

	final_df=get_target_all_data(best_index,present_day,select_ma)

	return final_df


def get_best_etf_alldata(present_day,select_ma):  #就是把上面的函数和其他的函数直接运行在一起。

	best_etf=get_best_etf(present_day,select_ma)

	final_df=get_target_all_data(best_etf,present_day,select_ma)

	bestetf_df=filter_etf_with_tradingmoney(final_df,'code',present_day) 

	return bestetf_df

def get_target_present_high(target_code,high_date):

	present_high=Jqdata(target_code).get_present_high(high_date)

	return present_high


def get_target_etf_period_increase(new_etf_df,high_date):

	new_etf_df['最高涨幅']= round((new_etf_df['high'] - new_etf_df['low']) / new_etf_df['low'] * 100, 4)

	new_etf_df['目前涨幅']=round((new_etf_df['现价'] - new_etf_df['low']) / new_etf_df['low'] * 100, 4)

	new_etf_df['当日最高价']=new_etf_df['code'].apply(get_target_present_high,high_date=high_date)

	new_etf_df['当日最高涨幅']=round((new_etf_df['当日最高价'] - new_etf_df['low']) / new_etf_df['low'] * 100, 4)

	new_etf_df.drop(new_etf_df[pd.isna(new_etf_df['cl_1'])].index,inplace=True)  #必须去掉，这表示这个ETF刚刚上市，不具备统计意义。

	selectedrows=filter_etf_with_tradingmoney(new_etf_df,'code',high_date)  #会把成交量低的ETF干掉。

	selected_df = selectedrows[~selectedrows['name'].str.contains('国开|债|货币')].copy()

	return selected_df


def get_jqetf_period_increase(low_date,high_date):

	jqetf_df=get_etf_high_low_data_from_jq(low_date,high_date)

	new_etf_df=jqetf_df.drop(jqetf_df.columns[[9,10,11,12,13]],axis=1)

	jqetf_period=get_target_etf_period_increase(new_etf_df,high_date)

	#top_jqetf_period=jqetf_period[jqetf_period['最高涨幅']==jqetf_period['当日最高涨幅']]  #阶段性表行最好的ETF,这行代码不好的原因是这样限制了当日必须是最高涨幅的目标，但不同的强的目标，达到高点的日期不同，但这里日期是当日，这样就限制死了选择范围，会把一些好的排除掉，不如下面直接选择出100强来。

	top_jqetf_period=dataframe_ranking(jqetf_period,'最高涨幅').head(n=100) #只选择出排名前100的最高涨幅的etf

	return jqetf_period,top_jqetf_period


def get_jqetf_buying_list_one(top_jqetf_period,best_etf_alldata_from_jq): #必须在运行之前得到两个函数的返回值，get_best_etf_alldata_from_jq, get_jqetf_period_increase

	same_rows_1=get_same_rows_from_dataframe(best_etf_alldata_from_jq,top_jqetf_period,'code')

	same_rows_2=get_same_rows_from_dataframe(top_jqetf_period,best_etf_alldata_from_jq,'code')

	return same_rows_1,same_rows_2


def get_jqetf_buying_list_two(best_etf_alldata_from_jq,best_jqetf): #必须运行两个函数的返回值，get_best_jqetf 是找回在不同时间周期内排名比较好的ETF，get_best_etf_alldata_from_jq

	same_rows1=get_same_rows_from_dataframe(best_etf_alldata_from_jq,best_jqetf,'code')

	same_rows2=get_same_rows_from_dataframe(best_jqetf,best_etf_alldata_from_jq,'code')

	return same_rows1,same_rows2


def get_alletf_period_increase(low_date,high_date):

	alletf_df=get_etf_high_low_data(low_date,high_date)

	new_alletf_df=alletf_df.drop(alletf_df.columns[[9,10,11,12,13]],axis=1)

	final_alletf_df=get_target_etf_period_increase(new_alletf_df,high_date)

	return final_alletf_df


def get_best_etf_alldata_from_jq(present_day,select_ma):  #就是把上面的函数和其他的函数直接运行在一起。

	best_etf=get_best_etf_from_jq(present_day,select_ma)

	final_df=get_target_all_data(best_etf,present_day,select_ma)

	selected_df = final_df[~final_df['name'].str.contains('国开|债')].copy() #必须这么写，才能将字符里包含这两个关键字的行都去掉！ 

	bestetf_df=filter_etf_with_tradingmoney(selected_df,'code',present_day) 

	return bestetf_df


#淘汰其中接连三天高点连续下降的个股。

def remove_down_trend_stock(target_df):#针对选择出来的股票今天的表现，去掉高连降低连降表现不好的股票。这个会改变目前对象的数据，就是目标对象数据会发生变化。所以要谨慎。

	#因为我选择的是去做强势股，如果发现高连降的情况，就考虑卖出了,后续看错了再观察。特别是观察交易日当天的表现和之前交易日的比较。

	all_data=target_df.copy()  #先做一个副本，就不怕改变原始的数据了。

	conditions = [
	    (all_data['高连降'] == 'yes'),
	]

	all_data['sell'] = np.where(np.all(conditions, axis=0), 'yes', '')   #严格筛选将接连三天包括观察当天，三天高点低点都下降的标记为‘sell'，然后排除出去。

	best_vip_df=remove_sell_stock(all_data)   #留下来没有标记sell的

	return best_vip_df


#使用如下的函数，再去适应相应的寻找股票/指数/etf的目标。

def get_best_target(target_code_list,end_day,select_ma): #select_ma 以后决定用20天作为基准，这样尺度放宽些。
#找出股票/etf/指数中，选择出目前股价大于指定均线的股票，然后选择股票大于5天，10天，20天，属于多头排列的股票。
#然后从中选择出来这些股票中，在过去任意5个交易日当中任意往前加两天包括当天就是3天，任意考虑其中的三天，如果那三天范围内，连涨两天，且高点低点连涨两天都往上提升的股票（没有说必须收盘价天天上涨），选择出来。

	targetcode_list=[x for x in target_code_list if Jqdata(x).judge_above_ma(select_ma,end_day)]  #选择出符合大于选择的移动平均线的股票。比如先选择出大于10天线的股票。

	n1, n2, n3 = 5,10,20

	trade_days_list=get_previous_days(end_day)  #得到包括今天在内的过去5个交易日的日期。颠倒了时间顺序，下面的循环就从最近的日期开始检查。

	trade_days_list.reverse()   #必须这么写，不能写成 tradedays=trade_days_list.reverse()  这样就会错，不知道为什么。人工智能告诉我：reverse() 方法会就地修改列表，但它不会返回一个新的列表。因此，您不需要将 reverse() 方法的返回值赋给另一个变量。相反，您可以直接在原始列表上调用 reverse() 方法。

	#************* 这个地方原本是用了调酒，就是选择出来股票大于5天，10天，20天，属于多头排列的股票，把这个条件取消了，这样能找出不在多头排列形态,但是最近开始在涨的形态好的股票 **********************

	best_stock=[]

	for stockcode in targetcode_list:

		for tradeday in trade_days_list:

			targetjq=Jqdata(stockcode)

			if targetjq.judge_stock_potential(tradeday):

				best_stock.append(stockcode)

				#print(stockcode,tradeday)

				break

	return best_stock


def get_best_stock(total_money,end_day,select_ma):#选择出目前股价大于指定均线的股票，然后选择股票大于5天，10天，20天，多头排列的股票。然后从中选择出来这些股票中，在过去任意5个交易日当中任意往前加两天包括当天就是3天，如果那三天范围内，连涨两天，且高点低点连涨两天都往上提升的股票，选择出来。

	codelist=get_all_sec(end_day)

	yesterday_date=get_yesterday_date(end_day)

	selected_code_list=[x for x in codelist if not Jqdata(x).judge_recent_public_stock(end_day)] #去掉上市少于1年的股票。

	target_code_list=[x for x in selected_code_list if get_sec_marketsize(x,yesterday_date) > 100 ]  #总市值低于100亿的不要看了。

	#target_code_list=filter_stocklist_with_tradingmoney(codelist,end_day) #先过滤掉不符合成交额需要的股票。

	best_stock=get_best_target(target_code_list,end_day,select_ma)

	regular_stock_all,mistake_removed_rows=regular_check_stock_all_indicator(best_stock,total_money,present_day,select_ma) 

	return best_stock,regular_stock_all,mistake_removed_rows

def select_suoliang_stock(regular_stock_all):

	selected_df=regular_stock_all[regular_stock_all['vlchg']<1].copy()  #选择出缩量的股票。 

	return selected_df


def select_cross_stock(regular_stock_all):

	selected_df=regular_stock_all[regular_stock_all['cross']=='yes'].copy()  #选择出十字星的股票

	return selected_df


def select_duotou_stock(regular_stock_all):

	selected_df=regular_stock_all[regular_stock_all['duot']=='yes'].copy()  #选择出多头排列的股票

	return selected_df

def target_stock_all_indicator(source_df,total_money,present_day,select_ma):

	target_stock_list=list(source_df['code'])

	target_stock_df=regular_check_stock_all_indicator(target_stock_list,total_money,present_day,select_ma) 

	return target_stock_df


def check_best_stock_in_target_index(target_index_code,best_stock): #获得目标股票列表中哪些属于目标指数的成分股。重点看待！

	target_index_stock=Jqdata(target_index_code).get_index_securities()

	index_stock_buying=[x for x in best_stock if x in target_index_stock]

	index_stock_buying=[transform_to_normal_code(x) for x in index_stock_buying]

	return index_stock_buying


def get_best_index(end_day,select_ma): #同上。

	indexlist=filter_allindex(end_day)

	best_index=get_best_target(indexlist,end_day,select_ma)

	return best_index


def get_best_etf_from_jq(end_day,select_ma):

	etflist=get_all_etf(end_day)

	best_etf=get_best_target(etflist,end_day,select_ma)

	best_etf=[transform_to_normal_code(x) for x in best_etf]

	return best_etf

def daily_all_jqetf_monitor(total_money,datenow):  #这是线上的实时的监测etf的数据。监测的是所有的etf的市场表现。会把成交量低的ETF干掉。

	etflist=get_all_etf(datenow)

	etflist=[transform_to_normal_code(x) for x in etflist]

	new_etfdf_list=[[single_etf,Jqdata(single_etf).security_name()]+get_single_etf_data(single_etf,total_money,datenow) for single_etf in etflist]

	etfdf_columns=['code', 'name', 'b_vl', 'close','Lup','Ldown','MHg', 'MCg', 'cl_1', 'cl_2', 'cl_5', 'cl_10', 'cl_20', 'cl_30', 'cl_60', 'YTD', 'vlchg', 'vlrk', 'h_p_t', 'warn', 'zf', 'zfr']

	etf_df=list2dataframe(new_etfdf_list,etfdf_columns)

	etf_df = etf_df.drop(etf_df[etf_df['b_vl'] == 1000000].index)

	#etf_df.sort_values(by=[rankingkeyword],ascending=False,inplace=True)

	selectedrows=filter_etf_with_tradingmoney(etf_df,'code',datenow)  #会把成交量低的ETF干掉。

	selected_df = selectedrows[~selectedrows['name'].str.contains('国开|债')].copy()

	return selected_df


def get_best_jqetf(all_jqetf_df,top_percent,target_number):#是找回在不同时间周期内排名比较好的ETF

	vipjqetf=find_top_sec_from_df(all_jqetf_df,top_percent).copy()

	vip1=dataframe_ranking(vipjqetf,'cl_10').head(n=target_number)

	vip2=dataframe_ranking(vipjqetf,'cl_20').head(n=target_number)

	samevip=get_same_rows_from_dataframe(vip1,vip2,'code')

	return samevip


def get_best_etf(end_day,select_ma):

	etflist=[ single_etf_list[0] for single_etf_list in etfdf_list ]

	bestetf=get_best_target(etflist,end_day,select_ma)

	return bestetf


def monitor_best_stock(best_stock,total_money,end_day,select_ma):

	tradingdata_df=Trading(best_stock).get_trading_data(total_money,end_day,select_ma)

	tradingdata_df.loc[tradingdata_df['zhangting'] == 'yes', '买点'] ='早'

	conditions_2=[
	    (tradingdata_df['close'] < tradingdata_df['standard_ma']),
	    (tradingdata_df['连跌']=='yes'),
	     (tradingdata_df['低连降']=='yes'),
	     (tradingdata_df['高点跌幅']>10)
	]

	tradingdata_df['sell'] = np.where(np.all(conditions_2, axis=0), 'yes', '')


	tradingdata_df.drop(['zhangting','standard_ma'], axis=1, inplace=True)  #删除不需要的列。

	indicator_df=Jq_codelist(best_stock).check_stock_indicator_stocklist(total_money,end_day)

	closechange_df=get_target_stocklist_closechange_data(best_stock,total_money,end_day)

	final_df=make_stock_df_in_one(tradingdata_df,indicator_df,closechange_df)


	return final_df


def select_stock_with_company_industry_keyword(data,company_keyword_list,industry_keyword_list):

	# Assuming your DataFrame is named 'data'

	# Define the keywords
	#keywords_a = ["生物", "医", "药"]
	#keywords_b = ["生物", "医", "药", "制剂"]

	# Filter rows based on condition 1
	condition_a = data['name'].str.contains('|'.join(company_keyword_list))

	# Filter rows based on condition 2
	condition_b = data['industry'].str.contains('|'.join(industry_keyword_list))

	# Combine conditions using OR operation
	filtered_df = data[condition_a | condition_b]

	return filtered_df



def best_of_best_stock(targetdf):  #在select_bull_stock后面使用这个，选择出来最好中的低连升，因为select_bull_stock选择出来的就是高连胜。

	alldata=targetdf.copy() #先做一个副本。

	bestofbest_stock=alldata[alldata['低连升']=='yes']

	return bestofbest_stock


#主要是将之前交易日的股票代码获取的数据，放在交易日当天用相关指标观察，选择适合的代码进行当日交易。
def monitor_mvp_stock(mvp_code_list,total_money,present_day,select_ma):#这是检查选择出来的股票当天的表现，好在收盘前进行交易买卖。

	allstock_df,mistake_removed_rows=regular_check_stock_all_indicator(mvp_code_list,total_money,present_day,select_ma)

	nosell_mvp=remove_sell_stock(allstock_df).copy()  #留下不用卖的。

	dataframe_ranking(nosell_mvp,'buy')

	#best_vip_df=remove_down_trend_stock(nosell_mvp.copy())  #这一步会改变nosell_mvp的内容的，所以要保留nosell_mvp,就必须用一个副本.copy()替换掉nosell_mvp

	best_vip_df=remove_down_trend_stock(nosell_mvp)  #从不用卖的当中去掉高连降的股票。剩下的就是最好的。

	dataframe_ranking(best_vip_df,'buy')

	removed_rows=get_different_rows_from_dataframe(nosell_mvp,best_vip_df,'code')  #观察到哪些是高连降的股票，以免错杀，有时间可以看下。因为高连降的股票并不表示就一定不好！

	return nosell_mvp,best_vip_df,removed_rows,mistake_removed_rows  #要把可能被误删的行列出来，否则失去好机会可能！


def trading_allsec(total_money,end_day,select_ma):

	allsec_codelist=get_all_sec(end_day)

	allsec_code=[transform_to_normal_code(x) for x in allsec_codelist]

	buy_sec_df,buy_sec_code=Trading(allsec_code).trading_target_stock_list(total_money,end_day,select_ma)


	return buy_sec_df,buy_sec_code



def select_bull_stock(total_money,end_day,select_ma):#直接获取前一个函数的结果，进行计算统计。寻找符合交易策略的股票中的多头排列图形的股票。

	buy_sec_df,buy_sec_code=trading_allsec(total_money,end_day,select_ma)

	n1, n2, n3, = 5,10,20

	#股价在5天，10天，20天线上多头排列。

	#这个条件表达式必须这么写，不能连写，否则出错！

	condition = (buy_sec_df['close'] > buy_sec_df["MA{}".format(n1)]) & (buy_sec_df["MA{}".format(n1)] > buy_sec_df["MA{}".format(n2)]) & (buy_sec_df["MA{}".format(n2)] > buy_sec_df["MA{}".format(n3)])

	buy_sec_df_filter=buy_sec_df[condition]

	potential_bull_stock=list(buy_sec_df_filter['code'])

	#buy_sec_market_df_filter=buy_sec_market_df.loc[buy_sec_market_df['code'].isin(potential_bull_stock)]

	#buy_sec_indicator_df_filter=buy_sec_indicator_df.loc[buy_sec_indicator_df['code'].isin(potential_bull_stock)]

	return buy_sec_df_filter,potential_bull_stock


def check_stock_all_indicator(mystockcode,buy_sec_df,buy_sec_market_df,buy_sec_indicator_df):  #结果要从上一个函数里面获取到。但这个数据可能是滞后的。注意设置日期。

	target_stock_code=transform_to_normal_code(mystockcode)

	row1=buy_sec_df[buy_sec_df['code']==target_stock_code]

	row2=buy_sec_market_df[buy_sec_market_df['code']==target_stock_code]

	row3=buy_sec_indicator_df[buy_sec_indicator_df['code']==target_stock_code]

	display(freeze_columns(row1))

	display(freeze_columns(row2))

	display(freeze_columns(row3))  #以比较好的形式打印出dataframe.


def regular_check_stock_all_indicator(my_code,total_money,end_day,select_ma):#只适合股票。

	if type(my_code)==str: #只要是字符串，就变成list.

		stock_codelist=[Jqdata(my_code).jqcode] #必须变成列表形式。

	else:

		stock_codelist=my_code

	regular_buy_sec_df=Trading(stock_codelist).trading_strategy_for_stock(total_money,end_day,select_ma)  #就是列出来需要观察的股票的各种指标，不管好坏！

	#这一步获取股票的交易行情数据，会把一部分如果没有长时间的数据的股票筛选掉，所以要注意，哪些被筛选掉了。并不是因为不符合交易策略，而只是因为上市时间不够长，没有250天那么久之类的原因，导致被下面函数删除掉了。

	regular_stock_closechange_df=get_target_stocklist_closechange_data(stock_codelist,total_money,end_day)  #观察的代码由上一行代码决定。以下同。

	regular_stock_indicator_df=Jq_codelist(stock_codelist).check_stock_indicator_stocklist(total_money,end_day)

	regular_stock_all=make_stock_df_in_one(regular_buy_sec_df,regular_stock_indicator_df,regular_stock_closechange_df)

	mistake_removed_rows=get_different_rows_from_dataframe(regular_buy_sec_df,regular_stock_closechange_df,'code')  #被误删除的行，要看看。

	if not mistake_removed_rows.empty:

		print('\n*************   需要看mistake_removed 里面的数据，有数据！ ***********\n')

	else:

		print('\n*******************   没有误删除的行，不用看mistake_removed 那个数据行了。********************\n')

	#display(freeze_columns(regular_buy_sec_df))

	#display(freeze_columns(regular_stock_closechange_df))

	#display(freeze_columns(regular_stock_indicator_df))

	return regular_stock_all,mistake_removed_rows


def regular_check_holdingstock(holdingstockdf,total_money,end_day,select_ma):


	holdingstockdf.drop(holdingstockdf[pd.isna(holdingstockdf['code'])].index,inplace=True)

	holdingstock_codelist=get_codelist_from_df(holdingstockdf)

	holdingstock_codelist=[code for code in holdingstock_codelist if not (code.startswith('1') or code.startswith('5'))]

	all_holdingsec_df,mistake_removed_holdingsec=regular_check_stock_all_indicator(holdingstock_codelist,total_money,end_day,select_ma)

	all_no_sell_stock=remove_sell_stock(all_holdingsec_df).copy()  #选择出来不用卖的。

	best_holdingsec=remove_down_trend_stock(all_no_sell_stock)  #将其中高连降的也去掉，就剩下比较好的了。

	bad_holdingsec=get_different_rows_from_dataframe(all_holdingsec_df,best_holdingsec,'code')  #这样就得到哪些是不好的，被排除出去的股票。可能有的就是应该卖的，有的应该是属于高连降的情况。

	#dataframe_ranking(holdingsec_df,'sell')

	#dataframe_ranking(holdingsec_closechange_df,'cl_5')

	#dataframe_ranking(holdingsec_indicator_df,'涨幅')

	return all_holdingsec_df,best_holdingsec,bad_holdingsec,mistake_removed_holdingsec  #可能有被误删除但符合交易策略的行，要注意！



def dataframe_sorting(df,column): #排序的同时，不替换原有的datafrmae,inplace=False

	final_df=df.sort_values(by=[column],ascending=True,inplace=False).copy()   #排序是升序

	#df.reset_index(inplace=True)   #每次排序后，将索引重新生成。

	return final_df


def display_df_in_my_way(source_df,ranking_word):  #快速得让所想观察的df根据自己想要的关键字进行排序，并易于观察。

	final_df=dataframe_sorting(source_df,ranking_word)

	display(freeze_columns(final_df))

def display_df_in_opposite_way(source_df,ranking_word): #根据降序让观察的df进行关键字排序。

	final_df=dataframe_ranking(source_df,ranking_word)

	display(freeze_columns(final_df))


def trading_target_etf_list(target_list,total_money,end_day,select_ma):

	targetlist=[transform_to_normal_code(x) for x in target_list]

	etfcodelist=filter_etflist_with_tradingmoney(targetlist,'code',end_day)  #过滤掉成交额低的。

	final_df=Trading(etfcodelist).trading_strategy_for_stock(total_money,end_day,select_ma)

	#target_etf_df=trading_strategy(targetlist,total_money,end_day,select_ma)

	#final_df=filter_etf_with_tradingmoney(target_etf_df,'code',end_day)

	buyetf_df=final_df[final_df['buy']=='yes']

	if not buyetf_df.empty:

		final_buyetf_df=dataframe_ranking(buyetf_df.copy(),'低连升')

		buyetf_code=get_codelist_from_df(buyetf_df)

		buyetf_market_df=get_target_etflist_closechange_data(buyetf_code,total_money,end_day)

		final_buyetf_market_df=dataframe_ranking(buyetf_market_df.copy(),'cl_5')

		buyetf_indicator_df=Jq_codelist(buyetf_code).check_stock_indicator_stocklist(total_money,end_day)

		final_buyetf_indicator_df=dataframe_ranking(buyetf_indicator_df.copy(),'最高')

	else:

		final_buyetf_df=pd.DataFrame()

		final_buyetf_market_df=pd.DataFrame()

		final_buyetf_indicator_df=pd.DataFrame()


	return final_buyetf_df,final_buyetf_market_df,final_buyetf_indicator_df


def trading_alletf(total_money,end_day,select_ma):

	alletf_df=list2dataframe(etfdf_list,etfdf_columns)

	alletf_code=get_codelist_from_df(alletf_df)

	buyalletf_df,buyalletf_market_df,buyalletf_indicator_df=trading_target_etf_list(alletf_code,total_money,end_day,select_ma)

	return buyalletf_df,buyalletf_market_df,buyalletf_indicator_df


def trading_for_index_stock(index_code,total_money,end_day,select_ma):  #能够帮助我快速了解某个指数下面的股票的买卖建议。

	indexstock_codelist=Jqdata(index_code).get_index_securities()

	index_stock_codelist=[transform_to_normal_code(x) for x in indexstock_codelist]

	buyindex_stock_df,buyindex_stock_code=Trading(index_stock_codelist).trading_target_stock_list(total_money,end_day,select_ma)

	return buyindex_stock_df,buyindex_stock_code


class Trading_index():

	'''create class for trading strategy for index'''

	def __init__(self,index_code_list):

		self.index_code_list=index_code_list

	def get_index_list_data(self,end_day,select_ma): #这个函数是指数和etf通用的。这个很重要，实际应用中会看到。

		targetcode_list=self.index_code_list

		targetcode_list_info = []

		n1, n2, n3, n4, n5 = 5,10,20,100,250

		#targetcode_list=[transform_to_normal_code(x) for x in targetcodelist] #因为指数的代码如果没有前缀和后缀的话，容易跟股票代码混乱，所以，不要用transform_to_normal_code变化。


		for targetcode in targetcode_list:
		    targetjq = Jqdata(targetcode)
		    targetcode_name = targetjq.security_name()
		    presentclose = targetjq.get_present_close(end_day)
		    max_high, max_close = targetjq.get_max_high_close(end_day)
		    highest_gap = round((max_high - presentclose) / max_high, 4) * 100  #20天内从最高点回撤幅度大于10%就是可以考虑卖出的条件之一。

		    ma1,ma2,ma3,ma4,ma5 = targetjq.get_ma(n1, end_day),targetjq.get_ma(n2, end_day),targetjq.get_ma(n3, end_day),targetjq.get_ma(n4, end_day),targetjq.get_ma(n5, end_day)

		    buyin = ''
		    sellout = ''


		    close_up, close_down, high_up, high_down, low_up, low_down = targetjq.judge_stock_data(end_day)

		    standard_ma_day = targetjq.get_ma(select_ma, end_day)

		    targetcode=transform_to_normal_code(targetcode)

		    targetcode_info = [targetcode, targetcode_name, presentclose, buyin,
		                       sellout, standard_ma_day,ma1, ma2, ma3, ma4, ma5, close_up, close_down, high_up, high_down,
		                       low_up, low_down, highest_gap]

		    if not pd.isna(presentclose):  # Instead of dropping rows with missing values using df.drop, you can filter the rows during the loop itself to avoid creating unnecessary rows in the DataFrame.

		        targetcode_list_info.append(targetcode_info)

		targetcode_columns = ['code', 'name', 'close', 'buy', 'sell','standard_ma', "MA{}".format(n1), "MA{}".format(n2), "MA{}".format(n3), "MA{}".format(n4),"MA{}".format(n5), '连涨', '连跌', '高连升', '高连降', '低连升', '低连降', '高点跌幅']

		df = pd.DataFrame(targetcode_list_info, columns=targetcode_columns)

		df['name']=df['name'].apply(remove_string_from_index)

		df.reset_index(drop=True, inplace=True)

		return df


	def index_strategy_buy(self,end_day,select_ma): #选择大于标准日线及符合其他指标的指数。

		target_code_list=self.index_code_list

		targetcode_list=[x for x in target_code_list if Jqdata(x).judge_above_ma(select_ma,end_day)]  #选择出符合大于选择的移动平均线的股票。

		df=Trading_index(targetcode_list).get_index_list_data(end_day,select_ma)

		conditions = [
		    (df['连涨'] == 'yes'),
		    (df['高连升'] == 'yes')
		]

		df['buy'] = np.where(np.all(conditions, axis=0), 'yes', '')

		df.drop(['standard_ma'], axis=1, inplace=True)  #删除不需要的列。

		message = '本次选择的买卖标准是 **** {}  天线！**** '.format(select_ma)

		print(message)

		print('buy 代表指数点位大于 {} 天线，指数连续三天收盘价上涨，最高点每天都上涨。\n'.format(select_ma))

		print('sell 代表指数点位低于 {} 天线，指数连续三天收盘价下跌，最低点每天都下跌，并且过去20天内指数从最高点到现价已经下跌10%了！ \n'.format(select_ma))

		return df

	def index_buy_sell(self,end_day,select_ma):

		'''list out all the buy sell information about the index I want'''

		df=self.get_index_list_data(end_day,select_ma)

		conditions = [
		    (df['close'] > df['standard_ma']),
		    (df['连涨'] == 'yes'),
		    (df['高连升'] == 'yes')
		]

		df['buy'] = np.where(np.all(conditions, axis=0), 'yes', '')


		conditions_2=[
		    (df['close'] < df['standard_ma']),
		    (df['连跌']=='yes'),
		     (df['低连降']=='yes'),
		     (df['高点跌幅']>10)
		]

		df['sell'] = np.where(np.all(conditions_2, axis=0), 'yes', '')

		df.drop(['standard_ma'], axis=1, inplace=True)  #删除不需要的列。

		message = '本次选择的买卖标准是 **** {}  天线！**** '.format(select_ma)

		print(message)

		print('buy 代表指数点位大于 {} 天线，指数连续三天收盘价上涨，最高点每天都上涨。\n'.format(select_ma))

		print('sell 代表指数点位低于 {} 天线，指数连续三天收盘价下跌，最低点每天都下跌，并且过去20天内指数从最高点到现价已经下跌10%了！ \n'.format(select_ma))

		return df


	def indexlist_trading(self,end_day,select_ma):  #因为指数的代码如果没有前缀和后缀的话，容易跟股票代码混乱，所以，不要用transform_to_normal_code变化。

		final_df=self.index_buy_sell(end_day,select_ma)

		buyindex_df=final_df[final_df['buy']=='yes']

		if not buyindex_df.empty:

			buyindex_code=list(buyindex_df['code'])

			final_buyindex_df=dataframe_ranking(buyindex_df.copy(),'低连升')

			indexdf_list=[get_single_index_data(index_code,end_day) for index_code in buyindex_code]

			indexdf_columns=['code', 'name','close', 'MHg', 'MCg', 'cl_1', 'cl_2', 'cl_5', 'cl_10', 'cl_20', 'cl_30', 'cl_60', 'YTD', 'vlchg', 'vlrk', 'h_p_t', 'warn', 'zf', 'zfr']

			index_market_df=list2dataframe(indexdf_list,indexdf_columns)

			index_market_df['name']=index_market_df['name'].apply(remove_string_from_index)

			final_index_market_df=dataframe_ranking(index_market_df.copy(),'cl_5')

			index_indicator_df=check_index_indicator_indexlist(buyindex_code,end_day)

			final_index_indicator_df=dataframe_ranking(index_indicator_df.copy(),'最高涨幅')

		else:

			final_buyindex_df=pd.DataFrame()

			final_index_market_df=pd.DataFrame()

			final_index_indicator_df=pd.DataFrame()


		sellindex_df=final_df[final_df['sell']=='yes']


		if not sellindex_df.empty:

			final_sellindex_df=dataframe_ranking(sellindex_df.copy(),'低连降')

		else:

			final_sellindex_df=pd.DataFrame()
		

		return final_buyindex_df,final_index_market_df,final_index_indicator_df,final_sellindex_df



def trading_allindex(end_day,select_ma):

	index_list=filter_allindex(end_day)  #因为指数的代码如果没有前缀和后缀的话，容易跟股票代码混乱，所以，不要用transform_to_normal_code变化。

	buyindex_df,index_market_df,index_indicator_df,sellindex_df=Trading_index(index_list).indexlist_trading(end_day,select_ma)

	return buyindex_df,index_market_df,index_indicator_df,sellindex_df


def index_stock_trading_indicator(index_code,end_day): #to see how a index trading volume and value data. 这是在交易指数的ETF的时候观察成交量的变化的重要参考。

	yesterday_date=get_yesterday_date(end_day)

	indexstock_codelist=get_index_stocks(Jqdata(index_code).jqcode,end_day)  #这样写，才能查出的是当时的成分股的数据，否则成分股数据不准确。

	index_name=Jqdata(index_code).security_name()

	#index_stock_codelist=[transform_to_normal_code(x) for x in indexstock_codelist]

	index_df=get_target_index_weights(index_code,end_day)

	total_stock_marketsize=index_df['marketsize'].sum()  #all index stock market size

	total_stock_liudong_size=index_df['liutong_size'].sum()

	present_trading_money=get_trading_money(index_code,end_day)

	present_volume=Jqdata(index_code).get_present_volume(end_day)/10000 #成交量多少股，除以10000，变成多少万股。和股本的计算单位保持一致。

	indexstockdf=get_valuation(indexstock_codelist, end_date=yesterday_date, count=1, fields=['capitalization', 'circulating_cap'])  #所有成分股总股本，流通股本 万股

	allstock_total_guben=indexstockdf['capitalization'].sum()

	allstock_total_liutong_guben=indexstockdf['circulating_cap'].sum()

	volume_total_pct=(present_volume/allstock_total_guben)*100

	warningword_1=''

	warningword_2=''

	warningword_3=''

	warningword_4=''

	if volume_total_pct > 6:

		warningword_1='危'

	volume_liutong_pct=(present_volume/allstock_total_liutong_guben)*100

	if volume_liutong_pct > 6:

		warningword_2='危'

	money_total_size_pct=(present_trading_money/total_stock_marketsize)*100

	if money_total_size_pct >6:

		warningword_3='危'

	money_total_liutong_size_pct=(present_trading_money/total_stock_liudong_size)*100

	if money_total_liutong_size_pct > 6:

		warningword_4='危'

	dflist=[[index_code,index_name,present_volume,allstock_total_guben,volume_total_pct,warningword_1,allstock_total_liutong_guben,volume_liutong_pct,warningword_2,present_trading_money,total_stock_marketsize,money_total_size_pct,warningword_3,total_stock_liudong_size,money_total_liutong_size_pct,warningword_4]]

	dfcolumns=['code','name','成交量','成分股票总股本','占总股本比例','信号','成分股票流通股','占流通股比例','信号','成交额','成分股总市值','占总市值比例','信号','成分股总流通市值','占流通市值比例','信号']

	df=list2dataframe(dflist,dfcolumns)

	return df


def get_security_lowest_price_date(stockcode,startdate,present_day):#找到一个股票从指定开始日期到指定结束时间之间的最低价的日期。

	stock_jq=Jqdata(stockcode)

	public_date=stock_jq.get_security_public_date()

	if public_date > startdate: #股票上市时间晚于观察的开始日期，就按上市日期来计算。

		startdate=public_date

	low_price_df=get_price(stock_jq.jqcode,start_date=startdate,end_date=present_day,frequency='daily',fields=['low'])  

	low_price_df.drop(low_price_df[pd.isna(low_price_df['low'])].index, inplace=True) #把没有最低价的日期的行去掉，以免影响比较的下一步代码。

	if not low_price_df.empty:

		lowest_price_date = low_price_df[low_price_df['low'] == low_price_df['low'].min()].index[0].strftime('%Y-%m-%d')

	else:

		lowest_price_date='None'

	return lowest_price_date


def get_stocklist_lowest_price_date(stock_list,start_date,present_day):

	stocklist=[transform_to_normal_code(x) for x in stock_list]

	stock_info=[[stockcode,Jqdata(stockcode).security_name(),get_security_lowest_price_date(stockcode,start_date,present_day)] for stockcode in stocklist]

	stock_columns=['code','name','lowprice_date']

	df=list2dataframe(stock_info,stock_columns)

	return df

def get_stock_gap_to_ma(stockcode,present_day): #实时获取一个股票距离5，10，20天均线的差距。以此判断其调整的幅度情况。

	targetjq=Jqdata(stockcode)

	present_point,present_open,present_high,present_low=targetjq.get_stock_point(present_day)

	ma5,ma10,ma20=targetjq.get_ma(5,present_day),targetjq.get_ma(10,present_day),targetjq.get_ma(20,present_day)

	gap_to_ma5=round(round((present_point - ma5)/ma5,4)*100,2)

	gap_to_ma10=round(round((present_point - ma10)/ma10,4)*100,2) 

	gap_to_ma20=round(round((present_point - ma20)/ma20,4)*100,2)

	return gap_to_ma5,gap_to_ma10,gap_to_ma20


def get_stocklist_gap_to_ma(stockslist,present_day):

	stock_info=[]

	for stockcode in stockslist:

		myjq=Jqdata(stockcode)

		presentclose = myjq.get_present_close(present_day)

		max_high, max_close = myjq.get_max_high_close(present_day)

		highest_gap = round((max_high - presentclose) / max_high, 4) * 100 

		gap_to_ma5,gap_to_ma10,gap_to_ma20=get_stock_gap_to_ma(stockcode,present_day)
		
		single_stock_info=[transform_to_normal_code(stockcode),myjq.security_name(),myjq.security_industry(present_day),gap_to_ma5,gap_to_ma10,gap_to_ma20,highest_gap]

		stock_info.append(single_stock_info)

	finaldf_columns=['code','name','行业','距5日线','距10日线','距20日线','高点跌幅']

	finaldf=list2dataframe(stock_info,finaldf_columns)


	return finaldf



def watch_whole_market(startingdate,endingdate):  #startingdate指的是数据观察的开始时间，比如2023年的12月16日，我肯定是从年头开始观察的，就是2023-01-01，endingdate就是截止到查数据的时间。

	# 获取所有股票的数据
	stocks = get_all_securities(types=['stock'], date=endingdate)

	print('现有股票数量为 {} '.format(len(stocks)))

	stocks.reset_index(inplace=True)
	stocks.rename(columns={'index': 'code'}, inplace=True)
	stocks.drop(['end_date', 'type', 'name'], axis=1, inplace=True)

	# 将日期字符串转换为 Pandas 时间戳对象
	stocks['start_date'] = pd.to_datetime(stocks['start_date'])  #这里的start_date，是每只股票的上市时间。

	# 将时间戳对象格式化为 '2023-01-01' 格式
	stocks['start_date'] = stocks['start_date'].dt.strftime('%Y-%m-%d')

	# 如果上市日期早于2023-01-01，将start_date设为'2023-01-01'，否则保持不变
	stocks.loc[stocks['start_date'] < startingdate, 'start_date'] = startingdate

	# 遍历股票数据
	for index, row in stocks.iterrows():
	    # 获取股票的历史数据
	    stock_data = get_price(row['code'], start_date=row['start_date'], end_date=endingdate, frequency='daily', fields=['low'])
	    stock_data.drop(stock_data[pd.isna(stock_data['low'])].index, inplace=True)  # 很关键，如果一条数据都没有，那么这就是筛选的关键一步。
	    
	    # 剔除停牌的股票
	    if not stock_data.empty:
	        lowest_price_date = stock_data[stock_data['low'] == stock_data['low'].min()].index[0].strftime('%Y-%m-%d')
	        # 将最低价日期添加到 DataFrame 中的新列 'low_price_date'
	        stocks.at[index, 'low_price_date'] = lowest_price_date
	    else:
	        stocks.at[index, 'low_price_date'] = 'None'

	stocks = stocks[stocks['low_price_date'] != 'None']

	print('去掉没有最低价的股票后数量为 {}   可能有的退市或者停牌'.format(len(stocks)))


	# 假设您的 DataFrame 名称为 stocks，并且包含了 'low_price_date' 列
	# 创建 DataFrame 的副本以避免 SettingWithCopyWarning 警告
	stocks_copy = stocks.copy()

	# 将 'low_price_date' 列转换为日期时间格式
	stocks_copy['low_price_date'] = pd.to_datetime(stocks_copy['low_price_date'])

	# 按月统计每月最低价的股票数量
	monthly_stock_counts = stocks_copy['low_price_date'].dt.to_period('M').value_counts().sort_index()

	# 计算每月总股票数
	total_stock_counts = stocks_copy['low_price_date'].dt.to_period('M').count()

	# 计算每月最低价股票数量占总股票数的百分比
	monthly_stock_percentages = (monthly_stock_counts / total_stock_counts) * 100

	# 将每月最低价股票数量和占比合并为一个 DataFrame
	monthly_data = pd.DataFrame({'股票数量': monthly_stock_counts, '占比 (%)': monthly_stock_percentages})


	monthly_data['total']=monthly_data['股票数量'].sum()

	# 打印每月最低价股票数量和占比
	print('每月最低价股票数量和占比:')
	print(monthly_data)


	# 使用柱状图可视化每月的最低价股票数量
	plt.figure(figsize=(10, 6))
	monthly_stock_counts.plot(kind='bar', color='skyblue')
	plt.title('每月最低价股票数量')
	plt.xlabel('月份')
	plt.ylabel('股票数量')
	plt.show()


	# 假设您的 DataFrame 名称为 stocks，并且包含了 'low_price_date' 列
	# 将 'low_price_date' 列转换为日期时间格式
	stocks_copy_df=stocks.copy()
	stocks_copy_df['low_price_date'] = pd.to_datetime(stocks_copy_df['low_price_date'])

	# 计算每天的最低价股票数量
	daily_stock_counts = stocks_copy_df['low_price_date'].value_counts()

	# 将日期作为索引并按日期排序
	daily_stock_counts = daily_stock_counts.sort_index()

	print(daily_stock_counts)

	# 绘制折线图
	plt.figure(figsize=(40,24))
	plt.plot(daily_stock_counts.index, daily_stock_counts.values, marker='o', linestyle='-')
	plt.title('每天的最低价股票数量')
	plt.xlabel('日期')
	plt.ylabel('股票数量')
	plt.xticks(rotation=45)  # 旋转x轴标签，使其更易阅读
	plt.show()

	# 找到最低价数量最多的日期
	top_dates = daily_stock_counts.nlargest(10)
	print("最低价数量最多的日期及对应的股票数量：")
	print(top_dates)

	return stocks

def watch_whole_index_market(startingdate,endingdate):  #startingdate指的是数据观察的开始时间，比如2023年的12月16日，我肯定是从年头开始观察的，就是2023-01-01，endingdate就是截止到查数据的时间。

	# 获取所有指数的数据
	df = get_all_securities(types=['index'], date=endingdate)

	stocks = df[~df['display_name'].str.contains('债')].copy()  #把带债的指数去掉。

	print('现有指数数量为 {} '.format(len(stocks)))

	stocks.reset_index(inplace=True)
	stocks.rename(columns={'index': 'code'}, inplace=True)
	stocks.drop(['end_date', 'type', 'name'], axis=1, inplace=True)

	# 将日期字符串转换为 Pandas 时间戳对象
	stocks['start_date'] = pd.to_datetime(stocks['start_date'])  #这里的start_date，是每只指数的上市时间。

	# 将时间戳对象格式化为 '2023-01-01' 格式
	stocks['start_date'] = stocks['start_date'].dt.strftime('%Y-%m-%d')

	# 如果上市日期早于2023-01-01，将start_date设为'2023-01-01'，否则保持不变
	stocks.loc[stocks['start_date'] < startingdate, 'start_date'] = startingdate

	# 遍历股票数据
	for index, row in stocks.iterrows():
	    # 获取股票的历史数据
	    stock_data = get_price(row['code'], start_date=row['start_date'], end_date=endingdate, frequency='daily', fields=['low'])
	    stock_data.drop(stock_data[pd.isna(stock_data['low'])].index, inplace=True)  # 很关键，如果一条数据都没有，那么这就是筛选的关键一步。
	    
	    # 剔除停牌的股票
	    if not stock_data.empty:
	        lowest_price_date = stock_data[stock_data['low'] == stock_data['low'].min()].index[0].strftime('%Y-%m-%d')
	        # 将最低价日期添加到 DataFrame 中的新列 'low_price_date'
	        stocks.at[index, 'low_price_date'] = lowest_price_date
	    else:
	        stocks.at[index, 'low_price_date'] = 'None'

	stocks = stocks[stocks['low_price_date'] != 'None']

	print('去掉没有最低价的指数后数量为 {}   可能有的退市或者停牌'.format(len(stocks)))


	# 假设您的 DataFrame 名称为 stocks，并且包含了 'low_price_date' 列
	# 创建 DataFrame 的副本以避免 SettingWithCopyWarning 警告
	stocks_copy = stocks.copy()

	# 将 'low_price_date' 列转换为日期时间格式
	stocks_copy['low_price_date'] = pd.to_datetime(stocks_copy['low_price_date'])

	# 按月统计每月最低价的股票数量
	monthly_stock_counts = stocks_copy['low_price_date'].dt.to_period('M').value_counts().sort_index()

	# 计算每月总股票数
	total_stock_counts = stocks_copy['low_price_date'].dt.to_period('M').count()

	# 计算每月最低价股票数量占总股票数的百分比
	monthly_stock_percentages = (monthly_stock_counts / total_stock_counts) * 100

	# 将每月最低价股票数量和占比合并为一个 DataFrame
	monthly_data = pd.DataFrame({'指数数量': monthly_stock_counts, '占比 (%)': monthly_stock_percentages})


	monthly_data['total']=monthly_data['指数数量'].sum()

	# 打印每月最低价股票数量和占比
	print('每月最低价指数数量和占比:')
	print(monthly_data)


	# 使用柱状图可视化每月的最低价股票数量
	plt.figure(figsize=(10, 6))
	monthly_stock_counts.plot(kind='bar', color='skyblue')
	plt.title('每月最低价指数数量')
	plt.xlabel('月份')
	plt.ylabel('指数数量')
	plt.show()


	# 假设您的 DataFrame 名称为 stocks，并且包含了 'low_price_date' 列
	# 将 'low_price_date' 列转换为日期时间格式
	stocks_copy_df=stocks.copy()
	stocks_copy_df['low_price_date'] = pd.to_datetime(stocks_copy_df['low_price_date'])

	# 计算每天的最低价股票数量
	daily_stock_counts = stocks_copy_df['low_price_date'].value_counts()

	# 将日期作为索引并按日期排序
	daily_stock_counts = daily_stock_counts.sort_index()

	print(daily_stock_counts)

	# 绘制折线图
	plt.figure(figsize=(40,24))
	plt.plot(daily_stock_counts.index, daily_stock_counts.values, marker='o', linestyle='-')
	plt.title('每天的最低价指数数量')
	plt.xlabel('日期')
	plt.ylabel('指数数量')
	plt.xticks(rotation=45)  # 旋转x轴标签，使其更易阅读
	plt.show()

	# 找到最低价数量最多的日期
	top_dates = daily_stock_counts.nlargest(10)
	print("最低价数量最多的日期及对应的指数数量：")
	print(top_dates)

	return stocks